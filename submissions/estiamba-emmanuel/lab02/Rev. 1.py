"""
Rev. 1 - 6m x 6m x 6m Cube Structural Model Generator
======================================================
Generates a 3D frame model of a 6 m cube with:
    - Node coordinates (global X, Y, Z)
    - Member incidences (Beam / Column classification)
    - Pinned support conditions at the base nodes
    - Member local axis systems (Local-1 / Local-2 / Local-3) with beta angle
    - Nodal degrees of freedom (DOF) and global DOF numbering
    - Member end releases (moment release for pinned beam connections)
    - Excel workbook export (Nodes, Member Incidences, Supports, Node DOF,
      Member Local Axes, Member Releases)
    - 3D wireframe plot with global axes, local axis triads, and support
      symbols

Axis convention (matches common structural analysis solvers such as
RISA-3D / STAAD.Pro):
    Global X : horizontal (lateral)
    Global Z : horizontal (lateral)
    Global Y : vertical (global up)

Local axis convention (per member):
    Local-1 (x') : along the member axis, from node i to node j
    Local-2 (y') : "vertical" reference axis, rotated by the beta angle
    Local-3 (z') : completes the right-handed set (Local-1 x Local-2)

Revision history:
    Rev. 0 - Initial release: nodes, members, Excel export, 3D plot.
    Rev. 1 - Added pinned supports, beta angle, local/global axes, nodal
             DOF, member end releases, and updated the diagram and the
             Excel output accordingly.

Note on member end releases (see Section 2d): this revision assumes Beam
members are simple (pinned) connections that release the local-3 bending
moment (Mz) at both ends, while Column members remain fully fixed. Edit
the `member_releases` dictionary below if a different release pattern is
required for your model.
"""

import os
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D

L = 6.0  # cube side length (m)

# ---------------------------------------------------------------------
# 0. Output folders
# ---------------------------------------------------------------------
FIGURES_DIR = r"C:\Users\DC Gaming\python\figures"
LABWORKS_DIR = r"C:\Users\DC Gaming\python\Labworks"

os.makedirs(FIGURES_DIR, exist_ok=True)
os.makedirs(LABWORKS_DIR, exist_ok=True)

# ---------------------------------------------------------------------
# 1. Node coordinates (Node 1 at origin 0,0,0)
#    Bottom face (Y=0): nodes 1-4 ; Top face (Y=6): nodes 5-8
# ---------------------------------------------------------------------
nodes = {
    1: (0, 0, 0),
    2: (L, 0, 0),
    3: (L, 0, L),
    4: (0, 0, L),
    5: (0, L, 0),
    6: (L, L, 0),
    7: (L, L, L),
    8: (0, L, L),
}

BASE_NODES = (1, 2, 3, 4)  # bottom-face nodes -> pinned supports

# ---------------------------------------------------------------------
# 2. Member incidences (12 edges of the cube) -> (i, j, type)
#    Type is derived from geometry:
#      - Vertical members (same X,Z; different Y)   -> "Column"
#      - Horizontal members (same Y; different X/Z) -> "Beam"
# ---------------------------------------------------------------------
member_pairs = [
    # Bottom face
    (1, 2), (2, 3), (3, 4), (4, 1),
    # Top face
    (5, 6), (6, 7), (7, 8), (8, 5),
    # Vertical columns
    (1, 5), (2, 6), (3, 7), (4, 8),
]

def classify_member(i, j):
    xi, yi, zi = nodes[i]
    xj, yj, zj = nodes[j]
    if xi == xj and zi == zj and yi != yj:
        return "Column"
    return "Beam"

members = [(i, j, classify_member(i, j)) for i, j in member_pairs]

# Beta angle (degrees): rotation of the Local-2 / Local-3 axes about the
# member's Local-1 (axial) axis, measured from the default orientation
# established in Section 2c. Default is 0.0 deg for every member; edit
# the value for a given member index below if a different orientation
# is required.
BETA_DEG = {idx: 0.0 for idx in range(1, len(member_pairs) + 1)}

# ---------------------------------------------------------------------
# 2a. Support conditions (pinned at the base nodes)
#     Pinned support: translations restrained (Ux = Uy = Uz = 0),
#                      rotations free (Rx, Ry, Rz unrestrained)
# ---------------------------------------------------------------------
DOF_NAMES = ["Ux", "Uy", "Uz", "Rx", "Ry", "Rz"]

def get_support(node):
    """Return a dict of restrained (True) / free (False) status per DOF."""
    if node in BASE_NODES:
        return {"Ux": True, "Uy": True, "Uz": True,
                "Rx": False, "Ry": False, "Rz": False}  # pinned
    return {d: False for d in DOF_NAMES}  # no support -> all DOF free

supports = {n: get_support(n) for n in nodes}

# ---------------------------------------------------------------------
# 2b. Nodal degrees of freedom and global DOF numbering
#     Every node has 6 DOF (3 translations + 3 rotations). Restrained
#     DOF are excluded from the active DOF numbering and marked "R".
# ---------------------------------------------------------------------
node_dof_numbers = {}
_active_counter = 0
for n in sorted(nodes):
    dof_map = {}
    for d in DOF_NAMES:
        if supports[n][d]:
            dof_map[d] = "R"  # restrained
        else:
            _active_counter += 1
            dof_map[d] = _active_counter
    node_dof_numbers[n] = dof_map

TOTAL_DOF = len(nodes) * 6
ACTIVE_DOF = _active_counter
RESTRAINED_DOF = TOTAL_DOF - ACTIVE_DOF

# ---------------------------------------------------------------------
# 2c. Member local axes (Local-1 / Local-2 / Local-3)
#     Local-1 : member axial direction, from node i to node j
#     Local-2 : default reference is global Y (up); if the member itself
#               is parallel to global Y (a column), global X is used as
#               the reference instead. Beta then rotates Local-2/Local-3
#               about Local-1.
#     Local-3 : completes the right-handed set (Local-1 x Local-2)
# ---------------------------------------------------------------------
def local_axes(i, j, beta_deg=0.0):
    p_i = np.array(nodes[i], dtype=float)
    p_j = np.array(nodes[j], dtype=float)
    vec = p_j - p_i
    length = np.linalg.norm(vec)
    local_1 = vec / length

    global_y = np.array([0.0, 1.0, 0.0])
    if abs(np.dot(local_1, global_y)) > 0.999:
        ref = np.array([1.0, 0.0, 0.0])  # member is vertical -> use global X
    else:
        ref = global_y

    local_3 = np.cross(local_1, ref)
    local_3 /= np.linalg.norm(local_3)
    local_2 = np.cross(local_3, local_1)
    local_2 /= np.linalg.norm(local_2)

    # Rotate Local-2 / Local-3 about Local-1 by the beta angle
    # (Rodrigues' rotation formula)
    beta = np.radians(beta_deg)
    local_2_rot = (local_2 * np.cos(beta)
                   + np.cross(local_1, local_2) * np.sin(beta)
                   + local_1 * np.dot(local_1, local_2) * (1 - np.cos(beta)))
    local_2_rot /= np.linalg.norm(local_2_rot)
    local_3_rot = np.cross(local_1, local_2_rot)
    local_3_rot /= np.linalg.norm(local_3_rot)

    return local_1, local_2_rot, local_3_rot

member_local_axes = {
    idx: local_axes(i, j, BETA_DEG[idx])
    for idx, (i, j, _mtype) in enumerate(members, start=1)
}

# ---------------------------------------------------------------------
# 2d. Member end releases (force / moment releases in LOCAL coordinates)
#     False = fixed/rigid  (the force or moment is transferred)
#     True  = released      (pinned / free - no transfer)
#     Local DOF order: Fx (axial), Fy, Fz (shear), Mx (torsion),
#                       My, Mz (bending moments)
#
#     Assumption for this revision (see module docstring): Beam members
#     are modeled as pinned (simple) connections, releasing the local-3
#     bending moment (Mz) at both ends. Column members remain fully
#     fixed. Adjust `member_releases` below for a different pattern.
# ---------------------------------------------------------------------
LOCAL_DOF_NAMES = ["Fx", "Fy", "Fz", "Mx", "My", "Mz"]

def default_release(mtype):
    end = {d: False for d in LOCAL_DOF_NAMES}
    if mtype == "Beam":
        end["Mz"] = True  # pinned beam: local-z moment released
    return end

member_releases = {
    idx: {"i": default_release(mtype), "j": default_release(mtype)}
    for idx, (i, j, mtype) in enumerate(members, start=1)
}

# ---------------------------------------------------------------------
# 3. Build Excel workbook
# ---------------------------------------------------------------------
wb = openpyxl.Workbook()

header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
body_font = Font(name="Arial")
center = Alignment(horizontal="center")

def style_header(ws, headers):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

def style_body(ws, ncols):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ncols):
        for cell in row:
            cell.font = body_font
            cell.alignment = center

# --- Sheet 1: Nodes ---
ws_nodes = wb.active
ws_nodes.title = "Nodes"
style_header(ws_nodes, ["Node", "X (m)", "Y (m)", "Z (m)"])
for n, (x, y, z) in nodes.items():
    ws_nodes.append([n, x, y, z])
style_body(ws_nodes, 4)
for col_letter, width in zip("ABCD", [10, 10, 10, 10]):
    ws_nodes.column_dimensions[col_letter].width = width

# --- Sheet 2: Member Incidences ---
ws_mem = wb.create_sheet("Member Incidences")
style_header(ws_mem, ["Member", "i (Start Node)", "j (End Node)", "Type", "Beta (deg)"])
for idx, (i, j, mtype) in enumerate(members, start=1):
    ws_mem.append([idx, i, j, mtype, BETA_DEG[idx]])
style_body(ws_mem, 5)
for col_letter, width in zip("ABCDE", [10, 16, 16, 12, 12]):
    ws_mem.column_dimensions[col_letter].width = width

# --- Sheet 3: Supports ---
ws_sup = wb.create_sheet("Supports")
style_header(ws_sup, ["Node", "Support Type", "Ux", "Uy", "Uz", "Rx", "Ry", "Rz"])
for n in sorted(nodes):
    stype = "Pinned" if n in BASE_NODES else "Free (no support)"
    row = [n, stype] + ["Fixed" if supports[n][d] else "Free" for d in DOF_NAMES]
    ws_sup.append(row)
style_body(ws_sup, 8)
for col_letter, width in zip("ABCDEFGH", [8, 18, 8, 8, 8, 8, 8, 8]):
    ws_sup.column_dimensions[col_letter].width = width

# --- Sheet 4: Node DOF (global DOF numbering) ---
ws_dof = wb.create_sheet("Node DOF")
style_header(ws_dof, ["Node"] + DOF_NAMES)
for n in sorted(nodes):
    ws_dof.append([n] + [node_dof_numbers[n][d] for d in DOF_NAMES])
style_body(ws_dof, 7)
for col_letter, width in zip("ABCDEFG", [8, 8, 8, 8, 8, 8, 8]):
    ws_dof.column_dimensions[col_letter].width = width
ws_dof.append([])
ws_dof.append(["Total DOF:", TOTAL_DOF])
ws_dof.append(["Active (Free) DOF:", ACTIVE_DOF])
ws_dof.append(["Restrained DOF:", RESTRAINED_DOF])

# --- Sheet 5: Member Local Axes ---
ws_axes = wb.create_sheet("Member Local Axes")
style_header(ws_axes, ["Member", "i", "j", "Beta (deg)",
                       "Local-1 (X)", "Local-1 (Y)", "Local-1 (Z)",
                       "Local-2 (X)", "Local-2 (Y)", "Local-2 (Z)",
                       "Local-3 (X)", "Local-3 (Y)", "Local-3 (Z)"])
for idx, (i, j, mtype) in enumerate(members, start=1):
    l1, l2, l3 = member_local_axes[idx]
    ws_axes.append([idx, i, j, BETA_DEG[idx],
                     round(l1[0], 3), round(l1[1], 3), round(l1[2], 3),
                     round(l2[0], 3), round(l2[1], 3), round(l2[2], 3),
                     round(l3[0], 3), round(l3[1], 3), round(l3[2], 3)])
style_body(ws_axes, 13)
for col_letter in "ABCDEFGHIJKLM":
    ws_axes.column_dimensions[col_letter].width = 11

# --- Sheet 6: Member Releases ---
ws_rel = wb.create_sheet("Member Releases")
style_header(ws_rel, ["Member", "End", "Fx", "Fy", "Fz", "Mx", "My", "Mz"])
for idx, (i, j, mtype) in enumerate(members, start=1):
    for end_label, node_ref in (("i", i), ("j", j)):
        rel = member_releases[idx][end_label]
        row = [idx, f"{end_label} (Node {node_ref})"] + \
              ["Released" if rel[d] else "Fixed" for d in LOCAL_DOF_NAMES]
        ws_rel.append(row)
style_body(ws_rel, 8)
for col_letter, width in zip("ABCDEFGH", [10, 16, 10, 10, 10, 10, 10, 10]):
    ws_rel.column_dimensions[col_letter].width = width

excel_path = os.path.join(LABWORKS_DIR, "cube_model.xlsx")
wb.save(excel_path)
print(f"Saved: {excel_path}")

# ---------------------------------------------------------------------
# 4. 3D Plot (X & Z on the horizontal plane, Y vertical)
# ---------------------------------------------------------------------
fig = plt.figure(figsize=(9, 9))
ax = fig.add_subplot(111, projection="3d")

for idx, (i, j, mtype) in enumerate(members, start=1):
    x_vals = [nodes[i][0], nodes[j][0]]
    z_vals = [nodes[i][2], nodes[j][2]]
    y_vals = [nodes[i][1], nodes[j][1]]
    color = "darkorange" if mtype == "Column" else "steelblue"
    ax.plot(x_vals, z_vals, y_vals, color=color, linewidth=2)

    # --- Local axis triad at the member midpoint ---
    mid_x = (nodes[i][0] + nodes[j][0]) / 2
    mid_z = (nodes[i][2] + nodes[j][2]) / 2
    mid_y = (nodes[i][1] + nodes[j][1]) / 2
    l1, l2, l3 = member_local_axes[idx]
    scale = 0.6
    # Local-1 (axial) in red, Local-2 in green, Local-3 in blue
    ax.quiver(mid_x, mid_z, mid_y, l1[0], l1[2], l1[1],
              length=scale, color="red", linewidth=1)
    ax.quiver(mid_x, mid_z, mid_y, l2[0], l2[2], l2[1],
              length=scale, color="green", linewidth=1)
    ax.quiver(mid_x, mid_z, mid_y, l3[0], l3[2], l3[1],
              length=scale, color="blue", linewidth=1)

xs = [c[0] for c in nodes.values()]
zs = [c[2] for c in nodes.values()]
ys = [c[1] for c in nodes.values()]
ax.scatter(xs, zs, ys, color="red", s=60, depthshade=True)

for n, (x, y, z) in nodes.items():
    ax.text(x, z, y, f"  {n}", color="black", fontsize=10, fontweight="bold")

# --- Support symbols (triangles) at the pinned base nodes ---
for n in BASE_NODES:
    x, y, z = nodes[n]
    ax.scatter([x], [z], [y], marker="^", color="black", s=180, zorder=5)

# --- Global axis triad at the origin ---
axis_len = L * 0.35
ax.quiver(0, 0, 0, axis_len, 0, 0, color="black", linewidth=2, arrow_length_ratio=0.1)
ax.quiver(0, 0, 0, 0, axis_len, 0, color="black", linewidth=2, arrow_length_ratio=0.1)
ax.quiver(0, 0, 0, 0, 0, axis_len, color="black", linewidth=2, arrow_length_ratio=0.1)
ax.text(axis_len * 1.05, 0, 0, "Global X", fontsize=9, fontweight="bold")
ax.text(0, axis_len * 1.05, 0, "Global Z", fontsize=9, fontweight="bold")
ax.text(0, 0, axis_len * 1.05, "Global Y", fontsize=9, fontweight="bold")

ax.set_xlabel("X (m) - Lateral", fontsize=11, labelpad=10)
ax.set_ylabel("Z (m) - Lateral", fontsize=11, labelpad=10)
ax.set_zlabel("Y (m) - Vertical", fontsize=11, labelpad=10)

ax.set_title("6m x 6m x 6m Cube Model - Rev. 1\n"
             "(X, Z = Lateral | Y = Global Vertical | Triangles = Pinned Supports)",
             fontsize=12, fontweight="bold")

ax.set_xlim(-1, L + 1)
ax.set_ylim(-1, L + 1)
ax.set_zlim(0, L + 1)

ax.set_box_aspect([1, 1, 1])
ax.view_init(elev=20, azim=-60)

# Legend for members, supports, and local axes
legend_elems = [
    Line2D([0], [0], color="steelblue", lw=2, label="Beam"),
    Line2D([0], [0], color="darkorange", lw=2, label="Column"),
    Line2D([0], [0], marker="^", color="black", lw=0, markersize=10, label="Pinned Support"),
    Line2D([0], [0], color="red", lw=1, label="Local-1 (axial)"),
    Line2D([0], [0], color="green", lw=1, label="Local-2"),
    Line2D([0], [0], color="blue", lw=1, label="Local-3"),
]
ax.legend(handles=legend_elems, loc="upper left", fontsize=8)

plt.tight_layout()
plot_path = os.path.join(FIGURES_DIR, "cube_model.png")
plt.savefig(plot_path, dpi=150)
print(f"Saved: {plot_path}")

plt.show()