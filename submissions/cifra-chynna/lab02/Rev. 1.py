"""
Structural Solver - Rev. 1
6m x 6m x 6m Cube Frame Model
Features: Pinned supports, Beta angles, Local/Global axes, DOF labeling,
          Beam pinned releases, Structural diagram, Excel output
"""

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyArrowPatch
from mpl_toolkits.mplot3d import Axes3D
from mpl_toolkits.mplot3d.art3d import Poly3DCollection
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import os

HERE = os.path.dirname(os.path.abspath(__file__))

# =============================================================================
# MODEL DATA
# =============================================================================

REVISION = "Rev. 1"
CUBE_EDGE = 6.0  # meters
VERTICAL_AXIS = "Y"

# Node coordinates: {node_id: (X, Y, Z)}
NODES = {
    1: (0.0, 0.0, 0.0),
    2: (6.0, 0.0, 0.0),
    3: (6.0, 0.0, 6.0),
    4: (0.0, 0.0, 6.0),
    5: (0.0, 6.0, 0.0),
    6: (6.0, 6.0, 0.0),
    7: (6.0, 6.0, 6.0),
    8: (0.0, 6.0, 6.0),
}

# Support conditions: {node_id: support_type}
# Pinned: UX, UY, UZ restrained; RX, RY, RZ free
SUPPORTS = {
    1: "Pinned",
    2: "Pinned",
    3: "Pinned",
    4: "Pinned",
}

# Member incidences: {member_id: (node_i, node_j, type)}
MEMBERS = {
    1:  (1, 2, "Base Beam"),
    2:  (2, 3, "Base Beam"),
    3:  (3, 4, "Base Beam"),
    4:  (4, 1, "Base Beam"),
    5:  (5, 6, "Roof Beam"),
    6:  (6, 7, "Roof Beam"),
    7:  (7, 8, "Roof Beam"),
    8:  (8, 5, "Roof Beam"),
    9:  (1, 5, "Column"),
    10: (2, 6, "Column"),
    11: (3, 7, "Column"),
    12: (4, 8, "Column"),
}

# Beta angles by member type (degrees)
BETA_ANGLES = {
    "Base Beam": 0,
    "Roof Beam": 0,
    "Column":    90,
}

# Member releases: {member_id: {end: {dof: 'Fixed'/'Released'}}}
# For a beam pinned at x-direction (axial) and moment at z-direction:
# i-end pinned means Fx released; Mz released
# All members are Fixed-Fixed for now (fully restrained ends)
RELEASES = {}  # Empty means all Fixed

# DOF per node (6 DOF: UX, UY, UZ, RX, RY, RZ)
DOF_PER_NODE = 6
DOF_LABELS = ["UX", "UY", "UZ", "RX", "RY", "RZ"]
DOF_DESCRIPTIONS = [
    "Translation X", "Translation Y", "Translation Z",
    "Rotation X",    "Rotation Y",    "Rotation Z"
]


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def compute_member_length(ni, nj):
    """Return Euclidean length of member from node ni to nj."""
    xi, yi, zi = NODES[ni]
    xj, yj, zj = NODES[nj]
    return np.sqrt((xj - xi)**2 + (yj - yi)**2 + (zj - zi)**2)


def compute_local_axes(ni, nj, member_type, beta_deg):
    """
    Compute local axes (x, y, z) for a member.

    Convention (RISA/STAAD style):
      local x  : unit vector from node i to node j
      local y  : in the vertical plane containing local x, upward component
                 (for non-vertical members); for vertical members (columns),
                 determined by beta angle
      local z  : completes the right-handed set (local x cross local y)
      Beta angle rotates local y and z about local x.

    For columns (vertical members), the limiting case is used:
      local z is parallel to global Z.
    """
    xi, yi, zi = NODES[ni]
    xj, yj, zj = NODES[nj]
    dx = xj - xi
    dy = yj - yi
    dz = zj - zi
    L = np.sqrt(dx**2 + dy**2 + dz**2)
    ex = np.array([dx / L, dy / L, dz / L])   # local x

    # Global Y (vertical)
    g_Y = np.array([0.0, 1.0, 0.0])
    # Global Z (lateral)
    g_Z = np.array([0.0, 0.0, 1.0])

    beta = np.radians(beta_deg)

    # Check if member is vertical (parallel to global Y)
    if abs(abs(ex[1]) - 1.0) < 1e-9:
        # Vertical member (column): limiting case
        # local z parallel to global Z
        ez0 = g_Z
        ey0 = np.cross(ez0, ex)
        ey0 = ey0 / np.linalg.norm(ey0)
    else:
        # Non-vertical member
        # Reference ey0 in vertical plane containing ex
        # ey0 perpendicular to ex, with upward component
        p = g_Y - np.dot(g_Y, ex) * ex
        norm_p = np.linalg.norm(p)
        if norm_p < 1e-9:
            p = g_Z - np.dot(g_Z, ex) * ex
            norm_p = np.linalg.norm(p)
        ey0 = p / norm_p
        ez0 = np.cross(ex, ey0)
        ez0 = ez0 / np.linalg.norm(ez0)

    # Apply beta rotation about local x
    ey = np.cos(beta) * ey0 + np.sin(beta) * ez0
    ez = -np.sin(beta) * ey0 + np.cos(beta) * ez0

    return ex, ey, ez


def assign_dof(nodes, supports):
    """
    Assign global DOF numbers and equation numbers to all nodes.

    Returns:
        node_dof   : {node_id: [global_dof_1, ..., global_dof_6]}
        dof_status : {global_dof: 'Active' | 'Restrained'}
        dof_eq     : {global_dof: equation_number or None}
    """
    node_dof = {}
    dof_status = {}

    # Assign sequential DOFs
    for nid in sorted(nodes.keys()):
        base = (nid - 1) * DOF_PER_NODE + 1
        node_dof[nid] = list(range(base, base + DOF_PER_NODE))

    # Determine status: Pinned restrains UX(1), UY(2), UZ(3)
    for nid, dofs in node_dof.items():
        sup = supports.get(nid, None)
        for local_idx, gdof in enumerate(dofs):
            if sup == "Pinned" and local_idx < 3:  # UX, UY, UZ
                dof_status[gdof] = "Restrained"
            else:
                dof_status[gdof] = "Active"

    # Assign equation numbers to active DOFs
    eq = 1
    dof_eq = {}
    for gdof in sorted(dof_status.keys()):
        if dof_status[gdof] == "Active":
            dof_eq[gdof] = eq
            eq += 1
        else:
            dof_eq[gdof] = None

    return node_dof, dof_status, dof_eq


def compute_restraint_code(node_id, supports):
    """Return 6-character binary restraint code (1=restrained, 0=free)."""
    sup = supports.get(node_id, None)
    if sup == "Pinned":
        return "111000"   # UX UY UZ restrained; RX RY RZ free
    return "000000"


# =============================================================================
# COMPUTE MODEL DATA
# =============================================================================

def build_model():
    """Build and return all computed model data."""
    node_dof, dof_status, dof_eq = assign_dof(NODES, SUPPORTS)

    total_dof = len(NODES) * DOF_PER_NODE
    restrained_dof = sum(1 for s in dof_status.values() if s == "Restrained")
    active_dof = total_dof - restrained_dof

    # Member data
    member_data = {}
    for mid, (ni, nj, mtype) in MEMBERS.items():
        beta_deg = BETA_ANGLES[mtype]
        L = compute_member_length(ni, nj)
        ex, ey, ez = compute_local_axes(ni, nj, mtype, beta_deg)
        member_data[mid] = {
            "ni": ni, "nj": nj, "type": mtype,
            "length": L, "beta": beta_deg,
            "ex": ex, "ey": ey, "ez": ez,
        }

    return {
        "node_dof": node_dof,
        "dof_status": dof_status,
        "dof_eq": dof_eq,
        "total_dof": total_dof,
        "restrained_dof": restrained_dof,
        "active_dof": active_dof,
        "member_data": member_data,
    }


# =============================================================================
# 3D STRUCTURAL DIAGRAM
# =============================================================================

def draw_pinned_symbol(ax, x, y, z, size=0.18):
    """
    Draw a 3D pinned support symbol (pyramid/triangle) at (x, y, z).
    The pyramid points downward from the support node.
    """
    h = size * 2.2
    # Four base vertices below the node
    pts = np.array([
        [x - size, y - h, z - size],
        [x + size, y - h, z - size],
        [x + size, y - h, z + size],
        [x - size, y - h, z + size],
    ])
    apex = np.array([x, y, z])

    # Four triangular faces
    faces = [
        [apex, pts[0], pts[1]],
        [apex, pts[1], pts[2]],
        [apex, pts[2], pts[3]],
        [apex, pts[3], pts[0]],
        [pts[0], pts[1], pts[2], pts[3]],   # base (quad)
    ]
    poly = Poly3DCollection(faces, alpha=0.75,
                            facecolor='dimgray', edgecolor='black',
                            linewidth=0.5)
    ax.add_collection3d(poly)


def draw_local_axes(ax, ni, nj, mtype, beta_deg, scale=0.7):
    """Draw local x (red), y (green), z (purple) axes at member midpoint."""
    xi, yi, zi = NODES[ni]
    xj, yj, zj = NODES[nj]
    mx = (xi + xj) / 2
    my = (yi + yj) / 2
    mz = (zi + zj) / 2

    ex, ey, ez = compute_local_axes(ni, nj, mtype, beta_deg)

    arrow_kw = dict(mutation_scale=8, linewidth=1.2)
    # local x - red
    ax.quiver(mx, my, mz, ex[0], ex[1], ex[2],
              length=scale, color='red', arrow_length_ratio=0.35,
              linewidth=1.1)
    # local y - green
    ax.quiver(mx, my, mz, ey[0], ey[1], ey[2],
              length=scale, color='limegreen', arrow_length_ratio=0.35,
              linewidth=1.1)
    # local z - purple
    ax.quiver(mx, my, mz, ez[0], ez[1], ez[2],
              length=scale, color='mediumpurple', arrow_length_ratio=0.35,
              linewidth=1.1)


def plot_structural_model(model, output_path):
    """Generate and save the 3D structural model diagram."""
    fig = plt.figure(figsize=(18, 11), facecolor='white')

    # ── main 3D axis ──────────────────────────────────────────────────────────
    ax = fig.add_axes([0.04, 0.04, 0.70, 0.90], projection='3d')
    ax.set_facecolor('#f0f4f8')

    fig.suptitle(
        f"6m x 6m x 6m Cube - Structural Model, {REVISION}\n"
        "Pinned supports at nodes 1-4, member local axes and beta angles shown",
        fontsize=11, fontweight='bold', y=0.99
    )

    member_data = model["member_data"]
    node_dof    = model["node_dof"]

    # ── draw members ──────────────────────────────────────────────────────────
    for mid, md in member_data.items():
        ni, nj, mtype = md["ni"], md["nj"], md["type"]
        xi, yi, zi = NODES[ni]
        xj, yj, zj = NODES[nj]
        color = '#1f77b4' if "Beam" in mtype else '#2ca02c'
        lw = 2.8 if "Beam" in mtype else 2.4
        ax.plot([xi, xj], [yi, yj], [zi, zj],
                color=color, linewidth=lw, zorder=3)

        # Local axes arrows
        draw_local_axes(ax, ni, nj, mtype, md["beta"], scale=0.65)

    # ── draw nodes ────────────────────────────────────────────────────────────
    for nid, (x, y, z) in NODES.items():
        dof_range = node_dof[nid]
        dof_label = f"DOF {dof_range[0]}-{dof_range[-1]}"

        if nid in SUPPORTS:
            # Pinned support: filled red circle + pyramid
            ax.scatter([x], [y], [z], color='red', s=70, zorder=6,
                       depthshade=False)
            draw_pinned_symbol(ax, x, y, z, size=0.20)
            label = f"N{nid}\n{dof_label}"
        else:
            # Free node: salmon circle
            ax.scatter([x], [y], [z], color='salmon', s=60, zorder=6,
                       depthshade=False)
            label = f"N{nid}\n{dof_label}"

        # Offset label to avoid overlap
        offsets = {
            1: (-0.55, 0.0,  0.10),
            2: ( 0.25, 0.0, -0.30),
            3: ( 0.25, 0.0,  0.35),
            4: (-0.60, 0.0,  0.10),
            5: (-0.65, 0.0, -0.10),
            6: ( 0.10, 0.0, -0.30),
            7: ( 0.25, 0.0,  0.25),
            8: (-0.65, 0.0,  0.10),
        }
        ox, oy, oz = offsets.get(nid, (0.2, 0.0, 0.2))
        ax.text(x + ox, y + oy, z + oz, label,
                fontsize=6.8, ha='left', va='center',
                fontweight='bold', color='navy')

    # ── draw member labels with beta angle ────────────────────────────────────
    for mid, md in member_data.items():
        ni, nj = md["ni"], md["nj"]
        xi, yi, zi = NODES[ni]
        xj, yj, zj = NODES[nj]
        mx = (xi + xj) / 2
        my = (yi + yj) / 2
        mz = (zi + zj) / 2
        beta = md["beta"]
        if beta != 0:
            lbl = f"M{mid} (β={beta}°)"
        else:
            lbl = f"M{mid}"
        ax.text(mx, my, mz, lbl,
                fontsize=6.2, color='#333333', ha='center', va='bottom')

    # ── origin star ───────────────────────────────────────────────────────────
    ax.scatter([0], [0], [0], marker='*', color='green', s=120, zorder=8,
               depthshade=False)

    # ── axes formatting ───────────────────────────────────────────────────────
    ax.set_xlabel("X (m) - lateral", fontsize=8, labelpad=8)
    ax.set_ylabel("Y (m) - vertical", fontsize=8, labelpad=8)
    ax.set_zlabel("Z (m) - lateral", fontsize=8, labelpad=8)
    ax.set_xlim(-1, 7)
    ax.set_ylim(-1, 7)
    ax.set_zlim(-1, 7)
    ax.tick_params(labelsize=7)
    ax.view_init(elev=22, azim=-55)

    # ── legend ────────────────────────────────────────────────────────────────
    from matplotlib.lines import Line2D
    from matplotlib.patches import Patch

    legend_elements = [
        Line2D([0], [0], color='#1f77b4', lw=2.5, label='Beam'),
        Line2D([0], [0], color='#2ca02c', lw=2.5, label='Column'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='red',
               markersize=7, label='Supported node (pinned)'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='salmon',
               markersize=7, label='Free node'),
        Line2D([0], [0], marker='*', color='w', markerfacecolor='green',
               markersize=10, label='Origin (0, 0, 0)'),
        Line2D([0], [0], color='red',         lw=1.5, label='Local x axis'),
        Line2D([0], [0], color='limegreen',   lw=1.5, label='Local y axis'),
        Line2D([0], [0], color='mediumpurple',lw=1.5, label='Local z axis'),
    ]
    ax.legend(handles=legend_elements, loc='upper left',
              fontsize=7, framealpha=0.85, ncol=1,
              bbox_to_anchor=(-0.02, 1.0))

    # ── MODEL DATA info box ───────────────────────────────────────────────────
    info_ax = fig.add_axes([0.76, 0.10, 0.23, 0.82])
    info_ax.axis('off')
    info_ax.set_facecolor('#f9f9f9')

    info_lines = [
        ("MODEL DATA - REV. 1", True),
        ("", False),
        ("Geometry", True),
        (f"  Cube edge       {CUBE_EDGE:.1f} m", False),
        (f"  Nodes           {len(NODES)}", False),
        (f"  Members         {len(MEMBERS)}", False),
        (f"  Vertical axis   global {VERTICAL_AXIS}", False),
        ("", False),
        ("Supports", True),
        ("  Type            pinned", False),
        ("  Nodes           1, 2, 3, 4", False),
        ("  Restrained      UX, UY, UZ", False),
        ("  Released        RX, RY, RZ", False),
        ("", False),
        ("Degrees of freedom", True),
        (f"  DOF per node    {DOF_PER_NODE}", False),
        (f"  Total DOF       {model['total_dof']}", False),
        (f"  Restrained DOF  {model['restrained_dof']}", False),
        (f"  Active DOF      {model['active_dof']}", False),
        ("  Numbering       (node - 1) x 6 + 1..6", False),
        ("", False),
        ("Beta angles", True),
        ("  Base Beam       0 deg", False),
        ("  Roof Beam       0 deg", False),
        ("  Column          90 deg", False),
        ("", False),
        ("Local axes", True),
        ("  local x  start node i to end node j", False),
        ("  local y  in the vertical plane, upward", False),
        ("  local z  completes the right-handed set", False),
        ("  Vertical members follow the limiting", False),
        ("  case: local z parallel to global Z.", False),
    ]

    y_pos = 0.98
    for line, bold in info_lines:
        weight = 'bold' if bold else 'normal'
        info_ax.text(0.04, y_pos, line, transform=info_ax.transAxes,
                     fontsize=7.0, va='top', fontweight=weight,
                     fontfamily='monospace')
        y_pos -= 0.032

    # Box border
    rect = mpatches.FancyBboxPatch((0, 0), 1, 1,
                                   boxstyle="round,pad=0.01",
                                   linewidth=1.2, edgecolor='#555',
                                   facecolor='#fafafa',
                                   transform=info_ax.transAxes,
                                   zorder=-1)
    info_ax.add_patch(rect)

    plt.savefig(output_path, dpi=150, bbox_inches='tight',
                facecolor='white')
    plt.close(fig)
    print(f"  [OK] Diagram saved: {output_path}")


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def _hdr_fill(color_hex):
    return PatternFill("solid", fgColor=color_hex)

def _thin_border():
    s = Side(style='thin', color='999999')
    return Border(left=s, right=s, top=s, bottom=s)

def _med_border():
    s = Side(style='medium', color='444444')
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_header(ws, row, col, value, bg='1F3864', fg='FFFFFF',
                  bold=True, font_size=9):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill    = PatternFill("solid", fgColor=bg)
    cell.font    = Font(bold=bold, color=fg, name='Arial', size=font_size)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                               wrap_text=True)
    cell.border  = _thin_border()
    return cell

def _apply_data(ws, row, col, value, bg=None, bold=False,
                align='center', font_size=9, number_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font  = Font(bold=bold, name='Arial', size=font_size)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = _thin_border()
    if number_fmt:
        cell.number_format = number_fmt
    return cell

def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def write_excel(model, output_path):
    """Write all model data to a formatted Excel workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    node_dof    = model["node_dof"]
    dof_status  = model["dof_status"]
    dof_eq      = model["dof_eq"]
    member_data = model["member_data"]

    # ------------------------------------------------------------------
    # Sheet 1: Model Summary
    # ------------------------------------------------------------------
    ws1 = wb.create_sheet("model summary")
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 30

    # Title
    ws1.merge_cells('A1:B1')
    c = ws1.cell(row=1, column=1,
                 value=f"Structural Model Summary — {REVISION}")
    c.font      = Font(bold=True, size=13, name='Arial', color='1F3864')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill      = PatternFill("solid", fgColor='D6E4F7')

    _apply_header(ws1, 2, 1, "Item",  bg='2E5FA3')
    _apply_header(ws1, 2, 2, "Value", bg='2E5FA3')

    rows = [
        ("Revision",                  REVISION),
        ("Cube edge length (m)",       CUBE_EDGE),
        ("Number of nodes",            len(NODES)),
        ("Number of members",          len(MEMBERS)),
        ("Supported nodes",            len(SUPPORTS)),
        ("Support type",               "Pinned"),
        ("DOF per node",               DOF_PER_NODE),
        ("Total DOF",                  model["total_dof"]),
        ("Restrained DOF",             model["restrained_dof"]),
        ("Active DOF (equations)",     model["active_dof"]),
        ("Global vertical axis",       f"Y"),
        ("Global lateral axes",        "X and Z"),
        ("Beta angle, base beams (deg)", BETA_ANGLES["Base Beam"]),
        ("Beta angle, roof beams (deg)", BETA_ANGLES["Roof Beam"]),
        ("Beta angle, columns (deg)",    BETA_ANGLES["Column"]),
    ]

    alt = ['FFFFFF', 'EDF2FB']
    for i, (item, val) in enumerate(rows, start=3):
        bg = alt[i % 2]
        _apply_data(ws1, i, 1, item, bg=bg, align='left')
        _apply_data(ws1, i, 2, val,  bg=bg, align='center')

    col_width(ws1, 1, 32)
    col_width(ws1, 2, 20)

    # ------------------------------------------------------------------
    # Sheet 2: Node Coordinates & Supports
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("node")
    ws2.sheet_view.showGridLines = False

    headers = ["Node", "X (m)", "Y (m)", "Z (m)", "Support"]
    for c_idx, h in enumerate(headers, 1):
        _apply_header(ws2, 1, c_idx, h, bg='2E5FA3')

    for r, (nid, (x, y, z)) in enumerate(sorted(NODES.items()), start=2):
        sup = SUPPORTS.get(nid, "Free")
        bg  = 'FFE0E0' if sup == "Pinned" else 'FFFFFF'
        _apply_data(ws2, r, 1, nid, bg=bg)
        _apply_data(ws2, r, 2, x,   bg=bg)
        _apply_data(ws2, r, 3, y,   bg=bg)
        _apply_data(ws2, r, 4, z,   bg=bg)
        _apply_data(ws2, r, 5, sup, bg=bg,
                    bold=(sup == "Pinned"))

    for ci, w in enumerate([8, 10, 10, 10, 12], 1):
        col_width(ws2, ci, w)

    # ------------------------------------------------------------------
    # Sheet 3: Member Incidences
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("member incidences")
    ws3.sheet_view.showGridLines = False

    hdrs3 = ["Member", "Node i (Start)", "Node j (End)",
             "Type", "Length (m)", "Beta (deg)"]
    for ci, h in enumerate(hdrs3, 1):
        _apply_header(ws3, 1, ci, h, bg='2E5FA3')

    type_colors = {"Base Beam": 'E8F4FD', "Roof Beam": 'EDF7EE',
                   "Column": 'FEF9E7'}
    for r, (mid, md) in enumerate(sorted(member_data.items()), start=2):
        bg = type_colors.get(md["type"], 'FFFFFF')
        _apply_data(ws3, r, 1, mid,          bg=bg)
        _apply_data(ws3, r, 2, md["ni"],     bg=bg)
        _apply_data(ws3, r, 3, md["nj"],     bg=bg)
        _apply_data(ws3, r, 4, md["type"],   bg=bg, align='left')
        _apply_data(ws3, r, 5, md["length"], bg=bg)
        _apply_data(ws3, r, 6, md["beta"],   bg=bg)

    for ci, w in enumerate([10, 14, 14, 14, 13, 12], 1):
        col_width(ws3, ci, w)

    # ------------------------------------------------------------------
    # Sheet 4: Support Conditions
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("supports")
    ws4.sheet_view.showGridLines = False

    hdrs4 = ["Node", "X (m)", "Y (m)", "Z (m)", "Support Type",
              "UX", "UY", "UZ", "RX", "RY", "RZ", "Restraint Code"]
    for ci, h in enumerate(hdrs4, 1):
        _apply_header(ws4, 1, ci, h, bg='2E5FA3')

    for r, (nid, (x, y, z)) in enumerate(sorted(NODES.items()), start=2):
        sup = SUPPORTS.get(nid, None)
        sup_type = sup if sup else "Free"
        bg = 'FFE0E0' if sup else 'FFFFFF'
        _apply_data(ws4, r, 1, nid,      bg=bg)
        _apply_data(ws4, r, 2, x,        bg=bg)
        _apply_data(ws4, r, 3, y,        bg=bg)
        _apply_data(ws4, r, 4, z,        bg=bg)
        _apply_data(ws4, r, 5, sup_type, bg=bg, bold=bool(sup))

        if sup == "Pinned":
            for ci, lbl in enumerate(["Restrained"]*3 + ["Active"]*3, start=6):
                color = 'FF4444' if lbl == "Restrained" else '22AA22'
                c = ws4.cell(row=r, column=ci, value=lbl)
                c.font      = Font(bold=True, color=color, name='Arial', size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = _thin_border()
                c.fill      = PatternFill("solid", fgColor=bg)
        else:
            for ci in range(6, 12):
                _apply_data(ws4, r, ci, "Active", bg=bg)

        code = compute_restraint_code(nid, SUPPORTS)
        _apply_data(ws4, r, 12, code, bg=bg, bold=bool(sup))

    for ci, w in enumerate([8, 8, 8, 8, 14, 12, 12, 12, 12, 12, 12, 16], 1):
        col_width(ws4, ci, w)

    # ------------------------------------------------------------------
    # Sheet 5: Local Axes
    # ------------------------------------------------------------------
    ws5 = wb.create_sheet("local axes")
    ws5.sheet_view.showGridLines = False

    hdrs5 = ["Member", "Node i", "Node j", "Type", "Length (m)", "Beta (deg)",
             "local x - X", "local x - Y", "local x - Z",
             "local y - X", "local y - Y", "local y - Z",
             "local z - X", "local z - Y", "local z - Z"]
    for ci, h in enumerate(hdrs5, 1):
        _apply_header(ws5, 1, ci, h, bg='2E5FA3')

    type_colors5 = {"Base Beam": 'E8F4FD', "Roof Beam": 'EDF7EE',
                    "Column": 'FEF9E7'}
    for r, (mid, md) in enumerate(sorted(member_data.items()), start=2):
        bg = type_colors5.get(md["type"], 'FFFFFF')
        ex, ey, ez = md["ex"], md["ey"], md["ez"]
        vals = [mid, md["ni"], md["nj"], md["type"], md["length"], md["beta"],
                round(ex[0], 4), round(ex[1], 4), round(ex[2], 4),
                round(ey[0], 4), round(ey[1], 4), round(ey[2], 4),
                round(ez[0], 4), round(ez[1], 4), round(ez[2], 4)]
        for ci, v in enumerate(vals, 1):
            align = 'left' if ci == 4 else 'center'
            _apply_data(ws5, r, ci, v, bg=bg, align=align)

    widths5 = [10, 8, 8, 14, 12, 12,
               12, 12, 12, 12, 12, 12, 12, 12, 12]
    for ci, w in enumerate(widths5, 1):
        col_width(ws5, ci, w)

    # ------------------------------------------------------------------
    # Sheet 6: Node DOF Summary
    # ------------------------------------------------------------------
    ws6 = wb.create_sheet("node DOF")
    ws6.sheet_view.showGridLines = False

    hdrs6 = ["Node", "UX", "UY", "UZ", "RX", "RY", "RZ", "Raw DOF Range"]
    for ci, h in enumerate(hdrs6, 1):
        _apply_header(ws6, 1, ci, h, bg='2E5FA3')

    for r, nid in enumerate(sorted(NODES.keys()), start=2):
        dofs = node_dof[nid]
        sup = SUPPORTS.get(nid, None)
        bg  = 'FFE0E0' if sup else 'FFFFFF'

        _apply_data(ws6, r, 1, nid, bg=bg, bold=True)
        for ldi, gdof in enumerate(dofs):
            ci = ldi + 2
            status = dof_status[gdof]
            eq_num = dof_eq[gdof]
            if status == "Restrained":
                cell_val = "R"
            else:
                cell_val = eq_num if eq_num else gdof
            c = ws6.cell(row=r, column=ci, value=cell_val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = _thin_border()
            c.alignment = Alignment(horizontal='center', vertical='center')
            if status == "Restrained":
                c.font = Font(bold=True, color='CC0000', name='Arial', size=9)
            else:
                c.font = Font(color='005500', name='Arial', size=9)

        _apply_data(ws6, r, 8, f"{dofs[0]}-{dofs[-1]}", bg=bg)

    # Totals
    rr = len(NODES) + 3
    ws6.cell(row=rr,   column=1, value="Total DOF:").font = Font(bold=True, name='Arial', size=9)
    ws6.cell(row=rr,   column=2, value=model["total_dof"])
    ws6.cell(row=rr+1, column=1, value="Active (Free) DOF:").font = Font(bold=True, name='Arial', size=9)
    ws6.cell(row=rr+1, column=2, value=model["active_dof"])
    ws6.cell(row=rr+2, column=1, value="Restrained DOF:").font = Font(bold=True, name='Arial', size=9)
    ws6.cell(row=rr+2, column=2, value=model["restrained_dof"])

    for ci, w in enumerate([8, 8, 8, 8, 8, 8, 8, 14], 1):
        col_width(ws6, ci, w)

    # ------------------------------------------------------------------
    # Sheet 7: DOF Numbering Detail
    # ------------------------------------------------------------------
    ws7 = wb.create_sheet("DOF numbering")
    ws7.sheet_view.showGridLines = False

    hdrs7 = ["Node", "Local DOF", "DOF", "Description",
              "Global DOF No.", "Status", "Equation No."]
    for ci, h in enumerate(hdrs7, 1):
        _apply_header(ws7, 1, ci, h, bg='2E5FA3')

    r = 2
    for nid in sorted(NODES.keys()):
        dofs = node_dof[nid]
        for ldi, gdof in enumerate(dofs):
            status = dof_status[gdof]
            eq_num = dof_eq[gdof]
            sup = SUPPORTS.get(nid, None)
            bg  = 'FFE0E0' if status == "Restrained" else (
                  'E8F4FD' if sup else 'FFFFFF')
            _apply_data(ws7, r, 1, nid,                    bg=bg)
            _apply_data(ws7, r, 2, ldi + 1,                bg=bg)
            _apply_data(ws7, r, 3, DOF_LABELS[ldi],        bg=bg, bold=True)
            _apply_data(ws7, r, 4, DOF_DESCRIPTIONS[ldi],  bg=bg, align='left')
            _apply_data(ws7, r, 5, gdof,                   bg=bg)
            color_st = 'CC0000' if status == "Restrained" else '006600'
            c6 = ws7.cell(row=r, column=6, value=status)
            c6.font      = Font(bold=True, color=color_st, name='Arial', size=9)
            c6.fill      = PatternFill("solid", fgColor=bg)
            c6.alignment = Alignment(horizontal='center', vertical='center')
            c6.border    = _thin_border()
            eq_val = eq_num if eq_num else "-"
            _apply_data(ws7, r, 7, eq_val, bg=bg)
            r += 1

    for ci, w in enumerate([8, 10, 8, 16, 15, 14, 14], 1):
        col_width(ws7, ci, w)

    # ------------------------------------------------------------------
    # Sheet 8: Member Releases
    # ------------------------------------------------------------------
    ws8 = wb.create_sheet("member releases")
    ws8.sheet_view.showGridLines = False

    hdrs8 = ["Member", "End", "Fx", "Fy", "Fz", "Mx", "My", "Mz"]
    for ci, h in enumerate(hdrs8, 1):
        _apply_header(ws8, 1, ci, h, bg='2E5FA3')

    r = 2
    type_colors8 = {"Base Beam": 'E8F4FD', "Roof Beam": 'EDF7EE',
                    "Column": 'FEF9E7'}
    for mid, md in sorted(member_data.items()):
        bg = type_colors8.get(md["type"], 'FFFFFF')
        for end_idx, (end_node, end_label) in enumerate(
                [(md["ni"], f"i (Node {md['ni']})"),
                 (md["nj"], f"j (Node {md['nj']})")]):
            _apply_data(ws8, r, 1, mid,       bg=bg)
            _apply_data(ws8, r, 2, end_label, bg=bg, align='left')
            # Check releases dict; default Fixed
            rel = RELEASES.get(mid, {}).get(end_idx, {})
            for ci, dof_name in enumerate(["Fx","Fy","Fz","Mx","My","Mz"], 3):
                val = rel.get(dof_name, "Fixed")
                color = 'CC0000' if val == "Released" else '333333'
                c = ws8.cell(row=r, column=ci, value=val)
                c.font      = Font(color=color, name='Arial', size=9,
                                   bold=(val=="Released"))
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = _thin_border()
            r += 1

    for ci, w in enumerate([10, 16, 10, 10, 10, 10, 10, 10], 1):
        col_width(ws8, ci, w)

    # ── freeze header rows & set tab colors ──────────────────────────────────
    tab_colors = {
        "model summary":   "1F3864",
        "node":            "2E75B6",
        "member incidences":"2E75B6",
        "supports":        "C00000",
        "local axes":      "375623",
        "node DOF":        "7030A0",
        "DOF numbering":   "7030A0",
        "member releases": "833C00",
    }
    for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8]:
        ws.freeze_panes = 'A2'
        name = ws.title
        if name in tab_colors:
            ws.sheet_properties.tabColor = tab_colors[name]

    wb.save(output_path)
    print(f"  [OK] Excel saved: {output_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print(f"\n{'='*60}")
    print(f"  Structural Solver — {REVISION}")
    print(f"  6m x 6m x 6m Cube Frame Model")
    print(f"{'='*60}\n")

    # Build model
    model = build_model()

    print("  Model statistics:")
    print(f"    Nodes           : {len(NODES)}")
    print(f"    Members         : {len(MEMBERS)}")
    print(f"    Supports        : {len(SUPPORTS)} (Pinned)")
    print(f"    Total DOF       : {model['total_dof']}")
    print(f"    Restrained DOF  : {model['restrained_dof']}")
    print(f"    Active DOF      : {model['active_dof']}")
    print()

    # Output paths
    output_path = os.path.join(HERE, "structural_model_rev1.png")
    diagram_path = os.path.join(HERE, "structural_model_rev1.png")
    excel_path = os.path.join(HERE, "Book 1.xlsx")

    # Generate outputs
    print("  Generating structural diagram...")
    plot_structural_model(model, diagram_path)

    print("  Generating Excel report...")
    write_excel(model, excel_path)

    print(f"\n  Done. Output files:")
    print(f"    {diagram_path}")
    print(f"    {excel_path}")
    print(f"\n{'='*60}\n")

    return diagram_path, excel_path


if __name__ == "__main__":
    main()