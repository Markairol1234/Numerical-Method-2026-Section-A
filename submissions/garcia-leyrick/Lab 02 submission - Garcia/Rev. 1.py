"""
cube_geometry_rev2.py   (Rev. 2)

Rebuilds the 6 m x 6 m x 6 m cube structural model, matching the workbook
layout of the reference file "cube_model (1).xlsx":

    Sheets -> Model Summary, Nodes, Member Incidences, Supports,
              Local Axes, Node DOF, DOF Numbering, Member Releases

Axis convention (SAP2000 / ETABS / STAAD / RISA / Robot style):
    X, Z  -> horizontal / lateral (plan) directions
    Y     -> global vertical (up) direction

Rev. 2 change (relative to Rev. 1):
    * ALL member end connections are now FULLY FIXED (rigid) for every
      force and moment component (Fx, Fy, Fz, Mx, My, Mz), at both the
      i-end and j-end of every member (base beams, roof beams and
      columns alike). This matches the "Member Releases" sheet of the
      reference workbook, where every row reads "Fixed". Rev. 1's
      partial beam releases (Rx', Rz' pinned) have been removed -- the
      frame is now a fully rigid moment frame, only the base supports
      (nodes 1-4, pinned) allow rotation.

Node numbering (bottom story = Y=0 plan, top story = Y=L plan):
    Bottom (Y=0): 1 -> (0,0,0)  2 -> (L,0,0)  3 -> (L,0,L)  4 -> (0,0,L)
    Top    (Y=L): 5 -> (0,L,0)  6 -> (L,L,0)  7 -> (L,L,L)  8 -> (0,L,L)

Member numbering (matches the reference workbook):
    1-4   : base beams   (1-2, 2-3, 3-4, 4-1)
    5-8   : roof beams   (5-6, 6-7, 7-8, 8-5)
    9-12  : columns      (1-5, 2-6, 3-7, 4-8)

Outputs:
  1. cube_model_rev2.xlsx : 8-sheet workbook (Model Summary, Nodes,
     Member Incidences, Supports, Local Axes, Node DOF, DOF Numbering,
     Member Releases) -- values only, matching the reference file's
     plain (formula-free) convention.
  2. cube_6x6x6_rev2.png : 3D plot of the cube, Rev. 1 diagram style,
     updated to show fully fixed (rigid) member end connections.
"""

import subprocess
import sys

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from matplotlib.lines import Line2D
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401 (registers 3D projection)
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd

EDGE_LENGTH = 6.0  # meters
XLSX = "cube_model_rev2.xlsx"
PNG = "cube_6x6x6_rev2.png"
RECALC = "/mnt/skills/public/xlsx/scripts/recalc.py"

DOF_PER_NODE = 6
DOF_LABELS = ["UX", "UY", "UZ", "RX", "RY", "RZ"]
DOF_DESCRIPTIONS = {
    "UX": "Translation X",
    "UY": "Translation Y",
    "UZ": "Translation Z",
    "RX": "Rotation X",
    "RY": "Rotation Y",
    "RZ": "Rotation Z",
}
FORCE_LABELS = ["Fx", "Fy", "Fz", "Mx", "My", "Mz"]

# ---------------------------------------------------------------------------
# Geometry
# ---------------------------------------------------------------------------
NODE_COEFFS = {
    1: (0, 0, 0),
    2: (1, 0, 0),
    3: (1, 0, 1),
    4: (0, 0, 1),
    5: (0, 1, 0),
    6: (1, 1, 0),
    7: (1, 1, 1),
    8: (0, 1, 1),
}
NODE_COORDS = {n: tuple(c * EDGE_LENGTH for c in coeffs) for n, coeffs in NODE_COEFFS.items()}

# Member incidences (i = start node, j = end node), numbered to match the
# reference workbook: base beams, then roof beams, then columns.
MEMBERS_RAW = [
    (1, 2, "Base Beam"),
    (2, 3, "Base Beam"),
    (3, 4, "Base Beam"),
    (4, 1, "Base Beam"),
    (5, 6, "Roof Beam"),
    (6, 7, "Roof Beam"),
    (7, 8, "Roof Beam"),
    (8, 5, "Roof Beam"),
    (1, 5, "Column"),
    (2, 6, "Column"),
    (3, 7, "Column"),
    (4, 8, "Column"),
]

BETA_ANGLE = {"Base Beam": 0.0, "Roof Beam": 0.0, "Column": 90.0}
MEMBER_COLOR = {"Base Beam": "royalblue", "Roof Beam": "royalblue", "Column": "seagreen"}

SUPPORT_NODES = {1, 2, 3, 4}
SUPPORT_TYPE = "Pinned"
# Restrained flags per DOF (DOF_LABELS order): 1 = restrained, 0 = active/free
PINNED_RESTRAINT = [1, 1, 1, 0, 0, 0]
FREE_RESTRAINT = [0, 0, 0, 0, 0, 0]


def node_restraint(node):
  return PINNED_RESTRAINT if node in SUPPORT_NODES else FREE_RESTRAINT


def restraint_code(node):
  return "".join(str(f) for f in node_restraint(node))


# ---------------------------------------------------------------------------
# Local axis calculation
# ---------------------------------------------------------------------------
def local_axes_base(p_i, p_j):
  """Return unit (local_x, local_y, local_z) for a member from p_i to p_j,
  using global Y as the reference vector (global Z for vertical members,
  the limiting case), following the reference workbook's convention: the
  tabulated axes are this un-rotated (beta = 0) triad; the beta angle is
  reported alongside as a separate member property.
  """
  p_i = np.array(p_i, dtype=float)
  p_j = np.array(p_j, dtype=float)
  local_x = p_j - p_i
  local_x /= np.linalg.norm(local_x)

  global_y = np.array([0.0, 1.0, 0.0])
  vertical = np.isclose(abs(np.dot(local_x, global_y)), 1.0)
  reference = np.array([0.0, 0.0, 1.0]) if vertical else global_y

  local_z = np.cross(local_x, reference)
  local_z /= np.linalg.norm(local_z)
  local_y = np.cross(local_z, local_x)
  local_y /= np.linalg.norm(local_y)
  return local_x, local_y, local_z


# ---------------------------------------------------------------------------
# Assemble members
# ---------------------------------------------------------------------------
MEMBERS = []
for idx, (i_node, j_node, mtype) in enumerate(MEMBERS_RAW, start=1):
  beta = BETA_ANGLE[mtype]
  lx, ly, lz = local_axes_base(NODE_COORDS[i_node], NODE_COORDS[j_node])
  length = float(np.linalg.norm(np.array(NODE_COORDS[j_node]) - np.array(NODE_COORDS[i_node])))
  MEMBERS.append(
      {
          "id": idx,
          "i": i_node,
          "j": j_node,
          "type": mtype,
          "length": length,
          "beta": beta,
          "local_x": lx,
          "local_y": ly,
          "local_z": lz,
      }
  )

TOTAL_DOF = len(NODE_COEFFS) * DOF_PER_NODE
RESTRAINED_DOF = sum(sum(node_restraint(n)) for n in NODE_COEFFS)
ACTIVE_DOF = TOTAL_DOF - RESTRAINED_DOF

# ---------------------------------------------------------------------------
# Excel styles (match the reference workbook: plain header row, no title
# row, dark-blue header fill, white bold Arial header text, centered data)
# ---------------------------------------------------------------------------
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill(patternType="solid", fgColor="1F3864")
center = Alignment(horizontal="center", vertical="center")
data_font = Font(name="Arial")


def write_header(ws, headers):
  for col, text in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center


def write_row(ws, row, values):
  for col, val in enumerate(values, start=1):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = data_font
    cell.alignment = center


def set_widths(ws, widths):
  for col, width in zip("ABCDEFGHIJKLMNO", widths):
    ws.column_dimensions[col].width = width


wb = openpyxl.Workbook()

# ---------------------------------------------------------------------------
# 1. Model Summary
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Model Summary"
write_header(ws, ["Item", "Value"])
summary_rows = [
    ("Revision", "Rev. 2"),
    ("Cube edge length (m)", EDGE_LENGTH),
    ("Number of nodes", len(NODE_COEFFS)),
    ("Number of members", len(MEMBERS)),
    ("Supported nodes", len(SUPPORT_NODES)),
    ("Support type", SUPPORT_TYPE),
    ("DOF per node", DOF_PER_NODE),
    ("Total DOF", TOTAL_DOF),
    ("Restrained DOF", RESTRAINED_DOF),
    ("Active DOF (equations)", ACTIVE_DOF),
    ("Global vertical axis", "Y"),
    ("Global lateral axes", "X and Z"),
    ("Beta angle, base beams (deg)", BETA_ANGLE["Base Beam"]),
    ("Beta angle, roof beams (deg)", BETA_ANGLE["Roof Beam"]),
    ("Beta angle, columns (deg)", BETA_ANGLE["Column"]),
    ("Member end connections", "Fixed (rigid) - all members"),
]
for i, (item, value) in enumerate(summary_rows, start=2):
  write_row(ws, i, [item, value])
set_widths(ws, [30, 22])

# ---------------------------------------------------------------------------
# 2. Nodes
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Nodes")
write_header(ws, ["Node", "X (m)", "Y (m)", "Z (m)", "Support"])
for i, node in enumerate(sorted(NODE_COEFFS), start=2):
  x, y, z = NODE_COORDS[node]
  support = SUPPORT_TYPE if node in SUPPORT_NODES else "Free"
  write_row(ws, i, [node, x, y, z, support])
set_widths(ws, [10, 10, 10, 10, 12])

# ---------------------------------------------------------------------------
# 3. Member Incidences
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Member Incidences")
write_header(ws, ["Member", "Node i (Start)", "Node j (End)", "Type", "Length (m)", "Beta (deg)"])
for m in MEMBERS:
  write_row(ws, m["id"] + 1, [m["id"], m["i"], m["j"], m["type"], m["length"], m["beta"]])
set_widths(ws, [10, 16, 16, 12, 12, 12])

# ---------------------------------------------------------------------------
# 4. Supports
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Supports")
write_header(ws, ["Node", "X (m)", "Y (m)", "Z (m)", "Support Type",
                   "UX", "UY", "UZ", "RX", "RY", "RZ", "Restraint Code"])
for i, node in enumerate(sorted(NODE_COEFFS), start=2):
  x, y, z = NODE_COORDS[node]
  support = SUPPORT_TYPE if node in SUPPORT_NODES else "Free"
  restraint = node_restraint(node)
  flags = ["Restrained" if f else "Active" for f in restraint]
  write_row(ws, i, [node, x, y, z, support] + flags + [restraint_code(node)])
set_widths(ws, [8, 9, 9, 9, 13, 11, 11, 11, 11, 11, 11, 14])

# ---------------------------------------------------------------------------
# 5. Local Axes
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Local Axes")
write_header(ws, ["Member", "Node i", "Node j", "Type", "Length (m)", "Beta (deg)",
                   "local x - X", "local x - Y", "local x - Z",
                   "local y - X", "local y - Y", "local y - Z",
                   "local z - X", "local z - Y", "local z - Z"])
for m in MEMBERS:
  lx, ly, lz = m["local_x"], m["local_y"], m["local_z"]
  vals = [m["id"], m["i"], m["j"], m["type"], m["length"], m["beta"]]
  for vec in (lx, ly, lz):
    vals += [round(float(v)) if np.isclose(v, round(v)) else round(float(v), 4) for v in vec]
  write_row(ws, m["id"] + 1, vals)
set_widths(ws, [11] * 15)

# ---------------------------------------------------------------------------
# 6. Node DOF
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Node DOF")
write_header(ws, ["Node", "UX", "UY", "UZ", "RX", "RY", "RZ", "Raw DOF Range"])
for i, node in enumerate(sorted(NODE_COEFFS), start=2):
  restraint = node_restraint(node)
  first = (node - 1) * DOF_PER_NODE + 1
  last = first + DOF_PER_NODE - 1
  vals = [node]
  for k, f in enumerate(restraint):
    vals.append("R" if f else first + k)
  vals.append(f"{first}-{last}")
  write_row(ws, i, vals)

summary_start = 2 + len(NODE_COEFFS) + 1
write_row(ws, summary_start, ["Total DOF:", TOTAL_DOF])
write_row(ws, summary_start + 1, ["Active (Free) DOF:", ACTIVE_DOF])
write_row(ws, summary_start + 2, ["Restrained DOF:", RESTRAINED_DOF])
for r in range(summary_start, summary_start + 3):
  ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
set_widths(ws, [8, 8, 8, 8, 8, 8, 8, 14])

# ---------------------------------------------------------------------------
# 7. DOF Numbering
# ---------------------------------------------------------------------------
ws = wb.create_sheet("DOF Numbering")
write_header(ws, ["Node", "Local DOF", "DOF", "Description", "Global DOF No.", "Status", "Equation No."])
row = 2
eq_counter = 0
for node in sorted(NODE_COEFFS):
  restraint = node_restraint(node)
  base = (node - 1) * DOF_PER_NODE
  for k, label in enumerate(DOF_LABELS):
    restrained = bool(restraint[k])
    global_dof = base + k + 1
    if restrained:
      eq = "-"
    else:
      eq_counter += 1
      eq = eq_counter
    write_row(ws, row, [node, k + 1, label, DOF_DESCRIPTIONS[label], global_dof,
                         "Restrained" if restrained else "Active", eq])
    row += 1
set_widths(ws, [8, 10, 8, 14, 15, 12, 13])

# ---------------------------------------------------------------------------
# 8. Member Releases  (Rev. 2: fully fixed / rigid at every end, all DOF)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Member Releases")
write_header(ws, ["Member", "End", "Fx", "Fy", "Fz", "Mx", "My", "Mz"])
row = 2
for m in MEMBERS:
  for end_label, node in (("i", m["i"]), ("j", m["j"])):
    write_row(ws, row, [m["id"], f"{end_label} (Node {node})"] + ["Fixed"] * 6)
    row += 1
set_widths(ws, [10, 16, 10, 10, 10, 10, 10, 10])

wb.save(XLSX)
print(f"Workbook written to {XLSX}")

# ---------------------------------------------------------------------------
# Recalculate (no formulas in this workbook, but keep the check for parity)
# ---------------------------------------------------------------------------
result = subprocess.run([sys.executable, RECALC, XLSX], capture_output=True, text=True)
print(result.stdout)
if result.returncode != 0:
  print(result.stderr)
  raise SystemExit("Recalculation failed - see error above.")

# ---------------------------------------------------------------------------
# Read back and plot the cube in 3D (Rev. 1 diagram style, Rev. 2 data:
# fully fixed / rigid member connections)
# ---------------------------------------------------------------------------
df_nodes = pd.read_excel(XLSX, sheet_name="Nodes")
nodes = {int(r["Node"]): (r["X (m)"], r["Y (m)"], r["Z (m)"]) for _, r in df_nodes.iterrows()}

fig = plt.figure(figsize=(11, 9))
ax = fig.add_subplot(111, projection="3d")

AXIS_COLOR = {"local_x": "tab:red", "local_y": "tab:green", "local_z": "mediumpurple"}
AXIS_LEN = 1.1

for m in MEMBERS:
  x1, y1, z1 = nodes[m["i"]]
  x2, y2, z2 = nodes[m["j"]]
  color = MEMBER_COLOR[m["type"]]
  ax.plot([x1, x2], [z1, z2], [y1, y2], color=color, linewidth=2.4, zorder=1)

  mx, my, mz = (x1 + x2) / 2, (y1 + y2) / 2, (z1 + z2) / 2
  for key in ("local_x", "local_y", "local_z"):
    dx, dy, dz = m[key]
    ax.quiver(mx, mz, my, dx, dz, dy, length=AXIS_LEN, color=AXIS_COLOR[key],
              linewidth=1.4, arrow_length_ratio=0.35, zorder=4)

  # small filled square = rigid (fixed) end connection marker at both ends
  for node in (m["i"], m["j"]):
    nx, ny, nz = nodes[node]
    dirx = (x2 - x1, y2 - y1, z2 - z1)
    norm = np.linalg.norm(dirx)
    if node == m["i"]:
      px, py, pz = nx + dirx[0] / norm * 0.35, ny + dirx[1] / norm * 0.35, nz + dirx[2] / norm * 0.35
    else:
      px, py, pz = nx - dirx[0] / norm * 0.35, ny - dirx[1] / norm * 0.35, nz - dirx[2] / norm * 0.35
    ax.scatter([px], [pz], [py], color="black", marker="s", s=18, zorder=5)

  if m["beta"] != 0:
    ax.text(mx, mz, my + 0.55, f"M{m['id']} (\u03b2={m['beta']:.0f}\u00b0)",
             fontsize=7.5, color="dimgray", zorder=5)
  else:
    ax.text(mx - 0.1, mz, my + 0.15, f"M{m['id']}", fontsize=7.5, color="dimgray", zorder=5)

support_xyz = np.array([[nodes[n][0], nodes[n][2], nodes[n][1]] for n in SUPPORT_NODES])
free_nodes = [n for n in nodes if n not in SUPPORT_NODES]
free_xyz = np.array([[nodes[n][0], nodes[n][2], nodes[n][1]] for n in free_nodes])

ax.scatter(*support_xyz.T, color="crimson", s=90, zorder=6, edgecolor="black", linewidth=0.6)
ax.scatter(*free_xyz.T, color="crimson", s=60, zorder=6, edgecolor="black", linewidth=0.4)

for n in SUPPORT_NODES:
  x, y, z = nodes[n]
  size = 0.45
  apex = np.array([x, z, y])
  base1 = np.array([x - size, z - size * 0.6, y - size])
  base2 = np.array([x + size, z - size * 0.6, y - size])
  base3 = np.array([x, z + size * 0.8, y - size])
  for b1, b2 in [(base1, base2), (base2, base3), (base3, base1)]:
    ax.plot([apex[0], b1[0]], [apex[1], b1[1]], [apex[2], b1[2]], color="black", linewidth=1.0, zorder=3)
    ax.plot([b1[0], b2[0]], [b1[1], b2[1]], [b1[2], b2[2]], color="black", linewidth=1.0, zorder=3)

ax.scatter([0], [0], [0], color="gold", marker="*", s=180, zorder=7, edgecolor="black", linewidth=0.5)

for n, (x, y, z) in nodes.items():
  first = (n - 1) * DOF_PER_NODE + 1
  last = first + DOF_PER_NODE - 1
  label = f"  N{n}\n  DOF {first}-{last}"
  ax.text(x, z, y + 0.35, label, fontsize=8, color="black", zorder=6)

GLOBAL_LEN = 1.6
ax.quiver(0, 0, 0, GLOBAL_LEN, 0, 0, color="black", linewidth=1.6, arrow_length_ratio=0.25, zorder=8)
ax.quiver(0, 0, 0, 0, GLOBAL_LEN, 0, color="black", linewidth=1.6, arrow_length_ratio=0.25, zorder=8)
ax.quiver(0, 0, 0, 0, 0, GLOBAL_LEN, color="black", linewidth=1.6, arrow_length_ratio=0.25, zorder=8)
ax.text(GLOBAL_LEN + 0.1, 0, 0, "X", fontsize=10, fontweight="bold")
ax.text(0, GLOBAL_LEN + 0.1, 0, "Z", fontsize=10, fontweight="bold")
ax.text(0, 0, GLOBAL_LEN + 0.1, "Y", fontsize=10, fontweight="bold")

ax.set_xlabel("X (m) - lateral")
ax.set_ylabel("Z (m) - lateral")
ax.set_zlabel("Y (m) - vertical")
ax.set_title(
    "6m x 6m x 6m Cube - Structural Model, Rev. 2\n"
    "Pinned supports at nodes 1-4, all member end connections fully fixed (rigid)",
    fontsize=12, fontweight="bold",
)

pad = 1.2
ax.set_xlim(-pad, EDGE_LENGTH + pad)
ax.set_ylim(-pad, EDGE_LENGTH + pad)
ax.set_zlim(-pad, EDGE_LENGTH + pad)
ax.set_box_aspect([1, 1, 1])
ax.view_init(elev=20, azim=-60)

legend_elems = [
    Line2D([0], [0], color="royalblue", lw=2, label="Beam (base / roof)"),
    Line2D([0], [0], color="seagreen", lw=2, label="Column"),
    Line2D([0], [0], marker="o", color="w", markerfacecolor="crimson", markeredgecolor="black",
           markersize=9, label="Supported node (pinned)"),
    Line2D([0], [0], marker="o", color="w", markerfacecolor="crimson", markeredgecolor="black",
           markersize=6, label="Free node"),
    Line2D([0], [0], marker="s", color="w", markerfacecolor="black", markeredgecolor="black",
           markersize=7, label="Fixed (rigid) member end"),
    Line2D([0], [0], marker="*", color="w", markerfacecolor="gold", markeredgecolor="black",
           markersize=14, label="Origin (0, 0, 0)"),
    Line2D([0], [0], color=AXIS_COLOR["local_x"], lw=2, label="Local x axis"),
    Line2D([0], [0], color=AXIS_COLOR["local_y"], lw=2, label="Local y axis"),
    Line2D([0], [0], color=AXIS_COLOR["local_z"], lw=2, label="Local z axis"),
]
ax.legend(handles=legend_elems, loc="upper left", fontsize=8, framealpha=0.9)

model_text = (
    "MODEL DATA - REV. 2\n\n"
    "Geometry\n"
    f"  Cube edge      {EDGE_LENGTH:.1f} m\n"
    f"  Nodes          {len(NODE_COEFFS)}\n"
    f"  Members        {len(MEMBERS)}\n"
    "  Vertical axis  global Y\n\n"
    "Supports\n"
    "  Type           pinned\n"
    "  Nodes          1, 2, 3, 4\n"
    "  Restrained     UX, UY, UZ\n"
    "  Released       RX, RY, RZ\n\n"
    "Degrees of freedom\n"
    f"  DOF per node   {DOF_PER_NODE}\n"
    f"  Total DOF      {TOTAL_DOF}\n"
    f"  Restrained DOF {RESTRAINED_DOF}\n"
    f"  Active DOF     {ACTIVE_DOF}\n\n"
    "Beta angles\n"
    "  Base/Roof beam 0 deg\n"
    "  Column         90 deg\n\n"
    "Member end connections\n"
    "  ALL members    Fixed (rigid)\n"
    "  Fx,Fy,Fz,Mx,My,Mz fixed at\n"
    "  both i- and j- ends\n\n"
    "Local axes\n"
    "  local x   start node i to end node j\n"
    "  local y   in the vertical plane, upward\n"
    "  local z   completes the right-handed set\n"
    "  Vertical members follow the limiting case:\n"
    "  local z parallel to global Z"
)
fig.text(0.70, 0.5, model_text, fontsize=7.6, family="monospace", va="center", ha="left",
          bbox=dict(boxstyle="round,pad=0.6", facecolor="aliceblue", edgecolor="steelblue"))

fig.subplots_adjust(left=0.02, right=0.68, top=0.92, bottom=0.05)
fig.savefig(PNG, dpi=150)
plt.close(fig)
print(f"Saved {PNG}")
