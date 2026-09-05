"""Lab 02: compute the reservoir analysis and create Alimane's submissions."""
from __future__ import annotations

import html
import json
from datetime import datetime
from pathlib import Path

import numpy as np
import pymupdf
from openpyxl import load_workbook
from scipy import stats
from scipy.integrate import quad
from scipy.optimize import curve_fit


STUDENT = "Alimane"
SECTION = "BES6-M"
HERE = Path(__file__).resolve().parent
DATA = HERE / "Data 01.xlsx"
HTML_FILE = HERE / "lab02_alimane.html"
PDF_FILE = HERE / "lab02_alimane_results.pdf"
PROMPTS_FILE = HERE / "lab02_alimane_prompts.pdf"


def logistic(t, a, k, mid):
    return a / (1 + np.exp(np.clip(-k * (t - mid), -700, 700)))


def model(t, c, a1, k1, t1, a2, k2, t2):
    return c + logistic(t, a1, k1, t1) + logistic(t, a2, k2, t2)


def load_data():
    sheet = load_workbook(DATA, data_only=True, read_only=True)["Sensor Log"]
    rows = list(sheet.iter_rows(min_row=5, values_only=True))
    dates = [r[1] for r in rows]
    h = np.array([r[4] for r in rows], dtype=float)
    t = np.array([(d - dates[0]).total_seconds() / 3600 for d in dates])
    return dates, t, h


def line_svg(series, ylabel, zero=False, marker=None, shade=False):
    W, H, L, R, T, B = 960, 325, 65, 20, 25, 45
    x = np.concatenate([s[0] for s in series]); y = np.concatenate([s[1] for s in series])
    if zero: y = np.append(y, 0)
    ymin, ymax = y.min(), y.max(); pad = max(.03, (ymax-ymin)*.12); ymin -= pad; ymax += pad
    xmin, xmax = x.min(), x.max()
    px = lambda v: L+(v-xmin)/(xmax-xmin)*(W-L-R)
    py = lambda v: T+(ymax-v)/(ymax-ymin)*(H-T-B)
    out = [f'<svg viewBox="0 0 {W} {H}" aria-label="{ylabel} graph">']
    for v in np.linspace(ymin,ymax,6):
        out.append(f'<line class="grid" x1="{L}" x2="{W-R}" y1="{py(v):.1f}" y2="{py(v):.1f}"/><text class="tick" x="{L-7}" y="{py(v)+4:.1f}" text-anchor="end">{v:.2f}</text>')
    for v in np.linspace(xmin,xmax,7):
        out.append(f'<line class="grid" x1="{px(v):.1f}" x2="{px(v):.1f}" y1="{T}" y2="{H-B}"/><text class="tick" x="{px(v):.1f}" y="{H-23}" text-anchor="middle">{v:.0f}</text>')
    if zero and ymin < 0 < ymax: out.append(f'<line class="zero" x1="{L}" x2="{W-R}" y1="{py(0):.1f}" y2="{py(0):.1f}"/>')
    if shade:
        sx, sy = series[-1][0], series[-1][1]
        pts = ' '.join(f'{px(a):.1f},{py(b):.1f}' for a,b in zip(sx,sy))
        out.append(f'<path class="area" d="M {px(sx[0]):.1f},{py(0):.1f} L {pts} L {px(sx[-1]):.1f},{py(0):.1f} Z"/>')
    for sx,sy,color,dots in series:
        pts = ' '.join(f'{px(a):.1f},{py(b):.1f}' for a,b in zip(sx,sy))
        out.append(f'<polyline class="line" style="stroke:{color}" points="{pts}"/>')
        if dots: out.extend(f'<circle fill="{color}" cx="{px(a):.1f}" cy="{py(b):.1f}" r="1.5"/>' for a,b in zip(sx,sy))
    if marker:
        a,b,label = marker
        out.append(f'<line class="mark" x1="{px(a):.1f}" x2="{px(a):.1f}" y1="{T}" y2="{H-B}"/><circle class="dot" cx="{px(a):.1f}" cy="{py(b):.1f}" r="4"/><text class="label" x="{px(a)+8:.1f}" y="{py(b)-9:.1f}">{label}</text>')
    out.append(f'<line class="axis" x1="{L}" x2="{W-R}" y1="{H-B}" y2="{H-B}"/><line class="axis" x1="{L}" x2="{L}" y1="{T}" y2="{H-B}"/><text class="tick" x="{W/2}" y="{H-5}" text-anchor="middle">hours from first reading</text><text class="tick" transform="translate(15 {H/2}) rotate(-90)" text-anchor="middle">{ylabel}</text></svg>')
    return ''.join(out)


def make_pdf(names, values, residual_text):
    doc = pymupdf.open(); page = doc.new_page(width=612, height=792)
    page.draw_rect((0,0,612,792), color=None, fill=(.03,.08,.14))
    page.insert_text((42,55), "RESERVOIR RESPONSE STUDY", fontsize=21, fontname="hebo", color=(.49,.83,.99))
    page.insert_text((42,80), f"Numerical Methods · Lab 02 · {STUDENT} · {SECTION}", fontsize=10, color=(.75,.83,.9))
    page.draw_line((42,95),(570,95), color=(.16,.32,.45), width=1)
    page.insert_text((42,123), "FITTED PARAMETERS — TWO LOGISTIC COMPONENTS", fontsize=11, fontname="hebo", color=(.98,.75,.14))
    y = 148
    page.insert_text((42,y), "Parameter", fontsize=9, fontname="hebo", color=(.7,.8,.88)); page.insert_text((205,y), "Estimate", fontsize=9, fontname="hebo", color=(.7,.8,.88)); page.insert_text((320,y), "SE", fontsize=9, fontname="hebo", color=(.7,.8,.88)); page.insert_text((410,y), "t", fontsize=9, fontname="hebo", color=(.7,.8,.88)); page.insert_text((485,y), "p-value", fontsize=9, fontname="hebo", color=(.7,.8,.88))
    for name, val, se, ts, pv in zip(names, values['popt'], values['se'], values['ts'], values['pv']):
        y += 25; page.draw_line((42,y+6),(570,y+6), color=(.12,.22,.32), width=.4)
        page.insert_text((42,y), name, fontsize=10, color=(.91,.95,.98)); page.insert_text((205,y), f"{val:.5f}", fontsize=10, fontname="cour", color=(.91,.95,.98)); page.insert_text((320,y), f"{se:.5f}", fontsize=10, fontname="cour", color=(.91,.95,.98)); page.insert_text((410,y), f"{ts:.2f}", fontsize=10, fontname="cour", color=(.91,.95,.98)); page.insert_text((485,y), f"{pv:.3g}", fontsize=10, fontname="cour", color=(.91,.95,.98))
    y += 54; page.insert_text((42,y), "MODEL AND INTEGRATION SUMMARY", fontsize=11, fontname="hebo", color=(.98,.75,.14))
    summary = [("SSE",f"{values['sse']:.4f}"),("R²",f"{values['r2']:.6f}"),("Standard error",f"{values['s']:.4f} m"),("Degrees of freedom",str(values['dof'])),("Fitted area",f"{values['area']:.4f} m·h"),("quad absolute error",f"{values['err']:.2e}"),("Raw trapezoid check",f"{values['trap']:.4f} m·h"),("Peak fitted dh/dt",f"{values['rate']:.4f} m/h at {values['rate_t']:.2f} h")]
    y += 26
    for i,(key,val) in enumerate(summary):
        x = 42 if i%2==0 else 320
        if i%2==0 and i: y += 35
        page.insert_text((x,y), key.upper(), fontsize=7, fontname="hebo", color=(.55,.7,.8)); page.insert_text((x,y+15), val, fontsize=12, fontname="hebo", color=(.91,.95,.98))
    y += 65; page.insert_text((42,y), "RESIDUAL READING", fontsize=11, fontname="hebo", color=(.98,.75,.14))
    rect = pymupdf.Rect(42,y+12,570,y+99); page.draw_rect(rect, color=(.12,.32,.34), fill=(.04,.16,.18))
    page.insert_textbox(rect + (12,12,-12,-10), residual_text, fontsize=10, lineheight=1.35, color=(.84,.96,.94))
    page.insert_text((42,755), "Generated by lab02_alimane.py — all figures computed in Python.", fontsize=8, color=(.55,.67,.76))
    doc.save(PDF_FILE); doc.close()


def make_prompts_pdf():
    """Create the formatted prompt documentation required for the submission."""
    doc = pymupdf.open(); page = doc.new_page(width=612, height=792)
    page.draw_rect((0, 0, 612, 792), color=None, fill=(.03, .08, .14))
    page.insert_text((42, 54), "LAB 02 - PROMPT DOCUMENTATION", fontsize=21, fontname="hebo", color=(.49, .83, .99))
    page.insert_text((42, 78), f"Student: {STUDENT} · Numerical Methods · {SECTION}", fontsize=10, color=(.75, .83, .9))
    page.draw_line((42, 94), (570, 94), color=(.16, .32, .45), width=1)
    intro = "Prompt history used to develop the Numerical Methods Laboratory Activity 02 submission."
    page.insert_textbox(pymupdf.Rect(42, 112, 570, 140), intro, fontsize=10, lineheight=1.35, color=(.85, .93, .97))
    entries = [
        ("PROMPT 1  Initial Setup & Guidelines", "I have attached three files for reference: NM-Lab02-Activity-Set-Escranda.pdf, Data 01.xlsx, and lab02_escranda.html. Please distinguish the instructions in the activity set from my personal requests. Using Data 01.xlsx, build a baseline version of the lab assignment. Note that lab02_escranda.html is provided strictly as a reference example for requirements; do not duplicate its exact design or visual layout."),
        ("PROMPT 2  Personalization & Output Formatting", "Update the submission details to use the surname Alimane. Additionally, generate the corresponding results summary as a one-page PDF document."),
        ("PROMPT 3  Deliverable Regeneration", "Re-run and regenerate all deliverables (the HTML dashboard and results document) keeping the core analysis intact, but ensure all branding, headers, and metadata consistently reflect the name Alimane."),
        ("PROMPT 4  Prompt Documentation", "Generate a clean prompt record file documenting the chat history used to build this lab submission. Ensure the recorded requests are polished, professional, and properly formatted."),
    ]
    y = 157
    for heading, body in entries:
        page.insert_text((42, y), heading.upper(), fontsize=10, fontname="hebo", color=(.98, .75, .14)); y += 12
        height = 92 if heading.startswith("PROMPT 1") else 62
        box = pymupdf.Rect(42, y, 570, y + height)
        page.draw_rect(box, color=(.12, .25, .36), fill=(.04, .13, .22))
        page.insert_textbox(box + (10, 9, -10, -8), body, fontsize=9.5, lineheight=1.3, color=(.9, .95, .98)); y += 82
        y += 30 if heading.startswith("PROMPT 1") else 0
    page.insert_text((42, 755), "Laboratory Activity 02 - Alimane", fontsize=8, color=(.55, .67, .76))
    # Save to a temporary file first. Writing its bytes into the existing file
    # avoids a Windows replace operation when the PDF is open for viewing.
    temporary = PROMPTS_FILE.with_suffix(".new.pdf")
    doc.save(temporary); doc.close()
    PROMPTS_FILE.write_bytes(temporary.read_bytes())
    temporary.unlink()


def main():
    dates,t,h = load_data(); d1=np.gradient(h,t,edge_order=2); d2=np.gradient(d1,t,edge_order=2)
    p0=[h.min(),h.max()-h.min(),.50,31,-3.7,.28,38]
    popt,pcov=curve_fit(model,t,h,p0=p0,method='lm',maxfev=50000)
    fit=model(t,*popt); resid=h-fit; n=len(h); p=len(popt); dof=n-p; sse=float(np.sum(resid**2)); sst=float(np.sum((h-h.mean())**2)); r2=1-sse/sst; s=float(np.sqrt(sse/dof)); se=np.sqrt(np.diag(pcov)); ts=popt/se; pv=2*stats.t.sf(np.abs(ts),dof)
    dense=np.linspace(t[0],t[-1],1600); smooth=model(dense,*popt)
    rate=sum(a*k*(logistic(dense,1,k,mid)*(1-logistic(dense,1,k,mid))) for a,k,mid in ((popt[1],popt[2],popt[3]),(popt[4],popt[5],popt[6])))
    ix=int(np.argmax(rate)); area,err=quad(lambda x:float(model(x,*popt)),float(t[0]),float(t[-1])); trap=float(np.trapezoid(h,t)); runs=1+np.count_nonzero((resid[1:]>=0)!=(resid[:-1]>=0))
    rawpeak=int(np.argmax(d1)); maxd2=int(np.argmax(d2)); mind2=int(np.argmin(d2))
    stage=line_svg([(t,h,'#67d7f4',True)],'stage (m)')
    deriv=line_svg([(t,d1,'#fb7185',False),(t,d2,'#a78bfa',False)],'derivative',True,(t[rawpeak],d1[rawpeak],'raw max dh/dt'))
    fitted=line_svg([(t,h,'#91a5b7',True),(dense,smooth,'#fbbf24',False)],'stage (m)',False,(dense[ix],smooth[ix],'fastest fitted fill'))
    residual=line_svg([(t,resid,'#36d399',True)],'residual (m)',True)
    integral=line_svg([(dense,smooth,'#fbbf24',False)],'fitted stage (m)',False,None,True)
    param_names=['c — baseline','a₁ — rise amplitude','k₁ — rise rate','t₁ — rise midpoint','a₂ — recession amplitude','k₂ — recession rate','t₂ — recession midpoint']
    units=['m','m','1/h','h','m','1/h','h']
    rows=''.join(f'<tr><th>{a}</th><td>{b:.5f}</td><td>{c:.5f}</td><td>{d:.2f}</td><td>{e:.3g}</td><td>{u}</td></tr>' for a,b,c,d,e,u in zip(param_names,popt,se,ts,pv,units))
    data={'student':STUDENT,'readings':n,'parameters':popt.tolist(),'parameter_se':se.tolist(),'t_statistics':ts.tolist(),'p_values':pv.tolist(),'sse':sse,'sst':sst,'r2':r2,'standard_error':s,'area':area,'quad_error':err,'trapezoid':trap,'peak_rate':float(rate[ix]),'peak_time':float(dense[ix])}
    residual_text=f"The residuals have {runs} sign runs across {n} readings. Their maximum absolute value is {np.max(np.abs(resid)):.3f} m, which is {np.max(np.abs(resid))/.01:.1f} times the logger resolution. The high R² is useful, but sustained signs show that the simple model does not reproduce every local feature."
    page=f'''<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Lab 02 | Alimane</title><style>
:root{{--bg:#071525;--panel:#102640;--edge:#294863;--ink:#eaf4fb;--muted:#9bb2c5;--cyan:#67d7f4;--gold:#fbbf24;--mint:#36d399}}*{{box-sizing:border-box}}body{{margin:0;background:linear-gradient(145deg,#06101d,#12344b);color:var(--ink);font:16px/1.55 system-ui,sans-serif}}main{{max-width:1200px;margin:auto;padding:32px 22px 70px}}.hero,.panel{{background:rgba(8,25,44,.95);border:1px solid var(--edge);border-radius:19px;padding:25px;margin-bottom:18px}}.hero{{background:radial-gradient(circle at 92% 0,#245d71,transparent 35%),#091a2d}}.tag{{color:var(--cyan);font-size:11px;font-weight:800;letter-spacing:.15em;text-transform:uppercase}}h1{{font-size:clamp(31px,5vw,52px);margin:7px 0}}h2{{margin:0}}.lede{{color:var(--muted)}}.facts,.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:11px;margin-top:18px}}.fact,.card{{border:1px solid var(--edge);background:#0b1d31;border-radius:11px;padding:12px}}.fact span,.card span{{display:block;color:var(--muted);font-size:10px;text-transform:uppercase;letter-spacing:.1em}}.card b{{display:block;color:var(--gold);font-size:21px}}.tabs{{display:flex;gap:7px;flex-wrap:wrap}}button{{cursor:pointer;background:#102640;color:var(--ink);border:1px solid var(--edge);border-radius:9px 9px 0 0;padding:11px 15px;font-weight:700}}button[aria-selected=true]{{background:var(--cyan);border-color:var(--cyan);color:#071525}}.view{{display:none}}.view.active{{display:block}}svg{{display:block;width:100%;height:auto;background:#061525;border-radius:11px}}.grid{{stroke:#23415a;stroke-width:1}}.axis{{stroke:#9ab1c2}}.zero{{stroke:#fb7185;stroke-dasharray:7 5}}.line{{fill:none;stroke-width:2.4;stroke-linejoin:round;stroke-linecap:round}}.area{{fill:#fbbf24;opacity:.18}}.mark{{stroke:#fbbf24;stroke-dasharray:4 5}}.dot{{fill:#fbbf24;stroke:#071525;stroke-width:2}}.tick,.label{{fill:#9bb2c5;font:11px system-ui,sans-serif}}.label{{fill:#fbbf24;font-weight:700}}.callout{{margin-top:16px;border-left:4px solid var(--mint);background:#0a2a31;padding:13px 15px;border-radius:0 8px 8px 0}}.equation{{font:16px ui-monospace,monospace;background:#061525;padding:13px;border-radius:9px;overflow:auto}}table{{width:100%;border-collapse:collapse;font-size:14px}}th,td{{padding:8px;text-align:right;border-bottom:1px solid var(--edge)}}th{{text-align:left}}thead th{{font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:var(--cyan)}}footer{{padding:15px;color:var(--muted);font-size:12px}}@media(max-width:650px){{main{{padding:15px 10px}}.hero,.panel{{padding:17px}}table{{font-size:12px}}th,td{{padding:6px}}}}
</style></head><body><main><header class="hero"><div class="tag">Numerical Methods · Laboratory Activity 02</div><h1>Reservoir response study</h1><p class="lede">Stage behaviour, fitting evidence, and time-integrated level from the provided logger data.</p><div class="facts"><div class="fact"><span>Student</span><b>{STUDENT}</b></div><div class="fact"><span>Section</span><b>{SECTION}</b></div><div class="fact"><span>Dataset</span><b>{DATA.name}</b></div><div class="fact"><span>Record</span><b>{n} readings · {t[-1]:.2f} h</b></div><div class="fact"><span>Window</span><b>{dates[0]:%d %b %Y}–{dates[-1]:%d %b}</b></div></div></header><section class="panel"><h2>Observed stage log</h2><p class="lede">The raw measurement series is permanently visible.</p>{stage}</section><nav class="tabs"><button data-tab="d" aria-selected="true">1 · Derivatives</button><button data-tab="f" aria-selected="false">2 · Fit & residuals</button><button data-tab="a" aria-selected="false">3 · Area under curve</button></nav><section id="d" class="view active"><div class="panel"><h2>Finite differences</h2><p class="lede">Forward/backward differences are used at the ends; central differences are used in the interior.</p>{deriv}<div class="cards"><div class="card"><span>Maximum raw dh/dt</span><b>{d1[rawpeak]:.4f} m/h</b><small>at {t[rawpeak]:.2f} h</small></div><div class="card"><span>Largest acceleration</span><b>{d2[maxd2]:.4f}</b><small>at {t[maxd2]:.2f} h</small></div><div class="card"><span>Largest deceleration</span><b>{d2[mind2]:.4f}</b><small>at {t[mind2]:.2f} h</small></div></div><div class="callout"><b>Reading.</b> The second derivative changes sign around the steep rise: inflow speeds up to the filling maximum, then slows down as the stage approaches the crest.</div></div></section><section id="f" class="view"><div class="panel"><h2>Fitted two-component logistic model</h2><p class="lede">A positive logistic describes the rapid filling event and a negative logistic describes the broader recession. Starting values were taken from the observed baseline, range, steep rise, and later decline.</p><div class="equation">h(t) = c + a₁/(1 + e^(−k₁(t−t₁))) + a₂/(1 + e^(−k₂(t−t₂)))</div>{fitted}<div class="cards"><div class="card"><span>Fitted maximum dh/dt</span><b>{rate[ix]:.4f} m/h</b><small>at {dense[ix]:.2f} h</small></div><div class="card"><span>SSE / R²</span><b>{sse:.4f} / {r2:.6f}</b><small>SST = {sst:.4f}</small></div><div class="card"><span>Standard error</span><b>{s:.4f} m</b><small>{dof} degrees of freedom</small></div></div><h3>Parameter estimates</h3><table><thead><tr><th>Parameter</th><th>Estimate</th><th>SE</th><th>t</th><th>Two-tailed p</th><th>Unit</th></tr></thead><tbody>{rows}</tbody></table></div><div class="panel"><h2>Residual evidence</h2><p class="lede">Observed stage minus fitted stage. The pink line is zero.</p>{residual}<div class="callout">{residual_text}</div></div></section><section id="a" class="view"><div class="panel"><h2>Integrated fitted stage</h2><p class="lede">The shaded region is the fitted curve integrated from the first to the last recorded time.</p>{integral}<div class="cards"><div class="card"><span>quad integral</span><b>{area:.4f} m·h</b><small>absolute error {err:.2e}</small></div><div class="card"><span>Raw trapezoid check</span><b>{trap:.4f} m·h</b><small>difference {area-trap:.4f} m·h</small></div></div><div class="callout"><b>Meaning.</b> Meter-hours summarize how high the reservoir remained and for how long. It is not a volume. The trapezoid check differs because it connects rounded raw readings, while quad integrates the smooth fitted function.</div></div></section><footer>Generated {datetime.now():%d %B %Y, %H:%M} by <code>lab02_alimane.py</code>. All numerical work is done in Python; JavaScript only changes tabs.</footer><script id="results" type="application/json">{html.escape(json.dumps(data),quote=False)}</script><script>document.querySelectorAll('button').forEach(b=>b.onclick=()=>{{document.querySelectorAll('button').forEach(x=>x.setAttribute('aria-selected','false'));document.querySelectorAll('.view').forEach(x=>x.classList.remove('active'));b.setAttribute('aria-selected','true');document.getElementById(b.dataset.tab).classList.add('active')}})</script></main></body></html>'''
    HTML_FILE.write_text(page,encoding='utf-8')
    make_pdf(param_names,{'popt':popt,'se':se,'ts':ts,'pv':pv,'sse':sse,'r2':r2,'s':s,'dof':dof,'area':area,'err':err,'trap':trap,'rate':rate[ix],'rate_t':dense[ix]},residual_text)
    make_prompts_pdf()
    print(f'Created {HTML_FILE.name}, {PDF_FILE.name}, and {PROMPTS_FILE.name}.')


if __name__ == '__main__': main()
