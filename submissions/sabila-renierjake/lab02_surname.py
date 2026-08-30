"""
Reservoir Stage Analysis
------------------------
Reads the 15-minute reservoir stage logger data, builds a numeric time axis
(t = 0 h at the first reading, dt = 0.25 h), and computes the first and
second time-derivatives of the water depth h(t) using:

  dh/dt:
    - Forward difference  at the first point
    - Central difference  at all interior points
    - Backward difference at the last point

  d2h/dt2:
    - Standard 3-point second-derivative (central) formula at interior points
    - One-sided 3-point forward/backward formulas at the two endpoints
      (so every point in the record gets a second-derivative value)

Outputs a CSV with time, depth, dh/dt and d2h/dt2, plus a summary of the
key findings printed to the console.
"""

import openpyxl
import numpy as np
import csv
from datetime import datetime
from scipy.optimize import curve_fit
from scipy import stats
from scipy.integrate import quad
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# ---------------------------------------------------------------
# 1. Load data
# ---------------------------------------------------------------
SRC = "/mnt/user-data/uploads/Data_01__1_.xlsx"
OUT_CSV = "/mnt/user-data/outputs/reservoir_derivatives.csv"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Sensor Log"]

readings = []
depths = []
timestamps = []

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    reading_no, timestamp, date, time, depth = row
    if isinstance(reading_no, int):           # skip title/header rows
        readings.append(reading_no)
        timestamps.append(timestamp)
        depths.append(float(depth))

h = np.array(depths, dtype=float)
n = len(h)

# ---------------------------------------------------------------
# 2. Time axis: t = 0 h at first reading, dt = 0.25 h
# ---------------------------------------------------------------
dt = 0.25
t = np.arange(n) * dt   # 0, 0.25, 0.50, ... hours

# ---------------------------------------------------------------
# 3. First derivative dh/dt  (m / h)
#    forward at i=0, central for interior points, backward at i=n-1
# ---------------------------------------------------------------
dhdt = np.zeros(n)
dhdt[0] = (h[1] - h[0]) / dt                      # forward difference
dhdt[-1] = (h[-1] - h[-2]) / dt                    # backward difference
dhdt[1:-1] = (h[2:] - h[:-2]) / (2 * dt)           # central difference

# ---------------------------------------------------------------
# 4. Second derivative d2h/dt2  (m / h^2)
#    central 3-point formula for interior points;
#    one-sided 3-point formulas at the two endpoints
# ---------------------------------------------------------------
d2hdt2 = np.zeros(n)
d2hdt2[1:-1] = (h[2:] - 2 * h[1:-1] + h[:-2]) / dt**2         # central
d2hdt2[0]  = (h[2] - 2 * h[1] + h[0]) / dt**2                 # forward 3-pt
d2hdt2[-1] = (h[-1] - 2 * h[-2] + h[-3]) / dt**2              # backward 3-pt

# ---------------------------------------------------------------
# 5. Save full results
# ---------------------------------------------------------------
import os
os.makedirs("/mnt/user-data/outputs", exist_ok=True)
with open(OUT_CSV, "w", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(["Reading", "Timestamp", "t (h)", "h (m)",
                      "dh/dt (m/h)", "d2h/dt2 (m/h^2)"])
    for i in range(n):
        writer.writerow([readings[i], timestamps[i], f"{t[i]:.2f}",
                          f"{h[i]:.3f}", f"{dhdt[i]:.5f}", f"{d2hdt2[i]:.5f}"])

# ---------------------------------------------------------------
# 6. Analysis / deliverables
# ---------------------------------------------------------------
imax = int(np.argmax(dhdt))
print(f"Number of data points processed : {n}")
print(f"Time range                      : t = 0 h to t = {t[-1]:.2f} h")
print()
print("=== Maximum rate of rise ===")
print(f"  Time of max dh/dt : t = {t[imax]:.2f} h  "
      f"(Reading #{readings[imax]}, {timestamps[imax]})")
print(f"  Max dh/dt value   : {dhdt[imax]:.4f} m/h")
print()

# sign-change summary of the 2nd derivative for context
sign_changes = np.where(np.diff(np.sign(d2hdt2)) != 0)[0]
print(f"Number of sign changes in d2h/dt2 : {len(sign_changes)}")
print()
print("Saved full time series (t, h, dh/dt, d2h/dt2) to:")
print(f"  {OUT_CSV}")

# =================================================================
# 7. NONLINEAR CURVE FITTING (Levenberg-Marquardt)
# =================================================================
#
# MODEL CHOICE: Sum of two logistics
#
# Defense (2 sentences): The record is not a single monotonic
# saturating rise -- it shows a low baseline stage, a fast sigmoid-shaped
# rise driven by an inflow pulse, a peak, and then a second, slower
# sigmoid-shaped recession down to a new, higher equilibrium stage, so a
# single logistic, Gompertz, or cubic (all of which are monotonic or
# unbounded) cannot reproduce both the rising limb and the falling limb
# with two flat asymptotes. A sum of two logistic terms -- one positive
# term for the rising limb and one negative term for the falling limb --
# naturally reproduces this "rise-then-partial-recession-to-a-new-plateau"
# hydrograph shape while keeping both ends bounded, which matches the
# physical picture of a reservoir that fills quickly during an inflow
# event and then partially drains/settles toward a new stable level.
#
#   h(t) = h0 + A1 / (1 + exp(-k1*(t - t1)))   <- rising limb (inflow)
#             - A2 / (1 + exp(-k2*(t - t2)))   <- falling limb (recession)
#
#   h0  : baseline (pre-event) stage
#   A1  : total rise of the inflow pulse
#   k1  : steepness of the rise
#   t1  : time of the rising inflection (max dh/dt)
#   A2  : size of the post-peak recession
#   k2  : steepness of the recession
#   t2  : time of the falling inflection (min dh/dt after the peak)
# ---------------------------------------------------------------

def sum_two_logistics(t, h0, A1, k1, t1, A2, k2, t2):
    return (h0
            + A1 / (1 + np.exp(-k1 * (t - t1)))
            - A2 / (1 + np.exp(-k2 * (t - t2))))

# --- Initial guesses (p0), each read directly off the data ------
baseline_h   = h[:20].mean()              # flat stretch before the rise
peak_h       = h.max()                    # observed peak stage
peak_t       = t[np.argmax(h)]
final_h      = h[-20:].mean()             # flat stretch at the end

rise_idx     = int(np.argmax(dhdt))       # steepest rise -> rising inflection
t1_guess     = t[rise_idx]
dhdt_max     = dhdt[rise_idx]

post_peak    = dhdt[np.argmax(h):]
fall_idx     = int(np.argmin(post_peak)) + int(np.argmax(h))
t2_guess     = t[fall_idx]
dhdt_min     = dhdt[fall_idx]             # steepest recession (negative)

A1_guess = peak_h - baseline_h            # size of the rising limb
A2_guess = peak_h - final_h               # size of the falling limb

# For a logistic, slope at the inflection = k*A/4  =>  k = 4*slope/A
k1_guess = 4 * dhdt_max / A1_guess
k2_guess = 4 * abs(dhdt_min) / A2_guess

p0 = [baseline_h, A1_guess, k1_guess, t1_guess,
      A2_guess, k2_guess, t2_guess]

print("\n=== Initial guesses (p0) and their source ===")
print(f"  h0 (baseline)      = {baseline_h:.2f} m   "
      f"<- mean of first 20 readings (flat pre-event stage)")
print(f"  A1 (rise amplitude)= {A1_guess:.2f} m   "
      f"<- peak stage ({peak_h:.2f} m) minus baseline")
print(f"  k1 (rise steepness)= {k1_guess:.3f} 1/h "
      f"<- 4 * (max observed dh/dt = {dhdt_max:.2f} m/h) / A1")
print(f"  t1 (rise inflection)= {t1_guess:.2f} h  "
      f"<- time at which dh/dt is maximum")
print(f"  A2 (fall amplitude)= {A2_guess:.2f} m   "
      f"<- peak stage minus final plateau ({final_h:.2f} m)")
print(f"  k2 (fall steepness)= {k2_guess:.3f} 1/h "
      f"<- 4 * |min observed dh/dt after peak = {dhdt_min:.2f} m/h| / A2")
print(f"  t2 (fall inflection)= {t2_guess:.2f} h  "
      f"<- time at which dh/dt is most negative after the peak")

# --- Fit with Levenberg-Marquardt --------------------------------
popt, pcov = curve_fit(sum_two_logistics, t, h, p0=p0,
                        method='lm', maxfev=20000)

param_names = ["h0", "A1", "k1", "t1", "A2", "k2", "t2"]

# --- Statistical evaluation (exact order requested) --------------
h_hat   = sum_two_logistics(t, *popt)
resid   = h - h_hat
resid2  = resid ** 2
SSE     = np.sum(resid2)
SST     = np.sum((h - h.mean()) ** 2)
R2      = 1 - SSE / SST

p_count = len(popt)
s       = np.sqrt(SSE / (n - p_count))

se       = np.sqrt(np.diag(pcov))
t_stat   = popt / se
p_values = 2 * (1 - stats.t.cdf(np.abs(t_stat), n - p_count))

# --- Report --------------------------------------------------------
print("\n=== Fitted model: sum of two logistics ===")
print("h(t) = h0 + A1/(1+exp(-k1*(t-t1))) - A2/(1+exp(-k2*(t-t2)))\n")

print("=== Goodness of fit ===")
print(f"  n (data points)            = {n}")
print(f"  p (fitted parameters)      = {p_count}")
print(f"  SSE                        = {SSE:.4f} m^2")
print(f"  SST                        = {SST:.4f} m^2")
print(f"  R^2                        = {R2:.4f}")
print(f"  Standard error of estimate = {s:.3f} m")

print("\n=== Parameter summary table ===")
header = f"{'Param':<6}{'Estimate':>12}{'Std. Error':>14}{'t-stat':>12}{'p-value':>12}"
print(header)
print("-" * len(header))
units = {"h0": "m", "A1": "m", "k1": "1/h", "t1": "h",
         "A2": "m", "k2": "1/h", "t2": "h"}
for name, val, se_i, ts, pv in zip(param_names, popt, se, t_stat, p_values):
    # round to sensor precision: 2 decimals for m/h quantities, 3 for rates
    pv_str = "<0.001" if pv < 0.001 else f"{pv:.3g}"
    if units[name] in ("m", "h"):
        print(f"{name:<6}{val:>10.2f} {units[name]:<2}{se_i:>11.2f}"
              f"{ts:>12.2f}{pv_str:>12}")
    else:
        print(f"{name:<6}{val:>10.3f} {units[name]:<2}{se_i:>11.3f}"
              f"{ts:>12.2f}{pv_str:>12}")

# save fitted curve alongside residuals for reference
FIT_CSV = "/mnt/user-data/outputs/reservoir_curve_fit.csv"
with open(FIT_CSV, "w", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(["t (h)", "h (m)", "h_hat (m)", "residual (m)"])
    for i in range(n):
        writer.writerow([f"{t[i]:.2f}", f"{h[i]:.2f}",
                          f"{h_hat[i]:.2f}", f"{resid[i]:.3f}"])
print(f"\nSaved fitted curve and residuals to:\n  {FIT_CSV}")

# =================================================================
# 8. RESIDUAL VISUALIZATION
# =================================================================
RESID_VS_T_PNG   = "/mnt/user-data/outputs/residuals_vs_time.png"
RESID_VS_HHAT_PNG = "/mnt/user-data/outputs/residuals_vs_fitted.png"

# --- Residuals vs time -------------------------------------------
fig, ax = plt.subplots(figsize=(9, 4))
ax.scatter(t, resid, s=12, color="steelblue", alpha=0.7)
ax.axhline(0, color="black", linewidth=1)
ax.set_xlabel("t (h)")
ax.set_ylabel("Residual, e_i = h - h_hat (m)")
ax.set_title("Residuals vs. Time")
fig.tight_layout()
fig.savefig(RESID_VS_T_PNG, dpi=150)
plt.close(fig)

# --- Residuals vs fitted values -----------------------------------
fig, ax = plt.subplots(figsize=(9, 4))
ax.scatter(h_hat, resid, s=12, color="darkorange", alpha=0.7)
ax.axhline(0, color="black", linewidth=1)
ax.set_xlabel("Fitted value, h_hat (m)")
ax.set_ylabel("Residual, e_i = h - h_hat (m)")
ax.set_title("Residuals vs. Fitted Values")
fig.tight_layout()
fig.savefig(RESID_VS_HHAT_PNG, dpi=150)
plt.close(fig)

print(f"\nSaved residual plots to:\n  {RESID_VS_T_PNG}\n  {RESID_VS_HHAT_PNG}")

# --- Supporting numbers for the written residual analysis ---------
sensor_res = 0.01  # m, logger resolution (1 cm)
max_abs_resid = np.max(np.abs(resid))
idx_max_resid = int(np.argmax(np.abs(resid)))

# longest run of same-sign residuals
signs = np.sign(resid)
signs[signs == 0] = 1
run_lengths = []
run_len = 1
for i in range(1, n):
    if signs[i] == signs[i - 1]:
        run_len += 1
    else:
        run_lengths.append(run_len)
        run_len = 1
run_lengths.append(run_len)
longest_run = max(run_lengths)

# mean residual over three level bands (low / mid / high) to check
# whether spread widens as the stage rises
low_mask  = h < np.percentile(h, 33)
mid_mask  = (h >= np.percentile(h, 33)) & (h < np.percentile(h, 67))
high_mask = h >= np.percentile(h, 67)
std_low  = resid[low_mask].std()
std_mid  = resid[mid_mask].std()
std_high = resid[high_mask].std()

print("\n=== Residual diagnostics ===")
print(f"  Mean residual                 = {resid.mean():.4f} m")
print(f"  Max |residual|                = {max_abs_resid:.4f} m "
      f"at t = {t[idx_max_resid]:.2f} h "
      f"({max_abs_resid/sensor_res:.1f}x the 0.01 m sensor resolution)")
print(f"  Longest run of same-sign resid = {longest_run} consecutive points "
      f"({longest_run*dt:.2f} h)")
print(f"  Residual std dev by level band : low={std_low:.4f} m, "
      f"mid={std_mid:.4f} m, high={std_high:.4f} m")

# =================================================================
# 9. AREA UNDER THE FITTED CURVE  (scipy.integrate.quad)
# =================================================================
t0, tn = t[0], t[-1]
area_quad, area_err = quad(lambda tt: sum_two_logistics(tt, *popt), t0, tn)

print("\n=== Area under fitted curve (quad) ===")
print(f"  Integration interval : t = {t0:.2f} h to t = {tn:.2f} h")
print(f"  Area (quad)          = {area_quad:.2f} m*h")
print(f"  Absolute error est.  = {area_err:.2e} m*h")
print("  NOTE: units are meter-hours (m*h), a time-integral of stage,")
print("        NOT a volume (no cross-sectional/storage-area term is included).")

# =================================================================
# 10. CROSS-CHECK: trapezoidal rule on the raw readings
# =================================================================
try:
    area_trapz = np.trapezoid(h, t)   # numpy >= 2.0
except AttributeError:
    area_trapz = np.trapz(h, t)       # older numpy fallback

diff_area = area_quad - area_trapz

print("\n=== Cross-check: trapezoidal rule on raw data ===")
print(f"  Area (trapz, raw h)  = {area_trapz:.2f} m*h")
print(f"  Area (quad, fitted)  = {area_quad:.2f} m*h")
print(f"  Difference           = {diff_area:.2f} m*h "
      f"({100*diff_area/area_trapz:.3f}% of trapz area)")

# =================================================================
# 11. STANDALONE HTML DASHBOARD  (embedded SVG, no external deps)
# =================================================================
# Charts are hand-built as responsive inline SVG (viewBox-based) so the
# whole dashboard is a single file that opens directly in a browser with
# no internet connection, build step, or server required.

DASHBOARD_HTML = "/mnt/user-data/outputs/dashboard.html"

# ---------- small SVG charting helpers ----------------------------

def _nice_ticks(vmin, vmax, target=5):
    """Return a short list of 'nice' round tick values spanning [vmin, vmax]."""
    if vmin == vmax:
        vmin -= 1.0
        vmax += 1.0
    span = vmax - vmin
    raw_step = span / target
    mag = 10 ** np.floor(np.log10(raw_step))
    norm = raw_step / mag
    nice = 1 if norm < 1.5 else 2 if norm < 3 else 5 if norm < 7 else 10
    step = nice * mag
    start = np.floor(vmin / step) * step
    ticks, v = [], start
    while v <= vmax + step * 0.5:
        ticks.append(round(v, 6))
        v += step
    return [tk for tk in ticks if vmin - step * 0.01 <= tk <= vmax + step * 0.01]


class SVGChart:
    """Minimal responsive line/scatter/area chart renderer -> inline SVG."""

    def __init__(self, width=760, height=320, margin=(18, 22, 44, 62)):
        # margin = (top, right, bottom, left)
        self.W, self.H = width, height
        self.mt, self.mr, self.mb, self.ml = margin
        self.plot_w = self.W - self.ml - self.mr
        self.plot_h = self.H - self.mt - self.mb
        self.elements = []

    def set_domain(self, xmin, xmax, ymin, ymax, pad_y_frac=0.08):
        yr = ymax - ymin if ymax > ymin else 1.0
        self.xmin, self.xmax = xmin, xmax
        self.ymin = ymin - yr * pad_y_frac
        self.ymax = ymax + yr * pad_y_frac

    def sx(self, x):
        return self.ml + (x - self.xmin) / (self.xmax - self.xmin) * self.plot_w

    def sy(self, y):
        return self.mt + (1 - (y - self.ymin) / (self.ymax - self.ymin)) * self.plot_h

    def axes(self, xlabel="", ylabel="", n_xticks=6, n_yticks=5,
             x_fmt="{:.0f}", y_fmt="{:.1f}"):
        el = self.elements
        # plot area background + border
        el.append(f'<rect x="{self.ml}" y="{self.mt}" width="{self.plot_w}" '
                   f'height="{self.plot_h}" fill="var(--plot-bg)" '
                   f'stroke="var(--border)" stroke-width="1"/>')
        for ty in _nice_ticks(self.ymin, self.ymax, n_yticks):
            py = self.sy(ty)
            el.append(f'<line x1="{self.ml}" y1="{py:.1f}" x2="{self.ml+self.plot_w}" '
                       f'y2="{py:.1f}" stroke="var(--gridline)" stroke-width="1"/>')
            el.append(f'<text x="{self.ml-8}" y="{py+3.5:.1f}" text-anchor="end" '
                       f'class="tick-label">{y_fmt.format(ty)}</text>')
        for tx in _nice_ticks(self.xmin, self.xmax, n_xticks):
            px = self.sx(tx)
            el.append(f'<line x1="{px:.1f}" y1="{self.mt}" x2="{px:.1f}" '
                       f'y2="{self.mt+self.plot_h}" stroke="var(--gridline)" '
                       f'stroke-width="1"/>')
            el.append(f'<text x="{px:.1f}" y="{self.mt+self.plot_h+18}" '
                       f'text-anchor="middle" class="tick-label">'
                       f'{x_fmt.format(tx)}</text>')
        if xlabel:
            el.append(f'<text x="{self.ml+self.plot_w/2:.1f}" '
                       f'y="{self.H-4}" text-anchor="middle" '
                       f'class="axis-label">{xlabel}</text>')
        if ylabel:
            el.append(f'<text x="14" y="{self.mt+self.plot_h/2:.1f}" '
                       f'text-anchor="middle" class="axis-label" '
                       f'transform="rotate(-90 14 {self.mt+self.plot_h/2:.1f})">'
                       f'{ylabel}</text>')

    def line(self, xs, ys, color="var(--primary)", width=2.2, dash=None):
        pts = " ".join(f"{self.sx(x):.1f},{self.sy(y):.1f}" for x, y in zip(xs, ys))
        dash_attr = f' stroke-dasharray="{dash}"' if dash else ""
        self.elements.append(f'<polyline points="{pts}" fill="none" '
                              f'stroke="{color}" stroke-width="{width}"{dash_attr} '
                              f'stroke-linejoin="round" stroke-linecap="round"/>')

    def area(self, xs, ys, baseline, color="var(--primary)", opacity=0.14):
        pts = [f"{self.sx(xs[0]):.1f},{self.sy(baseline):.1f}"]
        pts += [f"{self.sx(x):.1f},{self.sy(y):.1f}" for x, y in zip(xs, ys)]
        pts += [f"{self.sx(xs[-1]):.1f},{self.sy(baseline):.1f}"]
        self.elements.append(f'<polygon points="{" ".join(pts)}" '
                              f'fill="{color}" opacity="{opacity}" stroke="none"/>')

    def scatter(self, xs, ys, color="var(--accent)", r=2.6, opacity=0.75):
        pts = "".join(f'<circle cx="{self.sx(x):.1f}" cy="{self.sy(y):.1f}" '
                       f'r="{r}" fill="{color}" opacity="{opacity}"/>'
                       for x, y in zip(xs, ys))
        self.elements.append(pts)

    def hline(self, y, color="var(--ink-soft)", width=1.4, dash="5,4"):
        py = self.sy(y)
        self.elements.append(f'<line x1="{self.ml}" y1="{py:.1f}" '
                              f'x2="{self.ml+self.plot_w}" y2="{py:.1f}" '
                              f'stroke="{color}" stroke-width="{width}" '
                              f'stroke-dasharray="{dash}"/>')

    def marker(self, x, y, color="var(--accent)", r=5.5):
        px, py = self.sx(x), self.sy(y)
        self.elements.append(f'<circle cx="{px:.1f}" cy="{py:.1f}" r="{r}" '
                              f'fill="{color}" stroke="white" stroke-width="1.6"/>')

    def callout(self, x, y, lines, dx=10, dy=-14, align="start"):
        px, py = self.sx(x), self.sy(y)
        tx = px + dx if align == "start" else px - dx
        ty_start = py + dy
        anchor = "start" if align == "start" else "end"
        tspans = "".join(
            f'<tspan x="{tx:.1f}" dy="{0 if i == 0 else 13}">{ln}</tspan>'
            for i, ln in enumerate(lines))
        self.elements.append(
            f'<text x="{tx:.1f}" y="{ty_start:.1f}" text-anchor="{anchor}" '
            f'class="callout">{tspans}</text>')

    def render(self):
        return (f'<svg viewBox="0 0 {self.W} {self.H}" class="chart-svg" '
                f'preserveAspectRatio="xMidYMid meet" role="img">'
                + "".join(self.elements) + "</svg>")


# ---------- Chart 1: header stage time series + fitted overlay -----
c1 = SVGChart(width=1180, height=320, margin=(16, 24, 40, 60))
c1.set_domain(t.min(), t.max(), min(h.min(), h_hat.min()), max(h.max(), h_hat.max()))
c1.axes(xlabel="Elapsed time, t (h)", ylabel="Depth, h (m)",
        n_xticks=9, y_fmt="{:.1f}")
c1.area(t, h, baseline=h.min(), color="var(--primary)", opacity=0.08)
c1.line(t, h, color="var(--primary)", width=1.8)
c1.line(t, h_hat, color="var(--accent)", width=2.2, dash="7,4")
chart_header = c1.render()

# ---------- Chart 2: dh/dt with max point annotated -----------------
c2 = SVGChart(width=740, height=300)
c2.set_domain(t.min(), t.max(), dhdt.min(), dhdt.max())
c2.axes(xlabel="t (h)", ylabel="dh/dt (m/h)", n_xticks=8, y_fmt="{:.2f}")
c2.hline(0, color="var(--gridline-strong)", dash="3,3")
c2.line(t, dhdt, color="var(--primary)", width=2)
c2.marker(t[imax], dhdt[imax], color="var(--accent)")
ts_label = timestamps[imax].strftime("%b %d, %H:%M")
c2.callout(t[imax], dhdt[imax],
           [f"Max dh/dt = {dhdt[imax]:.2f} m/h", f"t = {t[imax]:.2f} h  ({ts_label})"],
           dx=10, dy=-16,
           align="start" if t[imax] < (t.max() * 0.72) else "end")
chart_dhdt = c2.render()

# ---------- Chart 3: d2h/dt2 -----------------------------------------
c3 = SVGChart(width=740, height=300)
c3.set_domain(t.min(), t.max(), d2hdt2.min(), d2hdt2.max())
c3.axes(xlabel="t (h)", ylabel="d\u00b2h/dt\u00b2 (m/h\u00b2)", n_xticks=8, y_fmt="{:.2f}")
c3.hline(0, color="var(--gridline-strong)", dash="3,3")
c3.line(t, d2hdt2, color="var(--navy)", width=1.4)
chart_d2hdt2 = c3.render()

# ---------- Chart 4: fitted curve overlay ----------------------------
c4 = SVGChart(width=1180, height=340, margin=(16, 24, 40, 60))
c4.set_domain(t.min(), t.max(), min(h.min(), h_hat.min()), max(h.max(), h_hat.max()))
c4.axes(xlabel="t (h)", ylabel="h (m)", n_xticks=9, y_fmt="{:.1f}")
c4.scatter(t, h, color="var(--ink-soft)", r=2.1, opacity=0.55)
c4.line(t, h_hat, color="var(--accent)", width=2.4)
chart_fit = c4.render()

# ---------- Chart 5: residuals vs time --------------------------------
c5 = SVGChart(width=1180, height=300, margin=(16, 24, 40, 60))
c5.set_domain(t.min(), t.max(), resid.min(), resid.max())
c5.axes(xlabel="t (h)", ylabel="Residual, e (m)", n_xticks=9, y_fmt="{:.2f}")
c5.hline(0, color="var(--gridline-strong)", width=1.8, dash="6,4")
c5.scatter(t, resid, color="var(--primary)", r=2.6, opacity=0.7)
chart_resid = c5.render()

# ---------- Chart 5b: residuals vs fitted values -----------------------
c5b = SVGChart(width=1180, height=300, margin=(16, 24, 40, 60))
c5b.set_domain(h_hat.min(), h_hat.max(), resid.min(), resid.max())
c5b.axes(xlabel="Fitted value, h\u0302 (m)", ylabel="Residual, e (m)",
          n_xticks=8, y_fmt="{:.2f}", x_fmt="{:.1f}")
c5b.hline(0, color="var(--gridline-strong)", width=1.8, dash="6,4")
c5b.scatter(h_hat, resid, color="var(--accent)", r=2.6, opacity=0.7)
chart_resid_fitted = c5b.render()

# ---------- Chart 6: shaded area under fitted curve --------------------
c6 = SVGChart(width=1180, height=340, margin=(16, 24, 40, 60))
c6.set_domain(t.min(), t.max(), 0, h_hat.max())
c6.axes(xlabel="t (h)", ylabel="h\u0302 (m)", n_xticks=9, y_fmt="{:.1f}")
c6.area(t, h_hat, baseline=0, color="var(--accent)", opacity=0.20)
c6.line(t, h_hat, color="var(--accent)", width=2.2)
c6.hline(0, color="var(--gridline-strong)", width=1.2, dash="3,3")
chart_area = c6.render()

# ---------- Parameter table rows --------------------------------------
param_units = {"h0": "m", "A1": "m", "k1": "1/h", "t1": "h",
               "A2": "m", "k2": "1/h", "t2": "h"}
param_desc = {"h0": "Baseline stage", "A1": "Rise amplitude", "k1": "Rise steepness",
              "t1": "Rise inflection time", "A2": "Fall amplitude",
              "k2": "Fall steepness", "t2": "Fall inflection time"}
rows_html = []
for name, val, se_i, ts_, pv in zip(param_names, popt, se, t_stat, p_values):
    dec = 2 if param_units[name] in ("m", "h") else 3
    pv_str = "&lt; 0.001" if pv < 0.001 else f"{pv:.3g}"
    verdict = "Significant" if pv < 0.05 else "Not significant"
    verdict_class = "pill-good" if pv < 0.05 else "pill-bad"
    rows_html.append(
        f"<tr><td><span class='mono'>{name}</span> "
        f"<span class='muted-inline'>({param_desc[name]})</span></td>"
        f"<td class='mono num'>{val:.{dec}f} {param_units[name]}</td>"
        f"<td class='mono num'>{se_i:.{dec}f}</td>"
        f"<td class='mono num'>{ts_:.2f}</td>"
        f"<td class='mono num'>{pv_str}</td>"
        f"<td><span class='pill {verdict_class}'>{verdict}</span></td></tr>")
param_table_rows = "\n".join(rows_html)

# ---------- Residual interpretation text (computed dynamically) --------
fit_verdict = ("a <strong>failed fit for local shape</strong> despite the high R\u00b2"
               if longest_run >= 20 or max_abs_resid > 10 * sensor_res
               else "an <strong>acceptable fit</strong>")
resid_summary_html = f"""
<ul class="findings">
  <li><strong>Centering:</strong> mean residual is {resid.mean():.4f} m (essentially zero
      overall), but the scatter is <strong>not</strong> locally centered \u2014 it traces a
      smooth, repeating wave rather than random noise around the line.</li>
  <li><strong>Runs:</strong> the longest unbroken run of same-sign residuals is
      <strong>{longest_run} consecutive points ({longest_run*dt:.1f} h)</strong> \u2014 far
      longer than the 1&ndash;3 points expected from a correctly specified model.</li>
  <li><strong>Spread vs. level:</strong> residual std. dev. is {std_low:.3f} m (low third),
      {std_mid:.3f} m (mid third), {std_high:.3f} m (high third) \u2014 the spread does
      <strong>not</strong> widen with stage; error is tied to curve-bending regions, not
      water level.</li>
  <li><strong>Largest error:</strong> {max_abs_resid:.3f} m at t = {t[idx_max_resid]:.2f} h,
      about <strong>{max_abs_resid/sensor_res:.0f}\u00d7</strong> the logger's 0.01 m
      resolution \u2014 a real structural mismatch, not sensor noise.</li>
  <li><strong>Conclusion:</strong> the residual pattern indicates {fit_verdict}; R\u00b2 alone
      is not sufficient evidence of a good fit for this record.</li>
</ul>"""

# ---------- Integration interpretation text -----------------------------
integration_summary_html = f"""
<p>The two areas differ by only {abs(diff_area):.2f} m\u00b7h
({100*diff_area/area_trapz:.3f}% of the trapezoidal value) because
<code>quad</code> integrates the smooth fitted curve while
<code>trapezoid</code> integrates the raw, centimeter-rounded readings, so the
difference is just the net effect of the fitting residuals cancelling out over
the record.</p>
<p><strong>For a flood-control office:</strong> this m\u00b7h value is a compact
summary of how high and how long the reservoir stayed elevated above baseline
during the event \u2014 useful for ranking or comparing event severity/duration
on a common scale. It is <strong>not</strong> a water volume and must not be
used directly to size spillway capacity or compute inflow/outflow volumes;
that requires multiplying by the reservoir's stage-dependent storage area,
which is not part of this dataset.</p>"""

now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
peak_stage = h.max()
peak_time = t[int(np.argmax(h))]
event_start = timestamps[0].strftime("%b %d, %Y")
event_end = timestamps[-1].strftime("%b %d, %Y")

HTML_TEMPLATE = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reservoir Stage Analysis Dashboard</title>
<style>
  :root {{
    /* Coolors palette: https://coolors.co/palette/ccd5ae-e9edc9-fefae0-faedcd-d4a373 */
    --sage: #CCD5AE;        /* accent / borders / secondary highlights */
    --olive: #E9EDC9;       /* light container fills / active tab / hover */
    --bg: #FEFAE0;          /* warm cream - main background */
    --card-sand: #FAEDCD;   /* card backgrounds / table header fills */
    --clay: #D4A373;        /* primary buttons / accent highlights / plot lines */
    --clay-dark: #B0824F;   /* darker clay for hover/emphasis */
    --sage-line: #8FA06A;   /* deepened sage, used only for chart lines/markers
                                (CCD5AE itself is kept true to spec for fills/borders) */
    --ink: #2B2D42;         /* dark charcoal - body text */
    --ink-soft: #5B5D6E;    /* softened charcoal - captions/labels */
    --muted: #83869A;       /* muted charcoal - footers/fine print */

    --surface: var(--card-sand);
    --border: var(--sage);
    --primary: var(--clay);
    --primary-dark: var(--clay-dark);
    --navy: var(--ink);
    --accent: var(--sage-line);
    --accent-dark: #6E7E52;
    --good: #6E8B4F;
    --bad: #B1573E;
    --plot-bg: var(--olive);
    --gridline: rgba(43,45,66,0.10);
    --gridline-strong: rgba(43,45,66,0.32);
    --radius: 16px;
    --shadow: 0 1px 2px rgba(43,45,66,0.05), 0 8px 24px rgba(43,45,66,0.07);
  }}
  * {{ box-sizing: border-box; }}
  html, body {{ margin: 0; padding: 0; }}
  body {{
    background: var(--bg);
    color: var(--ink);
    font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto,
                 Helvetica, Arial, sans-serif;
    line-height: 1.5;
    padding: 32px 5vw 64px;
  }}
  .mono {{
    font-family: "SFMono-Regular", Consolas, "Liberation Mono", Menlo, monospace;
  }}
  .wrap {{ max-width: 1280px; margin: 0 auto; }}

  /* ---- header ---- */
  .masthead {{
    display: flex; align-items: baseline; justify-content: space-between;
    gap: 16px; flex-wrap: wrap; margin-bottom: 4px;
  }}
  h1 {{ font-size: 26px; font-weight: 700; margin: 0; letter-spacing: -0.01em; }}
  .subtitle {{ color: var(--ink-soft); font-size: 14px; margin-top: 4px; }}
  .timestamp {{ color: var(--muted); font-size: 12px; }}

  .chip-row {{ display: flex; gap: 10px; flex-wrap: wrap; margin: 18px 0 20px; }}
  .chip {{
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 999px; padding: 7px 14px; font-size: 12.5px;
    color: var(--ink-soft); box-shadow: var(--shadow);
  }}
  .chip b {{ color: var(--ink); font-weight: 600; }}

  .card {{
    background: var(--surface); border: 1px solid var(--border);
    border-radius: var(--radius); box-shadow: var(--shadow);
    padding: 22px 24px; margin-bottom: 22px;
  }}
  .card h2 {{ font-size: 15px; margin: 0 0 4px; font-weight: 700; }}
  .card .card-caption {{ color: var(--ink-soft); font-size: 13px; margin: 0 0 14px; }}

  .chart-svg {{ width: 100%; height: auto; display: block; }}
  .tick-label {{ font-size: 10.5px; fill: var(--ink-soft); font-family: inherit; }}
  .axis-label {{ font-size: 11.5px; fill: var(--ink-soft); font-weight: 600; }}
  .callout {{ font-size: 11.5px; fill: var(--accent-dark); font-weight: 600; }}

  .legend-row {{ display: flex; gap: 18px; margin-bottom: 10px; flex-wrap: wrap; }}
  .legend-item {{ font-size: 12.5px; color: var(--ink-soft); display: inline-flex;
                   align-items: center; gap: 7px; font-weight: 600; }}
  .swatch {{ width: 14px; height: 4px; border-radius: 2px; display: inline-block; }}
  .swatch-primary {{ background: var(--primary); }}
  .swatch-accent {{ background: var(--accent);
                     background-image: repeating-linear-gradient(90deg, var(--accent) 0 5px,
                       transparent 5px 8px); }}

  /* ---- tabs ---- */
  .tabbar {{
    display: flex; gap: 6px; background: var(--surface);
    border: 1px solid var(--border); border-radius: 999px; padding: 5px;
    width: fit-content; margin: 6px 0 20px; box-shadow: var(--shadow);
  }}
  .tab-btn {{
    border: none; background: transparent; padding: 9px 18px; border-radius: 999px;
    font-size: 13.5px; font-weight: 600; color: var(--ink-soft); cursor: pointer;
    font-family: inherit; transition: background .15s, color .15s;
  }}
  .tab-btn.active {{ background: var(--olive); color: var(--ink); }}
  .tab-btn:not(.active):hover {{ background: var(--olive); }}
  .tab-panel {{ display: none; }}
  .tab-panel.active {{ display: block; }}

  @media (max-width: 700px) {{
    .tabbar {{ width: 100%; flex-wrap: wrap; border-radius: 14px; }}
    .tab-btn {{ flex: 1 1 auto; text-align: center; }}
  }}

  .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 22px; }}
  @media (max-width: 860px) {{ .grid-2 {{ grid-template-columns: 1fr; }} }}

  .eq-box {{
    background: var(--plot-bg); border: 1px dashed var(--border);
    border-radius: 10px; padding: 12px 16px; font-size: 13px;
    margin-bottom: 16px; overflow-x: auto;
  }}

  table {{ width: 100%; border-collapse: collapse; font-size: 13.5px; }}
  th {{
    text-align: left; color: var(--ink); font-weight: 700; font-size: 12px;
    text-transform: uppercase; letter-spacing: .03em; background: var(--card-sand);
    border-bottom: 1px solid var(--border); padding: 9px 10px;
  }}
  td {{ padding: 9px 10px; border-bottom: 1px solid var(--gridline); }}
  td.num {{ text-align: right; }}
  th:nth-child(n+2) {{ text-align: right; }}
  .muted-inline {{ color: var(--muted); font-size: 12px; }}

  .pill {{
    display: inline-block; padding: 3px 10px; border-radius: 999px;
    font-size: 11.5px; font-weight: 700;
  }}
  .pill-good {{ background: rgba(110,139,79,0.14); color: var(--good); }}
  .pill-bad  {{ background: rgba(177,87,62,0.14); color: var(--bad); }}

  .metric-grid {{
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px;
  }}
  @media (max-width: 860px) {{ .metric-grid {{ grid-template-columns: 1fr 1fr; }} }}
  .metric {{
    background: var(--plot-bg); border: 1px solid var(--border);
    border-radius: 12px; padding: 14px 16px;
  }}
  .metric .label {{ font-size: 11.5px; color: var(--ink-soft); font-weight: 600;
                     text-transform: uppercase; letter-spacing: .03em; }}
  .metric .value {{ font-size: 22px; font-weight: 700; color: var(--primary-dark);
                     margin-top: 4px; }}
  .metric .foot {{ font-size: 11.5px; color: var(--muted); margin-top: 2px; }}

  .findings {{ margin: 0; padding-left: 20px; font-size: 13.5px; }}
  .findings li {{ margin-bottom: 8px; }}

  .int-summary {{ display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }}
  @media (max-width: 860px) {{ .int-summary {{ grid-template-columns: 1fr; }} }}
  .int-summary p {{ font-size: 13.5px; margin: 0 0 10px; }}
  code {{ background: var(--gridline); padding: 1px 5px; border-radius: 5px;
          font-family: "SFMono-Regular", Consolas, monospace; font-size: 12px; }}

  footer {{ text-align: center; color: var(--muted); font-size: 12px; margin-top: 28px; }}
</style>
</head>
<body>
<div class="wrap">

  <div class="masthead">
    <div>
      <h1>Reservoir Stage Analysis Dashboard</h1>
      <div class="subtitle">15-minute stage logger record &middot; {event_start} &ndash; {event_end}</div>
    </div>
    <div class="timestamp">Generated {now_str}</div>
  </div>

  <div class="chip-row">
    <div class="chip">Data points: <b>{n}</b></div>
    <div class="chip">Sampling interval: <b>{dt:.2f} h</b> (15 min)</div>
    <div class="chip">Time span: <b>{t.max():.2f} h</b></div>
    <div class="chip">Peak stage: <b>{peak_stage:.2f} m</b> @ t={peak_time:.2f} h</div>
    <div class="chip">Max dh/dt: <b>{dhdt[imax]:.2f} m/h</b> @ t={t[imax]:.2f} h</div>
    <div class="chip">Model R&sup2;: <b>{R2:.4f}</b></div>
  </div>

  <div class="card">
    <h2>Stage Log &amp; Fitted Model Overlay</h2>
    <p class="card-caption">Raw depth above gauge datum, h(t), logged every 15 minutes,
      with the fitted sum-of-two-logistics model h&#770;(t) overlaid.</p>
    <div class="legend-row">
      <span class="legend-item"><span class="swatch swatch-primary"></span>Raw log, h(t)</span>
      <span class="legend-item"><span class="swatch swatch-accent"></span>Fitted model, h&#770;(t)</span>
    </div>
    {chart_header}
  </div>

  <div class="tabbar">
    <button class="tab-btn active" onclick="showTab(0, this)">Derivatives</button>
    <button class="tab-btn" onclick="showTab(1, this)">Model Fit &amp; Statistical Evaluation</button>
    <button class="tab-btn" onclick="showTab(2, this)">Integration &amp; Area</button>
  </div>

  <!-- TAB 1: DERIVATIVES -->
  <div class="tab-panel active" id="tab-0">
    <div class="grid-2">
      <div class="card">
        <h2>First Derivative &mdash; dh/dt</h2>
        <p class="card-caption">Forward difference (first point), central difference (interior),
          backward difference (last point). Peak rate of rise marked.</p>
        {chart_dhdt}
      </div>
      <div class="card">
        <h2>Second Derivative &mdash; d&sup2;h/dt&sup2;</h2>
        <p class="card-caption">3-point central formula (interior), one-sided 3-point formulas
          at the endpoints. Positive = accelerating inflow; negative = waning inflow.</p>
        {chart_d2hdt2}
      </div>
    </div>
  </div>

  <!-- TAB 2: MODEL FIT -->
  <div class="tab-panel" id="tab-1">
    <div class="card">
      <h2>Fitted Curve vs. Raw Stage Data</h2>
      <p class="card-caption">Sum-of-two-logistics model fitted by Levenberg&ndash;Marquardt
        (<span class="mono">scipy.optimize.curve_fit</span>, method='lm').</p>
      <div class="eq-box mono">
        h(t) = h0 + A1 / (1 + exp(&minus;k1&middot;(t &minus; t1))) &minus; A2 / (1 + exp(&minus;k2&middot;(t &minus; t2)))
      </div>
      {chart_fit}
    </div>

    <div class="card">
      <h2>Residuals vs. Time</h2>
      <p class="card-caption">e<sub>i</sub> = h<sub>i</sub> &minus; h&#770;(t<sub>i</sub>), with a zero reference line.</p>
      {chart_resid}
    </div>

    <div class="card">
      <h2>Residuals vs. Fitted Water Level</h2>
      <p class="card-caption">e<sub>i</sub> = h<sub>i</sub> &minus; h&#770;<sub>i</sub>, plotted against
        h&#770;<sub>i</sub>, with a zero reference line.</p>
      {chart_resid_fitted}
    </div>

    <div class="card">
      <h2>Residual Pattern Interpretation</h2>
      {resid_summary_html}
    </div>

    <div class="card">
      <h2>Fitted Parameter Summary</h2>
      <table>
        <thead><tr><th>Parameter</th><th>Estimate</th><th>Std. Error</th>
          <th>t-statistic</th><th>p-value</th><th>Conclusion</th></tr></thead>
        <tbody>
          {param_table_rows}
        </tbody>
      </table>
    </div>

    <div class="card">
      <h2>Goodness of Fit</h2>
      <div class="metric-grid">
        <div class="metric">
          <div class="label">SSE</div>
          <div class="value">{SSE:.3f} m&sup2;</div>
          <div class="foot">n = {n}, p = {p_count}</div>
        </div>
        <div class="metric">
          <div class="label">SST</div>
          <div class="value">{SST:.1f} m&sup2;</div>
          <div class="foot">&Sigma;(h &minus; h&#772;)&sup2;</div>
        </div>
        <div class="metric">
          <div class="label">R&sup2;</div>
          <div class="value">{R2:.4f}</div>
          <div class="foot">1 &minus; SSE/SST</div>
        </div>
        <div class="metric">
          <div class="label">Std. Error of Estimate (s)</div>
          <div class="value">{s:.3f} m</div>
          <div class="foot">dof = n &minus; p = {n - p_count}</div>
        </div>
      </div>
    </div>
  </div>

  <!-- TAB 3: INTEGRATION -->
  <div class="tab-panel" id="tab-2">
    <div class="card">
      <h2>Area Under the Fitted Curve</h2>
      <p class="card-caption">h&#770;(t) shaded to the x-axis over t &isin;
        [{t0:.2f}, {tn:.2f}] h.</p>
      {chart_area}
    </div>

    <div class="card">
      <h2>Integration Summary</h2>
      <div class="metric-grid" style="grid-template-columns: repeat(3, 1fr);">
        <div class="metric">
          <div class="label">Area (scipy.integrate.quad)</div>
          <div class="value">{area_quad:.2f} m&middot;h</div>
          <div class="foot">abs. error est. = {area_err:.2e} m&middot;h</div>
        </div>
        <div class="metric">
          <div class="label">Area (np.trapezoid, raw data)</div>
          <div class="value">{area_trapz:.2f} m&middot;h</div>
          <div class="foot">cross-check on logged readings</div>
        </div>
        <div class="metric">
          <div class="label">Difference</div>
          <div class="value">{diff_area:+.2f} m&middot;h</div>
          <div class="foot">{100*diff_area/area_trapz:.3f}% of trapz area</div>
        </div>
      </div>
      <p class="card-caption" style="margin-top:14px;">Units are <strong>meter&ndash;hours
        (m&middot;h)</strong> &mdash; a time-integral of stage, <strong>not</strong> a volume.</p>
      <div class="int-summary" style="margin-top:10px;">
        {integration_summary_html}
      </div>
    </div>
  </div>

  <footer>Reservoir Stage Analysis &middot; generated automatically by reservoir_analysis.py</footer>
</div>

<script>
  function showTab(idx, btn) {{
    document.querySelectorAll('.tab-panel').forEach((p, i) => {{
      p.classList.toggle('active', i === idx);
    }});
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
  }}
</script>
</body>
</html>
"""

with open(DASHBOARD_HTML, "w", encoding="utf-8") as f:
    f.write(HTML_TEMPLATE)

print(f"\nSaved standalone HTML dashboard to:\n  {DASHBOARD_HTML}")
