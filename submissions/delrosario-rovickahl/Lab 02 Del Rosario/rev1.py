import os
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.lines import Line2D
from mpl_toolkits.mplot3d.art3d import Poly3DCollection
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# 1. GEOMETRY & LOCAL AXES HELPERS
# =============================================================================
NODES = {
    1: [0.0, 0.0, 0.0],
    2: [6.0, 0.0, 0.0],
    3: [6.0, 0.0, 6.0],
    4: [0.0, 0.0, 6.0],
    5: [0.0, 6.0, 0.0],
    6: [6.0, 6.0, 0.0],
    7: [6.0, 6.0, 6.0],
    8: [0.0, 6.0, 6.0],
}

SUPPORTS = {
    1: [1, 1, 1, 0, 0, 0],
    2: [1, 1, 1, 0, 0, 0],
    3: [1, 1, 1, 0, 0, 0],
    4: [1, 1, 1, 0, 0, 0],
    5: [0, 0, 0, 0, 0, 0],
    6: [0, 0, 0, 0, 0, 0],
    7: [0, 0, 0, 0, 0, 0],
    8: [0, 0, 0, 0, 0, 0],
}

MEMBERS = [
    [1, 1, 2, "Beam", 0.0, "NONE", "NONE"],
    [2, 2, 3, "Beam", 0.0, "NONE", "NONE"],
    [3, 3, 4, "Beam", 0.0, "NONE", "NONE"],
    [4, 4, 1, "Beam", 0.0, "NONE", "NONE"],
    [5, 5, 6, "Beam", 0.0, "MX-MZ", "NONE"],
    [6, 6, 7, "Beam", 0.0, "NONE", "NONE"],
    [7, 7, 8, "Beam", 0.0, "NONE", "NONE"],
    [8, 8, 5, "Beam", 0.0, "NONE", "NONE"],
    [9, 1, 5, "Column", 90.0, "NONE", "NONE"],
    [10, 2, 6, "Column", 90.0, "NONE", "NONE"],
    [11, 3, 7, "Column", 90.0, "NONE", "NONE"],
    [12, 4, 8, "Column", 90.0, "NONE", "NONE"],
]


def calculate_local_axes(p1, p2, beta_deg=0.0):
    """Calculates member local orientation vectors (vx, vy, vz)."""
    p1, p2 = np.array(p1), np.array(p2)
    vx = p2 - p1
    vx = vx / np.linalg.norm(vx)

    global_y = np.array([0.0, 1.0, 0.0])

    if np.abs(np.dot(vx, global_y)) > 0.999:
        vz_init = np.array([0.0, 0.0, 1.0])
        vy_init = np.cross(vz_init, vx)
        vy_init /= np.linalg.norm(vy_init)
    else:
        vz_init = np.cross(vx, global_y)
        vz_init /= np.linalg.norm(vz_init)
        vy_init = np.cross(vz_init, vx)
        vy_init /= np.linalg.norm(vy_init)

    beta_rad = np.radians(beta_deg)
    vy = vy_init * np.cos(beta_rad) + vz_init * np.sin(beta_rad)
    vz = -vy_init * np.sin(beta_rad) + vz_init * np.cos(beta_rad)

    return vx, vy, vz


# =============================================================================
# 2. EXCEL GENERATION FUNCTION
# =============================================================================
def generate_excel_report(output_filename="Cube_Structural_Model_Rev1.xlsx"):
    """Generates and styles the formatted multi-tab Excel file."""

    # Tab Data Creation
    summary_df = pd.DataFrame(
        {
            "Parameter": [
                "Structure Type",
                "Cube Edge Length",
                "Total Nodes",
                "Total Members",
                "Global Axis System",
                "DOF per Node",
                "Total System DOFs",
                "Restrained DOFs",
                "Active System DOFs (Equations)",
            ],
            "Value": [
                "3D Frame Model (Rev 1)",
                "6.0 m",
                8,
                12,
                "Y-Up Vertical",
                6,
                48,
                12,
                36,
            ],
        }
    )

    nodes_df = pd.DataFrame(
        [
            {"Node": k, "X (m)": v[0], "Y (m)": v[1], "Z (m)": v[2]}
            for k, v in NODES.items()
        ]
    )

    members_df = pd.DataFrame(
        MEMBERS,
        columns=[
            "Member",
            "Node i (Start)",
            "Node j (End)",
            "Type",
            "Beta Angle (deg)",
            "Release i",
            "Release j",
        ],
    )

    supports_df = pd.DataFrame(
        [
            {
                "Node": k,
                "Support": "Pinned" if sum(v[:3]) == 3 else "Free",
                "UX": "Restrained" if v[0] else "Free",
                "UY": "Restrained" if v[1] else "Free",
                "UZ": "Restrained" if v[2] else "Free",
                "RX": "Restrained" if v[3] else "Free",
                "RY": "Restrained" if v[4] else "Free",
                "RZ": "Restrained" if v[5] else "Free",
            }
            for k, v in SUPPORTS.items()
        ]
    )

    # Requested Node DOF Summary Layout
    node_summary_rows = []
    for node_id, sup in SUPPORTS.items():
        node_summary_rows.append(
            {
                "Node": node_id,
                "Support": "Pinned" if sum(sup[:3]) == 3 else "Free",
                "Global DOF Range": f"DOF {(node_id-1)*6 + 1}-{node_id*6}",
                "UX": "Restrained" if sup[0] else "Active",
                "UY": "Restrained" if sup[1] else "Active",
                "UZ": "Restrained" if sup[2] else "Active",
                "RX": "Restrained" if sup[3] else "Active",
                "RY": "Restrained" if sup[4] else "Active",
                "RZ": "Restrained" if sup[5] else "Active",
                "Active DOFs": sum(1 for val in sup if val == 0),
            }
        )
    node_dof_summary_df = pd.DataFrame(node_summary_rows)

    # Detailed DOF Numbering
    dof_names = ["UX", "UY", "UZ", "RX", "RY", "RZ"]
    dof_descriptions = [
        "Translation X",
        "Translation Y",
        "Translation Z",
        "Rotation X",
        "Rotation Y",
        "Rotation Z",
    ]
    dof_rows = []
    eq_counter = 1

    for node_id in range(1, 9):
        sup = SUPPORTS[node_id]
        for local_idx in range(6):
            global_dof = (node_id - 1) * 6 + (local_idx + 1)
            is_restrained = sup[local_idx] == 1
            dof_rows.append(
                {
                    "Node": node_id,
                    "Local DOF": local_idx + 1,
                    "DOF": dof_names[local_idx],
                    "Description": dof_descriptions[local_idx],
                    "Global DOF No.": global_dof,
                    "Status": "Restrained" if is_restrained else "Active",
                    "Equation No.": "-" if is_restrained else str(eq_counter),
                }
            )
            if not is_restrained:
                eq_counter += 1
    dof_numbering_df = pd.DataFrame(dof_rows)

    # Export to Excel File
    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Model Summary", index=False)
        nodes_df.to_excel(writer, sheet_name="Nodes", index=False)
        members_df.to_excel(writer, sheet_name="Member Incidences", index=False)
        supports_df.to_excel(writer, sheet_name="Supports", index=False)
        node_dof_summary_df.to_excel(
            writer, sheet_name="Node DOF Summary", index=False
        )
        dof_numbering_df.to_excel(
            writer, sheet_name="DOF Numbering", index=False
        )

    # OpenPyXL Styling
    wb = openpyxl.load_workbook(output_filename)
    header_fill = PatternFill(
        start_color="1B365D", end_color="1B365D", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    active_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    active_font = Font(name="Calibri", size=10, color="375623")
    restrained_fill = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )
    restrained_font = Font(name="Calibri", size=10, color="C65911")
    border_style = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border_style
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

                if cell.value in ["Active", "Free"]:
                    cell.fill = active_fill
                    cell.font = active_font
                elif cell.value in ["Restrained", "Pinned"]:
                    cell.fill = restrained_fill
                    cell.font = restrained_font

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(
                max_len + 4, 14
            )

    wb.save(output_filename)
    print(
        f"Excel saved successfully at: {os.path.abspath(output_filename)}"
    )


# =============================================================================
# 3. MATPLOTLIB 3D VISUALIZATION FUNCTION
# =============================================================================
def plot_structural_model():
    """Renders the 3D structural model with gray theme."""
    bg_color = "#3A3D40"  # Medium Gray Canvas Background
    pane_color = "#2D3033"  # Slightly Darker Gray Panes
    grid_color = "#5A5E63"  # Subtle Grid Lines

    fig = plt.figure(figsize=(14, 9), facecolor=bg_color)
    ax = fig.add_subplot(111, projection="3d", facecolor=bg_color)

    # Style Panes and Grid
    for pane in (ax.xaxis.pane, ax.yaxis.pane, ax.zaxis.pane):
        pane.set_facecolor(pane_color)
        pane.set_edgecolor(grid_color)
    ax.grid(True, color=grid_color, linestyle="--", linewidth=0.6)

    # Draw Members and Local Vector Axes
    for row in MEMBERS:
        m_id, ni, nj, m_type, beta = (
            row[0],
            row[1],
            row[2],
            row[3],
            float(row[4]),
        )
        ci, cj = np.array(NODES[ni]), np.array(NODES[nj])
        color = "#58A6FF" if m_type == "Beam" else "#3FB950"

        # Structural Lines (Y-Up mapped to Z-axis in plot)
        ax.plot(
            [ci[0], cj[0]],
            [ci[2], cj[2]],
            [ci[1], cj[1]],
            color=color,
            linewidth=3.0,
            zorder=3,
        )

        # Quiver Local Vectors
        mid = (ci + cj) / 2.0
        vx, vy, vz = calculate_local_axes(ci, cj, beta)
        scale = 0.65
        ax.quiver(
            mid[0],
            mid[2],
            mid[1],
            vx[0],
            vx[2],
            vx[1],
            color="#FF7B72",
            length=scale,
            arrow_length_ratio=0.2,
        )
        ax.quiver(
            mid[0],
            mid[2],
            mid[1],
            vy[0],
            vy[2],
            vy[1],
            color="#7EE787",
            length=scale,
            arrow_length_ratio=0.2,
        )
        ax.quiver(
            mid[0],
            mid[2],
            mid[1],
            vz[0],
            vz[2],
            vz[1],
            color="#D2A8FF",
            length=scale,
            arrow_length_ratio=0.2,
        )

        lbl = f"M{m_id}" if beta == 0 else f"M{m_id} (β={int(beta)}°)"
        ax.text(
            mid[0] + 0.12,
            mid[2] + 0.12,
            mid[1] + 0.12,
            lbl,
            color="#FFFFFF",
            fontsize=8,
            fontweight="bold",
        )

    # Draw Nodes
    for n_id, coords in NODES.items():
        x, y, z = coords[0], coords[1], coords[2]
        is_supported = n_id in [1, 2, 3, 4]
        color = "#FF5555" if is_supported else "#FF7B72"

        ax.scatter(x, z, y, color=color, s=85, edgecolors="#FFFFFF", zorder=5)
        dof_text = f"N{n_id}\nDOF {(n_id-1)*6 + 1}-{n_id*6}"
        ax.text(
            x + 0.15,
            z + 0.15,
            y + 0.25,
            dof_text,
            color="#FFFFFF",
            fontsize=8,
            fontweight="bold",
        )

    # Draw Pin Support Bases
    for n_id in [1, 2, 3, 4]:
        apex_xyz = NODES[n_id]
        ax_x, ax_y, ax_z = apex_xyz[0], apex_xyz[2], apex_xyz[1]
        hw, height = 0.3, 0.7
        base_nodes = [
            [ax_x - hw, ax_y - hw, ax_z - height],
            [ax_x + hw, ax_y - hw, ax_z - height],
            [ax_x + hw, ax_y + hw, ax_z - height],
            [ax_x - hw, ax_y + hw, ax_z - height],
        ]
        apex = [ax_x, ax_y, ax_z]
        faces = [
            [base_nodes[0], base_nodes[1], apex],
            [base_nodes[1], base_nodes[2], apex],
            [base_nodes[2], base_nodes[3], apex],
            [base_nodes[3], base_nodes[0], apex],
            [base_nodes[0], base_nodes[1], base_nodes[2], base_nodes[3]],
        ]
        pyramid = Poly3DCollection(
            faces, facecolors="#57606A", edgecolors="#ADB5BD", alpha=0.9
        )
        ax.add_collection3d(pyramid)

    # Legend Formatting
    legend_elements = [
        Line2D([0], [0], color="#58A6FF", lw=2.5, label="Beam"),
        Line2D([0], [0], color="#3FB950", lw=2.5, label="Column"),
        Line2D(
            [0],
            [0],
            marker="o",
            color="w",
            markerfacecolor="#FF5555",
            markersize=7,
            label="Supported node (pinned)",
        ),
        Line2D(
            [0],
            [0],
            marker="o",
            color="w",
            markerfacecolor="#FF7B72",
            markersize=7,
            label="Free node",
        ),
        Line2D([0], [0], color="#FF7B72", lw=1.5, label="Local x axis"),
        Line2D([0], [0], color="#7EE787", lw=1.5, label="Local y axis"),
        Line2D([0], [0], color="#D2A8FF", lw=1.5, label="Local z axis"),
    ]
    ax.legend(
        handles=legend_elements,
        loc="upper left",
        fontsize=8,
        facecolor="#2D3033",
        edgecolor="#5A5E63",
        labelcolor="#FFFFFF",
    )

    # Axis Labels & Scene Setup
    ax.set_xlabel(
        "X (m) - lateral", labelpad=10, fontweight="bold", color="#FFFFFF"
    )
    ax.set_ylabel(
        "Z (m) - lateral", labelpad=10, fontweight="bold", color="#FFFFFF"
    )
    ax.set_zlabel(
        "Y (m) - vertical", labelpad=10, fontweight="bold", color="#FFFFFF"
    )
    ax.tick_params(colors="#FFFFFF")
    ax.set_xlim([-1, 7])
    ax.set_ylim([-1, 7])
    ax.set_zlim([-1, 7])

    plt.title(
        "6m x 6m x 6m Cube - Structural Model, Rev. 1\n"
        "Pinned supports at nodes 1-4, member local axes and beta angles shown",
        pad=15,
        fontsize=11,
        fontweight="bold",
        color="#FFFFFF",
    )
    ax.view_init(elev=20, azim=-55)
    plt.subplots_adjust(left=0.02, right=0.75, top=0.92, bottom=0.05)

    plt.show()


# =============================================================================
# 4. SCRIPT EXECUTION ENTRY POINT
# =============================================================================
if __name__ == "__main__":
    generate_excel_report("Cube_Structural_Model_Rev1.xlsx")
    plot_structural_model()