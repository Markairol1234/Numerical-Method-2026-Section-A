from pathlib import Path
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from matplotlib.lines import Line2D
from mpl_toolkits.mplot3d.art3d import Poly3DCollection


# ============================================================
# REV. 1 — 3D FRAME STRUCTURAL SOLVER
# ============================================================
# Model:
#   6 m x 6 m x 6 m cube
#
# Global axes:
#   X = lateral
#   Y = vertical
#   Z = lateral
#
# Node DOF:
#   UX, UY, UZ, RX, RY, RZ
#
# Supports:
#   Nodes 1-4 = pinned
#   UX, UY, UZ restrained
#   RX, RY, RZ released
#
# Member beta angles:
#   Beams   = 0 degrees
#   Columns = 90 degrees
#
# Local axes:
#   Local x = member i node to j node
#   Local y = transverse axis
#   Local z = completes the right-handed system
#
# Beam release:
#   Local Mz release is documented in the Excel output.
#   It is not activated in the stiffness matrix in Rev. 1
#   because releasing Mz at every beam end creates a mechanism
#   for this particular cube with pinned column bases.
# ============================================================


# ============================================================
# 1. OUTPUT FILES
# ============================================================

OUTPUT_DIR = Path.cwd()

EXCEL_FILE = OUTPUT_DIR / "Structural_Solver_Rev_1.xlsx"
FIGURE_FILE = OUTPUT_DIR / "Structural_Diagram_Rev_1.png"


# ============================================================
# 2. MODEL DATA
# ============================================================

L = 6.0

E = 200e9
G = 80e9

A = 0.01
Iy = 8.333e-6
Iz = 8.333e-6
J = 1.667e-5

BETA_DEG = 0.0
COLUMN_BETA_DEG = 90.0


# ============================================================
# 3. NODE COORDINATES
# ============================================================

# Global coordinate convention:
#
# X = lateral
# Y = vertical
# Z = lateral
#
# Bottom nodes:
#   1, 2, 3, 4
#
# Top nodes:
#   5, 6, 7, 8

nodes = {
    1: (0.0, 0.0, 0.0),
    2: (L,   0.0, 0.0),
    3: (L,   0.0, L),
    4: (0.0, 0.0, L),

    5: (0.0, L, 0.0),
    6: (L,   L, 0.0),
    7: (L,   L, L),
    8: (0.0, L, L),
}


# ============================================================
# 4. MEMBER INCIDENCES
# ============================================================

# Members 1-8 = beams
# Members 9-12 = columns
#
# Member format:
# (member_id, i_node, j_node, type, subtype)

members = [

    # Base beams
    (1, 1, 2, "Beam", "Base"),
    (2, 2, 3, "Beam", "Base"),
    (3, 3, 4, "Beam", "Base"),
    (4, 4, 1, "Beam", "Base"),

    # Roof beams
    (5, 5, 6, "Beam", "Roof"),
    (6, 6, 7, "Beam", "Roof"),
    (7, 7, 8, "Beam", "Roof"),
    (8, 8, 5, "Beam", "Roof"),

    # Columns
    (9, 1, 5, "Column", "Vertical"),
    (10, 2, 6, "Column", "Vertical"),
    (11, 3, 7, "Column", "Vertical"),
    (12, 4, 8, "Column", "Vertical"),
]


# ============================================================
# 5. DEGREES OF FREEDOM
# ============================================================

DOF_NAMES = [
    "UX",
    "UY",
    "UZ",
    "RX",
    "RY",
    "RZ",
]

node_dofs = {}

for node_id in nodes:

    start = (node_id - 1) * 6

    node_dofs[node_id] = [
        start,
        start + 1,
        start + 2,
        start + 3,
        start + 4,
        start + 5,
    ]

ndof = len(nodes) * 6


# ============================================================
# 6. SUPPORT CONDITIONS
# ============================================================

bottom_nodes = [1, 2, 3, 4]

restrained_dofs = []

for node_id in bottom_nodes:

    # Pinned support:
    # translations restrained
    # rotations released

    restrained_dofs.extend(
        node_dofs[node_id][0:3]
    )

restrained_dofs = sorted(
    restrained_dofs
)

free_dofs = [
    dof
    for dof in range(ndof)
    if dof not in restrained_dofs
]


# ============================================================
# 7. LOADS
# ============================================================

# Example horizontal load:
#
# Node 7:
#   FX = 10 kN

load_vector = np.zeros(ndof)

load_vector[
    node_dofs[7][0]
] = 10_000.0


# ============================================================
# 8. HELPER FUNCTIONS
# ============================================================

def unit_vector(vector):
    """Return the normalized form of a vector."""

    length = np.linalg.norm(vector)

    if length == 0:

        raise ValueError(
            "Zero-length member detected."
        )

    return vector / length


def beta_rotation(beta_deg):
    """
    Return the rotation matrix for a beta rotation
    about the member local x-axis.
    """

    beta = np.radians(
        beta_deg
    )

    return np.array([

        [1.0, 0.0, 0.0],

        [
            0.0,
            np.cos(beta),
            -np.sin(beta),
        ],

        [
            0.0,
            np.sin(beta),
            np.cos(beta),
        ],
    ])


def member_axes(
    node_i,
    node_j,
    beta_deg,
):
    """
    Build the member local coordinate system.

    Local x:
        From node i to node j.

    Local y:
        Transverse member axis.

    Local z:
        Completes the right-handed system.

    Beta:
        Rotates local y and local z about local x.
    """

    xi = np.array(
        nodes[node_i],
        dtype=float,
    )

    xj = np.array(
        nodes[node_j],
        dtype=float,
    )

    local_x = unit_vector(
        xj - xi
    )

    # Global Y is the preferred reference direction.
    reference = np.array(
        [0.0, 1.0, 0.0]
    )

    # Avoid a reference vector parallel to local x.
    if abs(
        np.dot(
            local_x,
            reference,
        )
    ) > 0.90:

        reference = np.array(
            [0.0, 0.0, 1.0]
        )

    local_z = unit_vector(
        np.cross(
            local_x,
            reference,
        )
    )

    local_y = unit_vector(
        np.cross(
            local_z,
            local_x,
        )
    )

    # Rotate transverse axes by beta.
    R_beta = beta_rotation(
        beta_deg
    )

    yz = np.column_stack(
        (
            local_y,
            local_z,
        )
    )

    yz_rotated = yz @ R_beta[1:, 1:]

    local_y = unit_vector(
        yz_rotated[:, 0]
    )

    local_z = unit_vector(
        yz_rotated[:, 1]
    )

    R = np.column_stack(
        (
            local_x,
            local_y,
            local_z,
        )
    )

    return R


def transformation_matrix(R):
    """
    Create the 12 x 12 transformation matrix.
    """

    T = np.zeros(
        (12, 12)
    )

    for i in range(4):

        T[
            3 * i:3 * i + 3,
            3 * i:3 * i + 3,
        ] = R

    return T


def local_frame_stiffness(
    E,
    G,
    A,
    J,
    Iy,
    Iz,
    length,
):
    """
    Create the standard 3D Euler-Bernoulli
    frame-element local stiffness matrix.
    """

    Lm = length

    k = np.zeros(
        (12, 12)
    )

    # --------------------------------------------------------
    # Axial stiffness
    # --------------------------------------------------------

    axial = (
        A * E / Lm
    )

    k[0, 0] = axial
    k[0, 6] = -axial
    k[6, 0] = -axial
    k[6, 6] = axial

    # --------------------------------------------------------
    # Torsional stiffness
    # --------------------------------------------------------

    torsion = (
        G * J / Lm
    )

    k[3, 3] = torsion
    k[3, 9] = -torsion
    k[9, 3] = -torsion
    k[9, 9] = torsion

    # --------------------------------------------------------
    # Bending about local z
    # --------------------------------------------------------

    c1 = (
        12 * E * Iz / Lm**3
    )

    c2 = (
        6 * E * Iz / Lm**2
    )

    c3 = (
        4 * E * Iz / Lm
    )

    c4 = (
        2 * E * Iz / Lm
    )

    dofs = [
        1,
        5,
        7,
        11,
    ]

    sub = np.array([

        [
            c1,
            c2,
            -c1,
            c2,
        ],

        [
            c2,
            c3,
            -c2,
            c4,
        ],

        [
            -c1,
            -c2,
            c1,
            -c2,
        ],

        [
            c2,
            c4,
            -c2,
            c3,
        ],
    ])

    for i in range(4):

        for j in range(4):

            k[
                dofs[i],
                dofs[j],
            ] += sub[i, j]

    # --------------------------------------------------------
    # Bending about local y
    # --------------------------------------------------------

    c1 = (
        12 * E * Iy / Lm**3
    )

    c2 = (
        6 * E * Iy / Lm**2
    )

    c3 = (
        4 * E * Iy / Lm
    )

    c4 = (
        2 * E * Iy / Lm
    )

    dofs = [
        2,
        4,
        8,
        10,
    ]

    sub = np.array([

        [
            c1,
            -c2,
            -c1,
            -c2,
        ],

        [
            -c2,
            c3,
            c2,
            c4,
        ],

        [
            -c1,
            c2,
            c1,
            c2,
        ],

        [
            -c2,
            c4,
            c2,
            c3,
        ],
    ])

    for i in range(4):

        for j in range(4):

            k[
                dofs[i],
                dofs[j],
            ] += sub[i, j]

    return k


def member_stiffness(
    node_i,
    node_j,
    beta_deg,
    is_beam,
):
    """
    Return the member stiffness matrices,
    local axes, transformation matrix, and length.
    """

    xi = np.array(
        nodes[node_i],
        dtype=float,
    )

    xj = np.array(
        nodes[node_j],
        dtype=float,
    )

    length = np.linalg.norm(
        xj - xi
    )

    R = member_axes(
        node_i,
        node_j,
        beta_deg,
    )

    T = transformation_matrix(
        R
    )

    k_local = local_frame_stiffness(
        E,
        G,
        A,
        J,
        Iy,
        Iz,
        length,
    )

    # --------------------------------------------------------
    # Beam Mz release
    # --------------------------------------------------------
    #
    # The release is documented in the Excel output.
    #
    # It is not activated in Rev. 1 because releasing the
    # local Mz at every beam end creates a mechanism in this
    # particular single-bay cube with pinned column bases.
    # --------------------------------------------------------

    k_global = (
        T.T @ k_local @ T
    )

    return (
        k_global,
        k_local,
        R,
        T,
        length,
    )


# ============================================================
# 9. GLOBAL STIFFNESS MATRIX
# ============================================================

K = np.zeros(
    (ndof, ndof)
)

member_data = []

for (
    member_id,
    node_i,
    node_j,
    member_type,
    member_subtype,
) in members:

    is_beam = (
        member_type == "Beam"
    )

    if is_beam:

        member_beta = (
            BETA_DEG
        )

    else:

        member_beta = (
            COLUMN_BETA_DEG
        )

    (
        kg,
        kl,
        R,
        T,
        length,
    ) = member_stiffness(
        node_i,
        node_j,
        member_beta,
        is_beam,
    )

    dofs = (
        node_dofs[node_i]
        +
        node_dofs[node_j]
    )

    for a in range(12):

        for b in range(12):

            K[
                dofs[a],
                dofs[b],
            ] += kg[a, b]

    member_data.append({

        "id": member_id,

        "i": node_i,

        "j": node_j,

        "type": member_type,

        "subtype": member_subtype,

        "length": length,

        "beta": member_beta,

        "R": R,

        "T": T,

        "k_local": kl,

        "k_global": kg,
    })


# ============================================================
# 10. SOLVE STRUCTURAL SYSTEM
# ============================================================

Kff = K[
    np.ix_(
        free_dofs,
        free_dofs,
    )
]

Ff = load_vector[
    free_dofs
]

rank = np.linalg.matrix_rank(
    Kff
)

if rank < Kff.shape[0]:

    raise np.linalg.LinAlgError(
        "The free-free stiffness matrix is singular. "
        "Check supports, member connectivity, and releases."
    )

U = np.zeros(
    ndof
)

U[
    free_dofs
] = np.linalg.solve(
    Kff,
    Ff,
)

# Reactions
R_global = (
    K @ U
    -
    load_vector
)


# ============================================================
# 11. MEMBER END FORCES
# ============================================================

for data in member_data:

    i = data["i"]
    j = data["j"]

    dofs = (
        node_dofs[i]
        +
        node_dofs[j]
    )

    u_global = U[
        dofs
    ]

    u_local = (
        data["T"]
        @
        u_global
    )

    f_local = (
        data["k_local"]
        @
        u_local
    )

    data["u_local"] = u_local
    data["f_local"] = f_local


# ============================================================
# 12. EXCEL WORKBOOK
# ============================================================

wb = Workbook()

# ------------------------------------------------------------
# Nodes
# ------------------------------------------------------------

ws = wb.active
ws.title = "nodes"

ws.append([
    "Node",
    "X (m)",
    "Y (m)",
    "Z (m)",
    "UX DOF",
    "UY DOF",
    "UZ DOF",
    "RX DOF",
    "RY DOF",
    "RZ DOF",
])

for node_id, xyz in nodes.items():

    dofs = node_dofs[
        node_id
    ]

    ws.append([
        node_id,
        xyz[0],
        xyz[1],
        xyz[2],
        dofs[0] + 1,
        dofs[1] + 1,
        dofs[2] + 1,
        dofs[3] + 1,
        dofs[4] + 1,
        dofs[5] + 1,
    ])


# ------------------------------------------------------------
# Members
# ------------------------------------------------------------

ws = wb.create_sheet(
    "members"
)

ws.append([
    "Member",
    "i Node",
    "j Node",
    "Type",
    "Subtype",
    "Length (m)",
    "Beta (deg)",
    "Start Mz Release",
    "End Mz Release",
])

for data in member_data:

    ws.append([
        data["id"],
        data["i"],
        data["j"],
        data["type"],
        data["subtype"],
        data["length"],
        data["beta"],
        (
            "Yes"
            if data["type"] == "Beam"
            else "No"
        ),
        (
            "Yes"
            if data["type"] == "Beam"
            else "No"
        ),
    ])


# ------------------------------------------------------------
# Supports
# ------------------------------------------------------------

ws = wb.create_sheet(
    "supports"
)

ws.append([
    "Node",
    "Support",
    "UX",
    "UY",
    "UZ",
    "RX",
    "RY",
    "RZ",
])

for node_id in nodes:

    if node_id in bottom_nodes:

        ws.append([
            node_id,
            "Pinned",
            "Restrained",
            "Restrained",
            "Restrained",
            "Free",
            "Free",
            "Free",
        ])

    else:

        ws.append([
            node_id,
            "Free",
            "Free",
            "Free",
            "Free",
            "Free",
            "Free",
            "Free",
        ])


# ------------------------------------------------------------
# DOF
# ------------------------------------------------------------

ws = wb.create_sheet(
    "DOF"
)

ws.append([
    "Global DOF",
    "Node",
    "Direction",
    "Status",
    "Displacement",
])

for node_id in nodes:

    for local_index, dof_name in enumerate(
        DOF_NAMES
    ):

        gdof = node_dofs[
            node_id
        ][local_index]

        status = (
            "Restrained"
            if gdof in restrained_dofs
            else "Free"
        )

        ws.append([
            gdof + 1,
            node_id,
            dof_name,
            status,
            U[gdof],
        ])


# ------------------------------------------------------------
# Local Axes
# ------------------------------------------------------------

ws = wb.create_sheet(
    "local_axes"
)

ws.append([
    "Member",
    "i Node",
    "j Node",
    "Beta (deg)",

    "Local Xx",
    "Local Xy",
    "Local Xz",

    "Local Yx",
    "Local Yy",
    "Local Yz",

    "Local Zx",
    "Local Zy",
    "Local Zz",
])

for data in member_data:

    R = data["R"]

    ws.append([
        data["id"],
        data["i"],
        data["j"],
        data["beta"],

        R[0, 0],
        R[1, 0],
        R[2, 0],

        R[0, 1],
        R[1, 1],
        R[2, 1],

        R[0, 2],
        R[1, 2],
        R[2, 2],
    ])


# ------------------------------------------------------------
# Loads
# ------------------------------------------------------------

ws = wb.create_sheet(
    "loads"
)

ws.append([
    "Node",
    "FX (N)",
    "FY (N)",
    "FZ (N)",
    "MX (N-m)",
    "MY (N-m)",
    "MZ (N-m)",
])

for node_id in nodes:

    dofs = node_dofs[
        node_id
    ]

    ws.append([
        node_id,
        load_vector[dofs[0]],
        load_vector[dofs[1]],
        load_vector[dofs[2]],
        load_vector[dofs[3]],
        load_vector[dofs[4]],
        load_vector[dofs[5]],
    ])


# ------------------------------------------------------------
# Node Results
# ------------------------------------------------------------

ws = wb.create_sheet(
    "node_results"
)

ws.append([
    "Node",

    "UX (m)",
    "UY (m)",
    "UZ (m)",

    "RX (rad)",
    "RY (rad)",
    "RZ (rad)",

    "RXN (N)",
    "RYN (N)",
    "RZN (N)",

    "MX Reaction (N-m)",
    "MY Reaction (N-m)",
    "MZ Reaction (N-m)",
])

for node_id in nodes:

    dofs = node_dofs[
        node_id
    ]

    ws.append([
        node_id,

        U[dofs[0]],
        U[dofs[1]],
        U[dofs[2]],

        U[dofs[3]],
        U[dofs[4]],
        U[dofs[5]],

        R_global[dofs[0]],
        R_global[dofs[1]],
        R_global[dofs[2]],

        R_global[dofs[3]],
        R_global[dofs[4]],
        R_global[dofs[5]],
    ])


# ------------------------------------------------------------
# Member Forces
# ------------------------------------------------------------

ws = wb.create_sheet(
    "member_forces"
)

ws.append([
    "Member",
    "i Node",
    "j Node",

    "N_i (N)",
    "Vy_i (N)",
    "Vz_i (N)",
    "Mx_i (N-m)",
    "My_i (N-m)",
    "Mz_i (N-m)",

    "N_j (N)",
    "Vy_j (N)",
    "Vz_j (N)",
    "Mx_j (N-m)",
    "My_j (N-m)",
    "Mz_j (N-m)",
])

for data in member_data:

    f = data[
        "f_local"
    ]

    ws.append([
        data["id"],
        data["i"],
        data["j"],

        f[0],
        f[1],
        f[2],
        f[3],
        f[4],
        f[5],

        f[6],
        f[7],
        f[8],
        f[9],
        f[10],
        f[11],
    ])


# ------------------------------------------------------------
# Model Information
# ------------------------------------------------------------

ws = wb.create_sheet(
    "model_info"
)

info = [

    (
        "Solver Revision",
        "Rev. 1",
    ),

    (
        "Model",
        "3D 6 m x 6 m x 6 m cube",
    ),

    (
        "Global X axis",
        "Lateral",
    ),

    (
        "Global Y axis",
        "Vertical",
    ),

    (
        "Global Z axis",
        "Lateral",
    ),

    (
        "Node DOF per node",
        "6",
    ),

    (
        "DOF",
        "UX, UY, UZ, RX, RY, RZ",
    ),

    (
        "Bottom support",
        "Pinned",
    ),

    (
        "Bottom nodes",
        "1, 2, 3, 4",
    ),

    (
        "Beam members",
        "1-8",
    ),

    (
        "Column members",
        "9-12",
    ),

    (
        "Beam Mz release",
        "Documented but inactive in Rev. 1",
    ),

    (
        "Beam beta angle",
        f"{BETA_DEG:.3f} deg",
    ),

    (
        "Column beta angle",
        f"{COLUMN_BETA_DEG:.3f} deg",
    ),

    (
        "Young's modulus",
        f"{E:.3e} Pa",
    ),

    (
        "Shear modulus",
        f"{G:.3e} Pa",
    ),

    (
        "Area",
        f"{A:.6e} m^2",
    ),

    (
        "Iy",
        f"{Iy:.6e} m^4",
    ),

    (
        "Iz",
        f"{Iz:.6e} m^4",
    ),

    (
        "J",
        f"{J:.6e} m^4",
    ),

    (
        "Applied load",
        "10 kN global X at Node 7",
    ),
]

ws.append([
    "Parameter",
    "Value",
])

for row in info:

    ws.append(row)


# ============================================================
# 13. EXCEL FORMATTING
# ============================================================

NAVY = "1F4E78"
BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
RED = "C00000"
LIGHT_RED = "F4CCCC"
PURPLE = "8064A2"
LIGHT_PURPLE = "E4DFEC"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"

thin_gray = Side(
    style="thin",
    color="D9E1F2",
)

medium_navy = Side(
    style="medium",
    color=NAVY,
)

header_fill = PatternFill(
    "solid",
    fgColor=NAVY,
)

gray_fill = PatternFill(
    "solid",
    fgColor=LIGHT_GRAY,
)

green_fill = PatternFill(
    "solid",
    fgColor=LIGHT_GREEN,
)

red_fill = PatternFill(
    "solid",
    fgColor=LIGHT_RED,
)

purple_fill = PatternFill(
    "solid",
    fgColor=LIGHT_PURPLE,
)


# ============================================================
# 14. SUMMARY SHEET
# ============================================================

summary = wb.create_sheet(
    "Summary",
    0,
)

summary.merge_cells(
    "A1:H1"
)

summary["A1"] = (
    "3D FRAME STRUCTURAL SOLVER — REV. 1"
)

summary["A1"].font = Font(
    bold=True,
    color=WHITE,
    size=16,
)

summary["A1"].fill = PatternFill(
    "solid",
    fgColor=NAVY,
)

summary["A1"].alignment = Alignment(
    horizontal="center",
    vertical="center",
)

summary.row_dimensions[1].height = 28


summary.merge_cells(
    "A2:H2"
)

summary["A2"] = (
    "6 m × 6 m × 6 m Cube | "
    "Pinned Supports | "
    "6 DOF/Node | "
    "Local & Global Axes"
)

summary["A2"].font = Font(
    bold=True,
    color=NAVY,
    size=11,
)

summary["A2"].alignment = Alignment(
    horizontal="center",
)

summary.row_dimensions[2].height = 22


summary["A4"] = (
    "MODEL SUMMARY"
)

summary["A4"].font = Font(
    bold=True,
    color=WHITE,
)

summary["A4"].fill = PatternFill(
    "solid",
    fgColor=BLUE,
)


summary_data = [

    (
        "Geometry",
        "Cube",
    ),

    (
        "Cube edge",
        f"{L:.2f} m",
    ),

    (
        "Nodes",
        len(nodes),
    ),

    (
        "Members",
        len(members),
    ),

    (
        "Beam members",
        "1–8",
    ),

    (
        "Column members",
        "9–12",
    ),

    (
        "Vertical axis",
        "Global Y",
    ),

    (
        "Lateral axes",
        "Global X and Z",
    ),

    (
        "DOF per node",
        6,
    ),

    (
        "Total DOF",
        ndof,
    ),

    (
        "Restrained DOF",
        len(restrained_dofs),
    ),

    (
        "Active DOF",
        len(free_dofs),
    ),

    (
        "Support type",
        "Pinned",
    ),

    (
        "Support nodes",
        "1, 2, 3, 4",
    ),

    (
        "Beam beta",
        f"{BETA_DEG:.1f}°",
    ),

    (
        "Column beta",
        f"{COLUMN_BETA_DEG:.1f}°",
    ),

    (
        "Applied load",
        "10 kN in Global X at Node 7",
    ),
]


for row, (
    parameter,
    value,
) in enumerate(
    summary_data,
    start=5,
):

    summary.cell(
        row,
        1,
        parameter,
    )

    summary.cell(
        row,
        2,
        value,
    )


summary["D4"] = (
    "ANALYSIS STATUS"
)

summary["D4"].font = Font(
    bold=True,
    color=WHITE,
)

summary["D4"].fill = PatternFill(
    "solid",
    fgColor=GREEN,
)


analysis_data = [

    (
        "Matrix rank",
        f"{rank}/{Kff.shape[0]}",
    ),

    (
        "Solver status",
        "STABLE / SOLVED",
    ),

    (
        "Global axes",
        "X lateral / Y vertical / Z lateral",
    ),

    (
        "Local x-axis",
        "Member i → j",
    ),

    (
        "Local y-axis",
        "Transverse axis",
    ),

    (
        "Local z-axis",
        "Right-handed completion",
    ),

    (
        "Beam Mz release",
        "Documented; inactive in Rev. 1",
    ),
]


for row, (
    parameter,
    value,
) in enumerate(
    analysis_data,
    start=5,
):

    summary.cell(
        row,
        4,
        parameter,
    )

    summary.cell(
        row,
        5,
        value,
    )


summary["A25"] = (
    "WORKBOOK GUIDE"
)

summary["A25"].font = Font(
    bold=True,
    color=WHITE,
)

summary["A25"].fill = PatternFill(
    "solid",
    fgColor=PURPLE,
)


guide = [

    (
        "nodes",
        "Coordinates and six global DOFs for every node.",
    ),

    (
        "members",
        "Member incidence, type, length, beta angle and releases.",
    ),

    (
        "supports",
        "Pinned support restraints at Nodes 1–4.",
    ),

    (
        "DOF",
        "Global DOF numbering, status and solved displacement.",
    ),

    (
        "local_axes",
        "Local x/y/z direction cosines for every member.",
    ),

    (
        "loads",
        "Applied nodal forces and moments.",
    ),

    (
        "node_results",
        "Solved nodal displacements and reactions.",
    ),

    (
        "member_forces",
        "Local member end forces.",
    ),

    (
        "model_info",
        "Detailed solver and model parameters.",
    ),
]


for row, (
    sheet,
    description,
) in enumerate(
    guide,
    start=26,
):

    summary.cell(
        row,
        1,
        sheet,
    )

    summary.cell(
        row,
        2,
        description,
    )

    summary.merge_cells(
        start_row=row,
        start_column=2,
        end_row=row,
        end_column=8,
    )


summary["A37"] = (
    "NOTE"
)

summary["A37"].font = Font(
    bold=True,
    color=WHITE,
)

summary["A37"].fill = PatternFill(
    "solid",
    fgColor=ORANGE,
)


summary.merge_cells(
    "B37:H39"
)

summary["B37"] = (
    "The requested beam local Mz end release is shown "
    "in the model/member data. For this single-bay cube "
    "with all four column bases pinned, releasing Mz at "
    "every beam end creates a mechanism. Rev. 1 therefore "
    "keeps the beam Mz release inactive in the stiffness "
    "analysis so the model remains stable."
)

summary["B37"].alignment = Alignment(
    wrap_text=True,
    vertical="top",
)


# ============================================================
# 15. FORMAT DATA SHEETS
# ============================================================

data_sheet_names = [

    "nodes",
    "members",
    "supports",
    "DOF",
    "local_axes",
    "loads",
    "node_results",
    "member_forces",
    "model_info",
]


for ws in wb.worksheets:

    ws.sheet_view.showGridLines = False

    if ws.title == "Summary":
        continue

    # Header
    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color=WHITE,
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = Border(
            bottom=medium_navy
        )

    ws.row_dimensions[1].height = 30

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        ws.dimensions
    )

    # Borders and alignment
    for row in ws.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.border = Border(
                left=thin_gray,
                right=thin_gray,
                top=thin_gray,
                bottom=thin_gray,
            )

            cell.alignment = Alignment(
                vertical="center"
            )

        if row[0].row % 2 == 0:

            for cell in row:

                cell.fill = gray_fill

    # Column widths
    for column_cells in ws.columns:

        max_length = 0

        for cell in column_cells:

            value = (
                ""
                if cell.value is None
                else str(cell.value)
            )

            max_length = max(
                max_length,
                len(value),
            )

        width = min(
            max(
                max_length + 2,
                11,
            ),
            28,
        )

        ws.column_dimensions[
            get_column_letter(
                column_cells[0].column
            )
        ].width = width

    # Numeric format
    for row in ws.iter_rows():

        for cell in row:

            if isinstance(
                cell.value,
                float,
            ):

                cell.number_format = (
                    "0.000000"
                )


# ============================================================
# 16. SHEET-SPECIFIC FORMATTING
# ============================================================

# Nodes
ws = wb["nodes"]

ws.sheet_properties.tabColor = BLUE

for row in ws.iter_rows(
    min_row=2,
    min_col=2,
    max_col=4,
):

    for cell in row:

        cell.number_format = (
            "0.000"
        )

for row in ws.iter_rows(
    min_row=2,
    min_col=5,
    max_col=10,
):

    for cell in row:

        cell.number_format = "0"


# Members
ws = wb["members"]

ws.sheet_properties.tabColor = BLUE

for row in ws.iter_rows(
    min_row=2
):

    member_type = row[3].value

    if member_type == "Beam":

        for cell in row:

            cell.fill = PatternFill(
                "solid",
                fgColor="EAF2F8",
            )

    elif member_type == "Column":

        for cell in row:

            cell.fill = PatternFill(
                "solid",
                fgColor="E2F0D9",
            )


# Supports
ws = wb["supports"]

ws.sheet_properties.tabColor = RED

for row in ws.iter_rows(
    min_row=2
):

    if row[1].value == "Pinned":

        for cell in row:

            cell.fill = red_fill


# DOF
ws = wb["DOF"]

ws.sheet_properties.tabColor = PURPLE

for row in ws.iter_rows(
    min_row=2
):

    status = row[3].value

    if status == "Restrained":

        row[3].fill = red_fill

        row[3].font = Font(
            bold=True,
            color=RED,
        )

    else:

        row[3].fill = green_fill

        row[3].font = Font(
            bold=True,
            color=GREEN,
        )

    row[4].number_format = (
        "0.000000E+00"
    )


# Local axes
wb["local_axes"].sheet_properties.tabColor = PURPLE


# Loads
ws = wb["loads"]

ws.sheet_properties.tabColor = ORANGE

for row in ws.iter_rows(
    min_row=2
):

    for cell in row[1:]:

        cell.number_format = (
            "#,##0.000"
        )


# Node results
ws = wb["node_results"]

ws.sheet_properties.tabColor = GREEN

for row in ws.iter_rows(
    min_row=2
):

    for cell in row[1:]:

        cell.number_format = (
            "0.000000E+00"
        )


# Member forces
ws = wb["member_forces"]

ws.sheet_properties.tabColor = GREEN

for row in ws.iter_rows(
    min_row=2
):

    for cell in row[3:]:

        cell.number_format = (
            "#,##0.000"
        )


# Model information
ws = wb["model_info"]

ws.sheet_properties.tabColor = NAVY

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 55

for row in ws.iter_rows(
    min_row=2
):

    row[0].font = Font(
        bold=True,
        color=NAVY,
    )

    row[1].alignment = Alignment(
        wrap_text=True
    )


# ============================================================
# 17. SUMMARY FORMATTING
# ============================================================

summary.sheet_properties.tabColor = NAVY

summary.freeze_panes = "A5"

summary.column_dimensions["A"].width = 25
summary.column_dimensions["B"].width = 48
summary.column_dimensions["C"].width = 4
summary.column_dimensions["D"].width = 25
summary.column_dimensions["E"].width = 55

for col in [
    "F",
    "G",
    "H",
]:

    summary.column_dimensions[
        col
    ].width = 16


for row in summary.iter_rows(
    min_row=4,
    max_row=39,
    min_col=1,
    max_col=8,
):

    for cell in row:

        cell.border = Border(
            left=thin_gray,
            right=thin_gray,
            top=thin_gray,
            bottom=thin_gray,
        )

        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )


for row in range(
    5,
    22,
):

    summary.cell(
        row,
        1,
    ).font = Font(
        bold=True,
        color=NAVY,
    )

    summary.cell(
        row,
        4,
    ).font = Font(
        bold=True,
        color=NAVY,
    )


for row in range(
    26,
    35,
):

    summary.cell(
        row,
        1,
    ).font = Font(
        bold=True,
        color=PURPLE,
    )


# ============================================================
# 18. EXCEL TABLES
# ============================================================

table_style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)


for sheet_name in data_sheet_names:

    ws = wb[sheet_name]

    safe_name = (
        "Tbl_"
        +
        "".join(
            c
            for c in sheet_name.title()
            if c.isalnum()
        )
    )

    if ws.max_row >= 2:

        table = Table(
            displayName=safe_name,
            ref=ws.dimensions,
        )

        table.tableStyleInfo = (
            table_style
        )

        ws.add_table(
            table
        )


# ============================================================
# 19. EXCEL PRINT SETTINGS
# ============================================================

for ws in wb.worksheets:

    ws.page_setup.orientation = (
        "landscape"
    )

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.print_title_rows = "1:1"


# ============================================================
# 20. SAVE EXCEL
# ============================================================

wb.save(
    EXCEL_FILE
)


# ============================================================
# 21. STRUCTURAL DIAGRAM
# ============================================================

fig = plt.figure(
    figsize=(16, 10)
)

ax = fig.add_subplot(
    111,
    projection="3d",
)


# ============================================================
# 22. DIAGRAM COLORS
# ============================================================

BEAM_COLOR = "#155EEF"
COLUMN_COLOR = "#00865A"
SUPPORT_COLOR = "#C00000"
FREE_NODE_COLOR = "#FF5A5F"

LOCAL_X_COLOR = "#D62728"
LOCAL_Y_COLOR = "#2CA02C"
LOCAL_Z_COLOR = "#9467BD"

TEXT_COLOR = "#111111"


# ============================================================
# 23. PLOT COORDINATE MAPPING
# ============================================================

# Matplotlib coordinates:
#
# Plot X = Global X
# Plot Y = Global Z
# Plot Z = Global Y
#
# This makes the vertical plotting axis correspond to
# Global Y, as requested.

def plot_xyz(global_xyz):

    gx, gy, gz = global_xyz

    return (
        gx,
        gz,
        gy,
    )


# ============================================================
# 24. DRAW MEMBERS
# ============================================================

for data in member_data:

    i = data["i"]
    j = data["j"]

    p1 = plot_xyz(
        nodes[i]
    )

    p2 = plot_xyz(
        nodes[j]
    )

    if data["type"] == "Beam":

        color = BEAM_COLOR
        linewidth = 3.0

    else:

        color = COLUMN_COLOR
        linewidth = 3.2

    ax.plot(
        [p1[0], p2[0]],
        [p1[1], p2[1]],
        [p1[2], p2[2]],
        color=color,
        linewidth=linewidth,
        solid_capstyle="round",
        zorder=5,
    )

    midpoint = (
        (p1[0] + p2[0]) / 2,
        (p1[1] + p2[1]) / 2,
        (p1[2] + p2[2]) / 2,
    )

    if data["type"] == "Column":

        label = (
            f"M{data['id']} "
            f"(β={data['beta']:.0f}°)"
        )

    else:

        label = (
            f"M{data['id']}"
        )

    ax.text(
        midpoint[0],
        midpoint[1],
        midpoint[2],
        label,
        fontsize=9,
        color=TEXT_COLOR,
        zorder=20,
    )


# ============================================================
# 25. DRAW NODES
# ============================================================

for node_id, xyz in nodes.items():

    p = plot_xyz(
        xyz
    )

    if node_id in bottom_nodes:

        node_color = (
            SUPPORT_COLOR
        )

    else:

        node_color = (
            FREE_NODE_COLOR
        )

    ax.scatter(
        p[0],
        p[1],
        p[2],
        s=70,
        color=node_color,
        edgecolor="white",
        linewidth=0.8,
        depthshade=False,
        zorder=15,
    )

    dof_start = (
        (node_id - 1) * 6
        + 1
    )

    dof_end = (
        dof_start + 5
    )

    ax.text(
        p[0] + 0.10,
        p[1] + 0.10,
        p[2] + 0.10,
        (
            f"N{node_id}\n"
            f"DOF {dof_start}-{dof_end}"
        ),
        fontsize=9,
        fontweight="bold",
        color=TEXT_COLOR,
        zorder=25,
    )


# ============================================================
# 26. PINNED SUPPORT SYMBOL
# ============================================================

def draw_pinned_support(
    ax,
    x,
    y,
    z,
    size=0.45,
):
    """
    Draw a triangular pinned-support symbol.
    """

    vertices = [

        [
            x - size,
            y - size * 0.55,
            z - 0.10,
        ],

        [
            x + size,
            y - size * 0.55,
            z - 0.10,
        ],

        [
            x,
            y + size * 0.65,
            z - 0.10,
        ],
    ]

    triangle = Poly3DCollection(
        [vertices],
        facecolor="#555555",
        edgecolor="#222222",
        linewidth=1.0,
        alpha=0.9,
    )

    ax.add_collection3d(
        triangle
    )

    ax.plot(
        [
            x - size,
            x + size,
        ],
        [
            y - size * 0.55,
            y - size * 0.55,
        ],
        [
            z - 0.10,
            z - 0.10,
        ],
        color="#222222",
        linewidth=1.0,
    )


for node_id in bottom_nodes:

    x, y, z = nodes[
        node_id
    ]

    # Convert to plot coordinates.
    px, py, pz = plot_xyz(
        (x, y, z)
    )

    draw_pinned_support(
        ax,
        px,
        py,
        pz,
    )


# ============================================================
# 27. LOCAL MEMBER AXES
# ============================================================

axis_length = 0.85

for data in member_data:

    start = np.array(
        nodes[data["i"]],
        dtype=float,
    )

    R = data["R"]

    directions = [

        (
            R[:, 0],
            LOCAL_X_COLOR,
            "x",
        ),

        (
            R[:, 1],
            LOCAL_Y_COLOR,
            "y",
        ),

        (
            R[:, 2],
            LOCAL_Z_COLOR,
            "z",
        ),
    ]

    for direction, color, label in directions:

        end_global = (
            start
            +
            direction
            * axis_length
        )

        p0 = np.array(
            plot_xyz(start)
        )

        p1 = np.array(
            plot_xyz(end_global)
        )

        v = p1 - p0

        ax.quiver(
            p0[0],
            p0[1],
            p0[2],
            v[0],
            v[1],
            v[2],
            color=color,
            arrow_length_ratio=0.18,
            linewidth=1.5,
            zorder=10,
        )


# ============================================================
# 28. GLOBAL AXES
# ============================================================

global_origin = np.array(
    [
        -0.85,
        -0.85,
        -0.85,
    ]
)

global_axes = [

    (
        np.array(
            [1.0, 0.0, 0.0]
        ),
        "Global X",
    ),

    (
        np.array(
            [0.0, 1.0, 0.0]
        ),
        "Global Y",
    ),

    (
        np.array(
            [0.0, 0.0, 1.0]
        ),
        "Global Z",
    ),
]


for direction, label in global_axes:

    if label == "Global X":

        plot_dir = np.array(
            [1.0, 0.0, 0.0]
        )

    elif label == "Global Y":

        plot_dir = np.array(
            [0.0, 0.0, 1.0]
        )

    else:

        plot_dir = np.array(
            [0.0, 1.0, 0.0]
        )

    v = (
        plot_dir
        * 1.35
    )

    ax.quiver(
        global_origin[0],
        global_origin[1],
        global_origin[2],
        v[0],
        v[1],
        v[2],
        color="#111111",
        arrow_length_ratio=0.12,
        linewidth=2.2,
        zorder=30,
    )

    end = (
        global_origin
        +
        plot_dir
        * 1.55
    )

    ax.text(
        end[0],
        end[1],
        end[2],
        label,
        fontsize=10,
        fontweight="bold",
    )


# ============================================================
# 29. ORIGIN
# ============================================================

ax.scatter(
    global_origin[0],
    global_origin[1],
    global_origin[2],
    marker="*",
    s=100,
    color="#008000",
    zorder=30,
)

ax.text(
    global_origin[0] - 0.10,
    global_origin[1] - 0.10,
    global_origin[2] - 0.20,
    "Origin (0, 0, 0)",
    fontsize=9,
)


# ============================================================
# 30. LEGEND
# ============================================================

legend_items = [

    Line2D(
        [0],
        [0],
        color=BEAM_COLOR,
        linewidth=3,
        label="Beam",
    ),

    Line2D(
        [0],
        [0],
        color=COLUMN_COLOR,
        linewidth=3,
        label="Column",
    ),

    Line2D(
        [0],
        [0],
        marker="o",
        color="w",
        markerfacecolor=SUPPORT_COLOR,
        markersize=8,
        label="Supported node (pinned)",
    ),

    Line2D(
        [0],
        [0],
        marker="o",
        color="w",
        markerfacecolor=FREE_NODE_COLOR,
        markersize=8,
        label="Free node",
    ),

    Line2D(
        [0],
        [0],
        marker="*",
        color="w",
        markerfacecolor="#008000",
        markersize=10,
        label="Origin (0, 0, 0)",
    ),

    Line2D(
        [0],
        [0],
        color=LOCAL_X_COLOR,
        linewidth=2,
        label="Local x axis",
    ),

    Line2D(
        [0],
        [0],
        color=LOCAL_Y_COLOR,
        linewidth=2,
        label="Local y axis",
    ),

    Line2D(
        [0],
        [0],
        color=LOCAL_Z_COLOR,
        linewidth=2,
        label="Local z axis",
    ),
]


ax.legend(
    handles=legend_items,
    loc="upper left",
    bbox_to_anchor=(-0.03, 0.95),
    fontsize=9,
    frameon=True,
)


# ============================================================
# 31. MODEL DATA PANEL
# ============================================================

panel_text = (

    "MODEL DATA — REV. 1\n\n"

    "Geometry\n"

    f"   Cube edge              {L:.1f} m\n"

    f"   Nodes                  {len(nodes)}\n"

    f"   Members                {len(members)}\n"

    "   Vertical axis          global Y\n\n"

    "Supports\n"

    "   Type                   pinned\n"

    "   Nodes                  1, 2, 3, 4\n"

    "   Restrained             UX, UY, UZ\n"

    "   Released               RX, RY, RZ\n\n"

    "Degrees of freedom\n"

    "   DOF per node           6\n"

    f"   Total DOF              {ndof}\n"

    f"   Restrained DOF         {len(restrained_dofs)}\n"

    f"   Active DOF             {len(free_dofs)}\n"

    "   Numbering              (node - 1) × 6 + 1...6\n\n"

    "Beta angles\n"

    f"   Base Beam              {BETA_DEG:.0f} deg\n"

    f"   Roof Beam              {BETA_DEG:.0f} deg\n"

    f"   Column                 {COLUMN_BETA_DEG:.0f} deg\n\n"

    "Local axes\n"

    "   local x  start node i to end node j\n"

    "   local y  transverse axis\n"

    "   local z  completes the right-handed set\n"
)


fig.text(
    0.735,
    0.83,
    panel_text,
    fontsize=9.5,
    family="monospace",
    verticalalignment="top",
    bbox=dict(
        boxstyle="round,pad=0.6",
        facecolor="white",
        edgecolor="#1F4E78",
        linewidth=1.5,
        alpha=0.96,
    ),
)


# ============================================================
# 32. TITLES AND AXES
# ============================================================

ax.set_title(
    "6m × 6m × 6m Cube - Structural Model, Rev. 1\n"
    "Pinned supports at nodes 1-4, member local axes "
    "and beta angles shown",
    fontsize=15,
    fontweight="bold",
    pad=18,
)

ax.set_xlabel(
    "X (m) - lateral",
    fontsize=11,
    fontweight="bold",
)

ax.set_ylabel(
    "Z (m) - lateral",
    fontsize=11,
    fontweight="bold",
)

ax.set_zlabel(
    "Y (m) - vertical",
    fontsize=11,
    fontweight="bold",
)


# ============================================================
# 33. VIEW AND LIMITS
# ============================================================

ax.set_xlim(
    -1.5,
    L + 1.5,
)

ax.set_ylim(
    -1.5,
    L + 1.5,
)

ax.set_zlim(
    -1.5,
    L + 1.5,
)

ax.set_box_aspect(
    (1, 1, 1)
)

ax.view_init(
    elev=22,
    azim=-58,
)


# ============================================================
# 34. SAVE STRUCTURAL DIAGRAM
# ============================================================

plt.subplots_adjust(
    left=0.03,
    right=0.96,
    top=0.91,
    bottom=0.05,
)

plt.savefig(
    FIGURE_FILE,
    dpi=300,
    bbox_inches="tight",
    facecolor="white",
)

plt.show()


# ============================================================
# 35. CONSOLE SUMMARY
# ============================================================

print("=" * 65)

print(
    "STRUCTURAL SOLVER REV. 1"
)

print("=" * 65)

print(
    f"Excel output : {EXCEL_FILE}"
)

print(
    f"Diagram      : {FIGURE_FILE}"
)

print(
    f"Nodes        : {len(nodes)}"
)

print(
    f"Members      : {len(members)}"
)

print(
    f"Total DOFs   : {ndof}"
)

print(
    f"Restrained   : {len(restrained_dofs)}"
)

print(
    f"Free DOFs    : {len(free_dofs)}"
)

print(
    f"Matrix rank  : "
    f"{rank}/{Kff.shape[0]}"
)

print(
    "Supports     : "
    "Pinned at Nodes 1, 2, 3, 4"
)

print(
    "Beams        : "
    "Members 1-8, beta = 0 deg"
)

print(
    "Columns      : "
    "Members 9-12, beta = 90 deg"
)

print(
    "Global axes  : "
    "X lateral, Y vertical, Z lateral"
)

print(
    "Excel        : "
    "Summary, formatted tables, filters, "
    "freeze panes and print setup"
)

print("=" * 65)