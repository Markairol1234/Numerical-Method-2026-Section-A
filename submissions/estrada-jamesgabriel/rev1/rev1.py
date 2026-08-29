import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# 1. DEFINE NODES, SUPPORTS & DOFs
# ==========================================
nodes = {
    'NodeID': [1, 2, 3, 4, 5, 6, 7, 8],
    'X': [0, 6, 6, 0, 0, 6, 6, 0],
    'Y': [0, 0, 0, 0, 6, 6, 6, 6],
    'Z': [0, 0, 6, 6, 0, 0, 6, 6],
    'Support': ['Pinned', 'Pinned', 'Pinned', 'Pinned', 'Free', 'Free', 'Free', 'Free'],
    'DOF_Start': [1, 7, 13, 19, 25, 31, 37, 43],
    'DOF_End': [6, 12, 18, 24, 30, 36, 42, 48]
}
df_nodes = pd.DataFrame(nodes)

# ==========================================
# 2. DEFINE MEMBERS, BETA ANGLES & END RELEASES
# ==========================================
members = [
    # Bottom Frame Members (M1 to M4) - Mz released
    {'Member': 'M1', 'i': 1, 'j': 2, 'BetaAngle': 0.0, 'Release_i': 'Mz', 'Release_j': 'Mz'},
    {'Member': 'M2', 'i': 2, 'j': 3, 'BetaAngle': 0.0, 'Release_i': 'None', 'Release_j': 'None'},
    {'Member': 'M3', 'i': 3, 'j': 4, 'BetaAngle': 0.0, 'Release_i': 'Mz', 'Release_j': 'Mz'},
    {'Member': 'M4', 'i': 4, 'j': 1, 'BetaAngle': 0.0, 'Release_i': 'None', 'Release_j': 'None'},
    # Top Frame Members (M5 to M8) - Mz released
    {'Member': 'M5', 'i': 5, 'j': 6, 'BetaAngle': 0.0, 'Release_i': 'Mz', 'Release_j': 'Mz'},
    {'Member': 'M6', 'i': 6, 'j': 7, 'BetaAngle': 0.0, 'Release_i': 'None', 'Release_j': 'None'},
    {'Member': 'M7', 'i': 7, 'j': 8, 'BetaAngle': 0.0, 'Release_i': 'Mz', 'Release_j': 'Mz'},
    {'Member': 'M8', 'i': 8, 'j': 5, 'BetaAngle': 0.0, 'Release_i': 'None', 'Release_j': 'None'},
    # Vertical Columns (M9 to M12) - Beta angle 90 deg
    {'Member': 'M9', 'i': 1, 'j': 5, 'BetaAngle': 90.0, 'Release_i': 'None', 'Release_j': 'None'},
    {'Member': 'M10', 'i': 2, 'j': 6, 'BetaAngle': 90.0, 'Release_i': 'None', 'Release_j': 'None'},
    {'Member': 'M11', 'i': 3, 'j': 7, 'BetaAngle': 90.0, 'Release_i': 'None', 'Release_j': 'None'},
    {'Member': 'M12', 'i': 4, 'j': 8, 'BetaAngle': 90.0, 'Release_i': 'None', 'Release_j': 'None'}
]
df_members = pd.DataFrame(members)

# ==========================================
# 3. EXPORT TO STRUCTURED EXCEL FILE
# ==========================================
file_name = "cube_structure_Rev1.xlsx"
with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    df_nodes.to_excel(writer, sheet_name='Nodes & DOFs', index=False)
    df_members.to_excel(writer, sheet_name='Member Incidences & Releases', index=False)

    for sheet_name in ['Nodes & DOFs', 'Member Incidences & Releases']:
        ws = writer.sheets[sheet_name]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

print(f"Excel model successfully saved as '{file_name}'")

# ==========================================
# 4. ADVANCED 3D PLOTTING (MATCHING PROFESSOR'S LAYOUT)
# ==========================================
fig = plt.figure(figsize=(16, 9))

# 3D Structural Subplot (Left side)
ax = fig.add_subplot(121, projection='3d')

# Plot Nodes
free_nodes = df_nodes[df_nodes['Support'] == 'Free']
pinned_nodes = df_nodes[df_nodes['Support'] == 'Pinned']

ax.scatter(free_nodes['X'], free_nodes['Y'], free_nodes['Z'], color='salmon', s=50, label='Free node')
ax.scatter(pinned_nodes['X'], pinned_nodes['Y'], pinned_nodes['Z'], color='red', s=80, label='Supported node (pinned)')

for _, node in df_nodes.iterrows():
    ax.text(node['X'], node['Y'] + 0.3, node['Z'], f"N{node['NodeID']}\nDOF {node['DOF_Start']}-{node['DOF_End']}", fontsize=8, color='black', ha='center')
    # Draw pinned support pyramid bases at Y=0
    if node['Support'] == 'Pinned':
        ax.plot([node['X']-0.3, node['X']+0.3, node['X']], [node['Y']-0.5, node['Y']-0.5, node['Y']], [node['Z'], node['Z'], node['Z']], color='gray', linewidth=2)

# Plot Members and Releases
for _, member in df_members.iterrows():
    node_i = df_nodes[df_nodes['NodeID'] == member['i']].iloc[0]
    node_j = df_nodes[df_nodes['NodeID'] == member['j']].iloc[0]
    
    is_column = 'M9' in member['Member'] or 'M10' in member['Member'] or 'M11' in member['Member'] or 'M12' in member['Member']
    line_color = 'forestgreen' if is_column else 'navy'
    ax.plot([node_i['X'], node_j['X']], [node_i['Y'], node_j['Y']], [node_i['Z'], node_j['Z']], color=line_color, linewidth=2)
    
    mid_x, mid_y, mid_z = (node_i['X'] + node_j['X']) / 2, (node_i['Y'] + node_j['Y']) / 2, (node_i['Z'] + node_j['Z']) / 2
    label_text = f"{member['Member']}"
    if member['Release_i'] == 'Mz':
        label_text += " [MZ]"
    if member['BetaAngle'] != 0:
        label_text += f" (β={int(member['BetaAngle'])}°)"
    ax.text(mid_x, mid_y, mid_z, f" {label_text}", fontsize=8, color='darkblue', fontweight='bold')

    # Hollow circle symbols for pinned member ends (Mz released)
    if member['Release_i'] == 'Mz':
        ax.scatter([node_i['X']], [node_i['Y']], [node_i['Z']], color='white', edgecolors='black', s=50, zorder=5)
    if member['Release_j'] == 'Mz':
        ax.scatter([node_j['X']], [node_j['Y']], [node_j['Z']], color='white', edgecolors='black', s=50, zorder=5)

# Global Origin & Coordinate Arrows
ax.quiver(0, 0, 0, 1.5, 0, 0, color='darkorange', linewidth=2, arrow_length_ratio=0.2)
ax.quiver(0, 0, 0, 0, 1.5, 0, color='blue', linewidth=2, arrow_length_ratio=0.2)
ax.quiver(0, 0, 0, 0, 0, 1.5, color='saddlebrown', linewidth=2, arrow_length_ratio=0.2)
ax.text(1.7, 0, 0, 'X', color='darkorange', fontweight='bold')
ax.text(0, 1.7, 0, 'Y', color='blue', fontweight='bold')
ax.text(0, 0, 1.7, 'Z', color='saddlebrown', fontweight='bold')
ax.scatter([0], [0], [0], color='green', marker='*', s=100, label='Origin (0, 0, 0)')

ax.set_xlabel('X (m) - lateral', fontweight='bold')
ax.set_ylabel('Y (m) - vertical', fontweight='bold')
ax.set_zlabel('Z (m) - lateral', fontweight='bold')
ax.set_title('6m x 6m x 6m Cube - Structural Model, Rev. 1\nPinned supports at nodes 1-4, global/local axes, beta angles, and MZ end releases', fontsize=10, fontweight='bold')
ax.set_xlim(-1, 7)
ax.set_ylim(-1, 7)
ax.set_zlim(-1, 7)
ax.legend(loc='upper left', fontsize=8)

# ==========================================
# 5. MODEL DATA TEXT PANEL (Right Side Box)
# ==========================================
ax_text = fig.add_subplot(122)
ax_text.axis('off')

model_text = """MODEL DATA - REV. 1
---------------------------------------------
Geometry
  Cube edge          6.0 m
  Nodes              8
  Members            12

Global axes
  X                  lateral
  Y                  vertical (up)
  Z                  lateral

Supports
  Type               pinned
  Nodes              1, 2, 3, 4
  Restrained         UX, UY, UZ
  Released           RX, RY, RZ

Nodal degrees of freedom
  DOF per node       6
  Total DOF          48
  Restrained DOF     12
  Active DOF         36
  Numbering          (node - 1) x 6 + 1..6

Member end releases
  Pinned members     1, 3, 5, 7
  Pattern            Pinned i and j
  Component          MZ, moment about local z
  Released end DOF   8
  Member end DOF     144 (12 per member)
  Symbol             hollow circle at the end

Beta angles
  Base Beam          0 deg
  Roof Beam          0 deg
  Column             90 deg

Local axes
  local x            start node i to end node j
  local y            in the vertical plane, upward
  local z            completes the right-handed set
  Vertical members follow the limiting
  case: local z parallel to global Z.
"""

ax_text.text(0.05, 0.95, model_text, fontsize=9, family='monospace', verticalalignment='top', 
             bbox=dict(boxstyle='round', facecolor='aliceblue', edgecolor='steelblue', alpha=0.9))

plt.tight_layout()
plt.savefig('cube_structure_Rev1.png', dpi=300, bbox_inches='tight')
plt.show()